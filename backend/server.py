from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Request, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, date
import jwt
from jwt.exceptions import InvalidTokenError
from passlib.context import CryptContext
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import io
from enum import Enum
import asyncio
import time
import json
import hmac
import base64
import hashlib
import re

# Multi-tenant services
from services.multi_tenant import tenant_context, FeatureFlagService, tenant_middleware

# Super Admin module (integrated tenant management) - follows the 3-Feature Upgrade Guide
from super_admin import super_admin_router, set_database as set_super_admin_db, set_feature_service as set_super_admin_feature_service, create_indexes as create_super_admin_indexes

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection (Atlas TLS: use certifi CA bundle — avoids SSL handshake errors on macOS)
from mongo_utils import motor_client_kwargs

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url, **motor_client_kwargs(mongo_url))
DEFAULT_DB_NAME = os.environ['DB_NAME']
SUPER_ADMIN_DB_NAME = os.environ.get("SUPER_ADMIN_DB_NAME", "prawn_erp_super_admin")
ENABLE_MULTI_DB_ROUTING = os.environ.get("ENABLE_MULTI_DB_ROUTING", "false").lower() in ("1", "true", "yes", "on")


def _derive_tenant_db_name(tenant_id: str) -> str:
    safe = "".join(ch for ch in str(tenant_id or "") if ch.isalnum() or ch in ("_", "-")).strip()
    if not safe:
        return DEFAULT_DB_NAME
    if safe in ("default", "cli_001"):
        # Backward compatibility for existing primary tenant data.
        return DEFAULT_DB_NAME
    return f"prawn_erp_{safe}"


class TenantAwareDatabase:
    """
    Routes db.collection operations to tenant-specific Mongo databases.
    Existing handlers can continue using `db.<collection>` without code changes.
    """
    def __init__(self, mongo_client, default_db_name: str):
        self._client = mongo_client
        self._default_db_name = default_db_name

    def _active_db_name(self) -> str:
        if not ENABLE_MULTI_DB_ROUTING:
            return self._default_db_name
        try:
            tenant_id = tenant_context.get_tenant()
        except Exception:
            tenant_id = None
        if not tenant_id:
            return self._default_db_name
        return _derive_tenant_db_name(tenant_id)

    def _active_db(self):
        return self._client[self._active_db_name()]

    def __getattr__(self, item):
        return getattr(self._active_db(), item)

    def __getitem__(self, item):
        return self._active_db()[item]


db = TenantAwareDatabase(client, DEFAULT_DB_NAME)

# Feature Flag Service
feature_service = FeatureFlagService(db)

# Cache resolved tenant DB for cross-tenant login fallback (best-effort, in-memory)
_LOGIN_TENANT_CACHE_TTL_SEC = 300
_login_tenant_cache: Dict[str, Dict[str, Any]] = {}

# Security
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480


def _decode_jwt_with_known_keys(token: str) -> dict:
    """Decode JWT using supported keys for compatibility across services."""
    keys_to_try = [SECRET_KEY, os.environ.get("JWT_SECRET_KEY")]
    last_error = None
    for key in keys_to_try:
        if not key:
            continue
        try:
            return jwt.decode(token, key, algorithms=[ALGORITHM])
        except Exception as e:
            last_error = e
            continue
    if last_error:
        raise last_error
    raise HTTPException(status_code=401, detail="Invalid authentication credentials")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Prawn ERP - Multi-Tenant")
api_router = APIRouter(prefix="/api")

# Mount uploads directory for static file serving
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# CORS — `allow_origins=["*"]` with `allow_credentials=True` is invalid for browsers on
# credentialed requests (e.g. Authorization). Dev servers also hop ports (3000 vs 3001).
_cors_extra = [
    o.strip()
    for o in os.environ.get("CORS_ORIGINS", "").split(",")
    if o.strip()
]
_cors_regex = os.environ.get(
    "CORS_ALLOW_ORIGIN_REGEX",
    r"^https?://(localhost|127\.0\.0\.1|192\.168\.\d{1,3}\.\d{1,3})(:\d+)?$",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_extra,
    allow_origin_regex=_cors_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# Multi-Tenant Middleware
app.middleware("http")(tenant_middleware)


def _safe_parse_datetime(value):
    """
    Best-effort datetime parser for documents that may contain ISO strings, None,
    or legacy/invalid values. Returns the original value if parsing fails.
    This prevents 500s when listing resources like agents or dashboard lots.
    """
    if not value or not isinstance(value, str):
        return value
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return value

# Enums
class UserRole(str, Enum):
    super_admin = "super_admin"
    admin = "admin"
    owner = "owner"
    procurement_manager = "procurement_manager"
    production_supervisor = "production_supervisor"
    cold_storage_incharge = "cold_storage_incharge"
    qc_officer = "qc_officer"
    sales_manager = "sales_manager"
    accounts_manager = "accounts_manager"
    worker = "worker"

class Species(str, Enum):
    vannamei = "Vannamei"
    black_tiger = "Black Tiger"
    sea_tiger = "Sea Tiger"

class FreshnessGrade(str, Enum):
    A = "A"
    B = "B"
    C = "C"
    rejected = "Rejected"

class PaymentStatus(str, Enum):
    pending = "pending"
    partial = "partial"
    paid = "paid"
    overdue = "overdue"

class ProcessType(str, Enum):
    heading = "heading"
    peeling = "peeling"
    deveining = "deveining"
    iqf = "IQF"
    blanching = "blanching"
    grading = "grading"

class ProductForm(str, Enum):
    hoso = "HOSO"
    hlso = "HLSO"
    pto = "PTO"
    pd = "PD"
    pdto = "PDTO"
    butterfly = "Butterfly"
    ring_cut = "Ring Cut"
    cooked = "Cooked"

class QCStatus(str, Enum):
    pending = "pending"
    passed = "passed"
    failed = "failed"
    on_hold = "on_hold"

class StorageStatus(str, Enum):
    occupied = "occupied"
    empty = "empty"
    reserved = "reserved"
    maintenance = "maintenance"

class ShipmentStatus(str, Enum):
    draft = "draft"
    confirmed = "confirmed"
    in_transit = "in_transit"
    delivered = "delivered"

class ApprovalStatus(str, Enum):
    pending = "pending"
    approved = "approved"
    rejected = "rejected"

class PriceCategory(str, Enum):
    vannamei_30_40 = "Vannamei 30/40"
    vannamei_40_60 = "Vannamei 40/60"
    vannamei_60_80 = "Vannamei 60/80"
    black_tiger_20_30 = "Black Tiger 20/30"
    black_tiger_30_40 = "Black Tiger 30/40"

    cancelled = "cancelled"


# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: UserRole
    phone: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    tenant_id: Optional[str] = None
    # Impersonation fields
    is_impersonated: bool = False
    impersonator: Optional[str] = None
    impersonator_name: Optional[str] = None
    session_id: Optional[str] = None

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: UserRole
    phone: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User
    features: Optional[Dict[str, bool]] = None
    tenant_id: Optional[str] = None
    lot_number_prefix: Optional[str] = None

class Agent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agent_code: str
    name: str
    phone: str
    gst: Optional[str] = None
    pan: Optional[str] = None
    commission_pct: float = 0.0
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AgentCreate(BaseModel):
    agent_code: str
    name: str
    phone: str
    gst: Optional[str] = None
    pan: Optional[str] = None
    commission_pct: float = 0.0
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc: Optional[str] = None

class ProcurementPayment(BaseModel):
    amount: float
    payment_mode: str
    payment_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    reference: Optional[str] = None

class ProcurementLot(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lot_number: str
    agent_id: str
    agent_name: str
    vehicle_number: str
    driver_name: str
    arrival_time: datetime
    species: Species
    count_per_kg: str
    boxes_count: int
    gross_weight_kg: float
    ice_weight_kg: float
    net_weight_kg: float
    no_of_tons: float
    no_of_trays: int = 0
    rate_per_kg: float
    total_amount: float
    advance_paid: float = 0.0
    balance_due: float
    ice_ratio_pct: Optional[float] = None
    freshness_grade: FreshnessGrade
    is_rejected: bool = False
    rejection_reason: Optional[str] = None
    payment_status: PaymentStatus = PaymentStatus.pending
    payments: List[ProcurementPayment] = []
    photos: List[str] = []
    is_update_pending_approval: bool = False
    approval_status: ApprovalStatus = ApprovalStatus.approved
    approval_notes: Optional[str] = None
    approved_by: Optional[str] = None
    notes: Optional[str] = None
    attachments: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None
    
    @field_validator('species', mode='before')
    @classmethod
    def normalize_species(cls, v):
        """Normalize species to match enum values (handle case variations)"""
        if v is None:
            return v
        if isinstance(v, Species):
            return v
        # Convert string to title case for matching
        v_str = str(v).strip()
        # Try to match with enum values
        for species in Species:
            if species.value.lower() == v_str.lower():
                return species
        return v  # Let Pydantic handle the validation error

class ProcurementLotCreate(BaseModel):
    agent_id: str
    vehicle_number: str
    driver_name: str
    arrival_time: datetime
    species: Species
    count_per_kg: str
    boxes_count: int
    no_of_trays: int = 0
    gross_weight_kg: float
    ice_weight_kg: float
    rate_per_kg: float
    advance_paid: float = 0.0
    ice_ratio_pct: Optional[float] = None
    freshness_grade: FreshnessGrade
    is_rejected: bool = False
    rejection_reason: Optional[str] = None
    notes: Optional[str] = None

class Worker(BaseModel):
    worker_code: str
    name: str
    kg_processed: float
    hours_worked: float

class PreprocessingBatch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    batch_number: str
    procurement_lot_id: str
    process_type: ProcessType
    input_weight_kg: float
    output_weight_kg: float
    waste_weight_kg: float
    yield_pct: float
    yield_alert: bool = False
    start_time: datetime
    end_time: Optional[datetime] = None
    duration_mins: Optional[float] = None
    workers: List[Worker] = []
    supervisor: str
    notes: Optional[str] = None
    attachments: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PreprocessingBatchCreate(BaseModel):
    procurement_lot_id: str
    process_type: ProcessType
    input_weight_kg: float
    output_weight_kg: float
    no_of_trays: int = 0
    start_time: datetime
    end_time: Optional[datetime] = None
    workers: List[Worker] = []
    supervisor: str
    notes: Optional[str] = None

class ProductionOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_number: str
    preprocessing_batch_ids: List[str]
    product_form: ProductForm
    target_size_count: str
    glazing_pct: Optional[float] = None
    block_weight_kg: Optional[float] = None
    no_of_blocks: int
    input_weight_kg: float
    output_weight_kg: float
    conversion_rate_pct: float
    qc_status: QCStatus = QCStatus.pending
    notes: Optional[str] = None
    attachments: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductionOrderCreate(BaseModel):
    preprocessing_batch_ids: List[str]
    product_form: ProductForm
    target_size_count: str
    glazing_pct: Optional[float] = None
    block_weight_kg: Optional[float] = None
    no_of_blocks: int
    input_weight_kg: float
    output_weight_kg: float
    notes: Optional[str] = None

class FinishedGood(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fg_code: str
    production_order_id: str
    product_form: ProductForm
    size_count: str
    weight_kg: float
    qc_status: QCStatus = QCStatus.pending
    storage_location: Optional[str] = None
    temperature_c: Optional[float] = None
    manufactured_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    expiry_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FinishedGoodCreate(BaseModel):
    production_order_id: str
    product_form: ProductForm
    size_count: str
    weight_kg: float
    storage_location: Optional[str] = None
    temperature_c: Optional[float] = None
    expiry_date: Optional[datetime] = None

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    message: str
    module: str
    target_roles: List[UserRole]
    is_read: bool = False
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    title: str
    message: str
    module: str
    target_roles: List[UserRole]

class DashboardStats(BaseModel):
    total_procurement_lots: int
    total_weight_procured_kg: float
    total_procurement_value: float
    active_preprocessing_batches: int
    active_production_orders: int
    finished_goods_inventory_kg: float
    pending_qc_items: int
    recent_activities: List[Dict[str, Any]]


# QC Module Models
class QCInspection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    inspection_code: str
    entity_type: str  # procurement_lot, finished_good, cold_storage_slot
    entity_id: str
    inspection_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    qc_officer: str
    parameters: Dict[str, Any] = {}  # Flexible JSONB-like field
    overall_grade: FreshnessGrade
    pass_fail: bool
    failure_reason: Optional[str] = None
    lab_report_ref: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QCInspectionCreate(BaseModel):
    entity_type: str
    entity_id: str
    qc_officer: str
    parameters: Dict[str, Any] = {}
    overall_grade: FreshnessGrade
    pass_fail: bool
    failure_reason: Optional[str] = None
    lab_report_ref: Optional[str] = None
    notes: Optional[str] = None

# Cold Storage Models
class ColdStorageChamber(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    chamber_code: str  # CH-01, CH-02
    chamber_name: str
    setpoint_temperature_c: float = -18.0
    capacity_kg: float
    status: str = "active"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ColdStorageChamberCreate(BaseModel):
    chamber_code: str
    chamber_name: str
    setpoint_temperature_c: float = -18.0
    capacity_kg: float

class ColdStorageSlot(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    slot_code: str  # CH-01-R01-S05
    chamber_id: str
    rack_number: int
    slot_number: int
    status: StorageStatus = StorageStatus.empty
    occupied_weight_kg: float = 0.0
    fg_id: Optional[str] = None  # Linked finished good
    intake_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ColdStorageSlotCreate(BaseModel):
    slot_code: str
    chamber_id: str
    rack_number: int
    slot_number: int

class ColdStorageInventory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    slot_id: str
    fg_id: str
    quantity_kg: float
    carton_count: int
    intake_date: datetime
    days_in_storage: int = 0
    fifo_alert: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ColdStorageInventoryCreate(BaseModel):
    slot_id: str
    fg_id: str
    quantity_kg: float
    carton_count: int

class TemperatureLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    chamber_id: str
    temperature_c: float
    alert: bool = False
    alert_reason: Optional[str] = None
    recorded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TemperatureLogCreate(BaseModel):
    chamber_id: str
    temperature_c: float

# Sales & Dispatch Models
class Buyer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    buyer_code: str
    company_name: str
    country: str
    ie_code: Optional[str] = None
    gst: Optional[str] = None
    contact_person: str
    phone: str
    email: EmailStr
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BuyerCreate(BaseModel):
    buyer_code: str
    company_name: str
    country: str
    ie_code: Optional[str] = None
    gst: Optional[str] = None
    contact_person: str
    phone: str
    email: EmailStr

class SalesOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_number: str
    buyer_id: str
    buyer_name: str
    quantity_kg: float
    rate_per_kg_usd: float
    currency: str = "USD"
    total_value_usd: float
    delivery_date: datetime
    payment_status: PaymentStatus = PaymentStatus.pending
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SalesOrderCreate(BaseModel):
    buyer_id: str
    quantity_kg: float
    rate_per_kg_usd: float
    currency: str = "USD"
    delivery_date: datetime
    notes: Optional[str] = None

class Shipment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shipment_number: str
    sales_order_id: str
    container_no: str
    seal_no: str
    shipping_line: str
    vessel_name: str
    port_of_loading: str
    port_of_discharge: str
    destination_country: str
    etd: datetime  # Estimated Time of Departure
    eta: datetime  # Estimated Time of Arrival
    bill_of_lading: Optional[str] = None
    shipment_status: ShipmentStatus = ShipmentStatus.draft
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ShipmentCreate(BaseModel):
    sales_order_id: str
    container_no: str
    seal_no: str
    shipping_line: str
    vessel_name: str
    port_of_loading: str
    port_of_discharge: str
    destination_country: str
    etd: datetime
    eta: datetime
    bill_of_lading: Optional[str] = None

# Wage & Billing Models
class WageBill(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_number: str
    bill_type: str  # VA, TDS, contractor, daily
    period_from: datetime
    period_to: datetime
    department: str
    gross_amount: float
    tds_deduction: float
    net_payable: float
    payment_status: PaymentStatus = PaymentStatus.pending
    payment_date: Optional[datetime] = None
    line_items: List[Dict[str, Any]] = []
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WageBillCreate(BaseModel):
    bill_type: str
    period_from: datetime
    period_to: datetime
    department: str
    gross_amount: float
    tds_deduction: float
    line_items: List[Dict[str, Any]] = []
    notes: Optional[str] = None

class WageBillLine(BaseModel):
    worker_code: str
    worker_name: str
    days_worked: float
    rate_per_day: float
    basic_amount: float
    va_allowance: float
    tds_deducted: float
    net_amount: float

class Attachment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str
    entity_id: str
    file_name: str
    file_url: str
    file_size_kb: float
    mime_type: str
    category: str  # invoice, weighment_slip, lab_report, gate_pass, photo, other
    description: Optional[str] = None
    uploaded_by: str
    is_deleted: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AttachmentCreate(BaseModel):
    entity_type: str
    entity_id: str
    file_name: str
    category: str
    description: Optional[str] = None

class Note(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str
    entity_id: str
    note_text: str
    is_pinned: bool = False
    is_admin_note: bool = False
    authored_by: str
    author_name: str
    is_deleted: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NoteCreate(BaseModel):
    entity_type: str
    entity_id: str
    note_text: str
    is_pinned: bool = False

# ══════════════════════════════════════════════════════════════════════════════
# Purchase Invoice Models (Amendment A4)
# ══════════════════════════════════════════════════════════════════════════════

class PurchaseInvoiceStatus(str, Enum):
    draft = "draft"
    approved = "approved"
    pushed = "pushed"

class PurchaseInvoiceLine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_id: str
    line_no: int
    variety: str
    count_value: str
    custom_variety_notes: Optional[str] = None
    custom_count_notes: Optional[str] = None
    quantity_kg: float
    rate: float
    amount: float = 0.0  # Computed: quantity_kg × rate

class PurchaseInvoiceLineCreate(BaseModel):
    line_no: int
    variety: str
    count_value: str
    custom_variety_notes: Optional[str] = None
    custom_count_notes: Optional[str] = None
    quantity_kg: float
    rate: float

class PurchaseInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_no: str
    invoice_date: date = Field(default_factory=lambda: date.today())
    lot_id: Optional[str] = None
    
    # Farmer/Supplier info
    farmer_name: str
    farmer_mobile: Optional[str] = None  # A4 PATCH 10A
    farmer_location: Optional[str] = None
    agent_ref_name: Optional[str] = None
    weighment_slip_no: Optional[str] = None
    weighment_slip_file_url: Optional[str] = None  # e.g. /uploads/weighment_slip_*.jpg (compressed JPEG)
    weighment_slip_mime_type: Optional[str] = None
    
    # Custom fields
    custom_field_1_label: Optional[str] = None
    custom_field_1_value: Optional[str] = None
    custom_field_2_label: Optional[str] = None
    custom_field_2_value: Optional[str] = None
    
    # Computed totals
    total_quantity_kg: float = 0.0
    subtotal: float = 0.0
    tds_rate_pct: float = 0.1  # 0.1% default
    tds_amount: float = 0.0
    rounded_off: float = 0.0
    grand_total: float = 0.0
    
    # Payment tracking
    advance_paid: float = 0.0
    balance_due: float = 0.0
    payment_status: PaymentStatus = PaymentStatus.pending
    
    # Workflow
    status: PurchaseInvoiceStatus = PurchaseInvoiceStatus.draft
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    pushed_at: Optional[datetime] = None
    pushed_by: Optional[str] = None
    
    # Manual audit tracking (A4 PATCH 10A)
    is_manually_recorded: bool = False
    manually_recorded_at: Optional[datetime] = None
    manually_recorded_by: Optional[str] = None
    
    # A5: Party fields
    party_id: Optional[str] = None  # FK to parties collection
    party_name_text: Optional[str] = None  # Denormalized snapshot of party name
    
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Line items (populated separately)
    line_items: List[PurchaseInvoiceLine] = []

class PurchaseInvoiceCreate(BaseModel):
    invoice_date: date = Field(default_factory=lambda: date.today())
    farmer_name: str
    farmer_mobile: Optional[str] = None  # A4 PATCH 10F
    farmer_location: Optional[str] = None
    agent_ref_name: Optional[str] = None
    weighment_slip_no: Optional[str] = None
    weighment_slip_file_url: Optional[str] = None
    weighment_slip_mime_type: Optional[str] = None
    custom_field_1_label: Optional[str] = None
    custom_field_1_value: Optional[str] = None
    custom_field_2_label: Optional[str] = None
    custom_field_2_value: Optional[str] = None
    tds_rate_pct: float = 0.1
    advance_paid: float = 0.0
    notes: Optional[str] = None
    line_items: List[PurchaseInvoiceLineCreate] = []
    # A5: Party fields
    party_id: Optional[str] = None
    party_name_text: Optional[str] = None
    same_as_farmer: bool = False


class PushInvoiceRequest(BaseModel):
    apply_digital_signature: bool = False


# ══════════════════════════════════════════════════════════════════════════════
# Amendment A5: Party Ledger Models
# ══════════════════════════════════════════════════════════════════════════════

class LedgerEntryType(str, Enum):
    bill = "bill"
    payment = "payment"
    opening = "opening"
    manual_debit = "manual_debit"
    manual_credit = "manual_credit"

class PaymentMode(str, Enum):
    cash = "cash"
    bank_transfer = "bank_transfer"
    cheque = "cheque"
    upi = "upi"

class Party(BaseModel):
    """Party Master - stores supplier/farmer party information"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    party_name: str  # e.g. "SAI RAM AQUA TRADERS"
    party_alias: Optional[str] = None  # e.g. "RAMA RAO GARU" - shown in brackets
    short_code: Optional[str] = None  # e.g. "SRAT" - used in PAID TO column
    mobile: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    pan_number: Optional[str] = None
    is_active: bool = True
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # Virtual field - computed from ledger, not stored
    current_fy_balance: Optional[float] = None

class PartyCreate(BaseModel):
    party_name: str
    party_alias: Optional[str] = None
    short_code: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    pan_number: Optional[str] = None
    notes: Optional[str] = None

class PartyLedgerAccount(BaseModel):
    """Party Ledger Account - one per party per financial year"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    party_id: str
    financial_year: str  # Format: "25-26" (FY starts 01-Apr, ends 31-Mar)
    
    opening_balance: float = 0.0  # Carry-forward from previous FY
    closing_balance: float = 0.0  # Computed: opening + bills - payments
    
    total_billed: float = 0.0   # SUM(invoice subtotals) this FY
    total_tds: float = 0.0      # SUM(tds_amounts) this FY (4dp precision)
    total_payments: float = 0.0 # SUM(payments) this FY
    
    is_locked: bool = False  # TRUE after FY ends
    locked_at: Optional[datetime] = None
    locked_by: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PartyLedgerEntry(BaseModel):
    """Ledger Entry - append-only chronological log of transactions"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ledger_id: str
    party_id: str
    entry_date: date
    entry_type: LedgerEntryType
    
    # Bill entry fields (entry_type = 'bill' | 'manual_debit')
    invoice_id: Optional[str] = None
    invoice_no: Optional[str] = None
    bill_subtotal: Optional[float] = None
    tds_rate_pct: Optional[float] = None
    tds_amount: Optional[float] = None  # Stored to 4dp
    tds_after_bill: Optional[float] = None
    
    # Payment entry fields (entry_type = 'payment' | 'manual_credit')
    payment_amount: Optional[float] = None
    payment_date: Optional[date] = None
    paid_to: Optional[str] = None
    payment_mode: Optional[PaymentMode] = None
    payment_reference: Optional[str] = None
    
    # Manual entry fields
    description: Optional[str] = None
    
    # Running balance AFTER this entry (denormalized for fast rendering)
    balance_after: float = 0.0
    
    # Sequential order within ledger
    entry_order: int = 1
    
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentCreate(BaseModel):
    """Create a payment entry"""
    party_id: str
    entry_date: date
    payment_amount: float
    payment_date: Optional[date] = None
    paid_to: Optional[str] = None
    payment_mode: Optional[PaymentMode] = None
    payment_reference: Optional[str] = None
    invoice_id: Optional[str] = None  # Optional: link to specific invoice
    notes: Optional[str] = None

class ManualEntryCreate(BaseModel):
    """Create a manual debit/credit entry"""
    party_id: str
    entry_date: date
    entry_type: str  # 'manual_debit' or 'manual_credit'
    amount: float
    description: str


# ══════════════════════════════════════════════════════════════════════════════
# A5 Helper Functions
# ══════════════════════════════════════════════════════════════════════════════

def get_financial_year(d: date) -> str:
    """
    Calculate financial year from date.
    FY starts 01-Apr, ends 31-Mar.
    e.g. 06-Apr-2025 → "25-26" | 01-Mar-2026 → "25-26" | 01-Apr-2026 → "26-27"
    """
    year = d.year
    month = d.month
    if month >= 4:  # Apr-Dec: current year to next year
        start_year = year % 100
        end_year = (year + 1) % 100
    else:  # Jan-Mar: previous year to current year
        start_year = (year - 1) % 100
        end_year = year % 100
    return f"{start_year:02d}-{end_year:02d}"

def get_fy_date_range(fy: str) -> tuple:
    """
    Get start and end dates for a financial year.
    e.g. "25-26" → (date(2025, 4, 1), date(2026, 3, 31))
    """
    parts = fy.split("-")
    start_year = 2000 + int(parts[0])
    end_year = 2000 + int(parts[1])
    return (date(start_year, 4, 1), date(end_year, 3, 31))


def get_previous_fy(fy: str) -> str:
    parts = fy.split("-")
    if len(parts) != 2:
        raise ValueError(f"Invalid FY format: {fy}")
    start_year = int(parts[0])
    prev_start = (start_year - 1) % 100
    prev_end = start_year % 100
    return f"{prev_start:02d}-{prev_end:02d}"


def get_next_fy(fy: str) -> str:
    parts = fy.split("-")
    if len(parts) != 2:
        raise ValueError(f"Invalid FY format: {fy}")
    start_year = int(parts[0])
    next_start = (start_year + 1) % 100
    next_end = (next_start + 1) % 100
    return f"{next_start:02d}-{next_end:02d}"


async def _get_carry_forward_opening_balance(party_id: str, fy: str) -> float:
    """Carry opening from previous FY closing balance when available."""
    try:
        prev_fy = get_previous_fy(fy)
    except Exception:
        return 0.0
    prev_ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": prev_fy},
        {"_id": 0, "closing_balance": 1},
    )
    return float((prev_ledger or {}).get("closing_balance") or 0.0)


async def _create_party_ledger_for_fy(
    party_id: str,
    fy: str,
    created_by: Optional[str] = None,
    opening_balance: Optional[float] = None,
    opening_entry_date: Optional[date] = None,
) -> dict:
    if opening_balance is None:
        opening_balance = await _get_carry_forward_opening_balance(party_id, fy)
    opening_balance = float(opening_balance or 0.0)
    now_utc = datetime.now(timezone.utc)
    ledger = {
        "id": str(uuid.uuid4()),
        "party_id": party_id,
        "financial_year": fy,
        "opening_balance": opening_balance,
        "closing_balance": opening_balance,
        "total_billed": 0.0,
        "total_tds": 0.0,
        "total_payments": 0.0,
        "is_locked": False,
        "created_at": now_utc,
        "updated_at": now_utc,
    }
    await db.party_ledger_accounts.insert_one(ledger)

    dt = opening_entry_date or date.today()
    opening_entry = {
        "id": str(uuid.uuid4()),
        "ledger_id": ledger["id"],
        "party_id": party_id,
        "entry_date": dt.isoformat() if hasattr(dt, "isoformat") else str(dt),
        "entry_type": "opening",
        "entry_order": 0,
        "description": f"Opening Balance FY {fy}",
        "balance_after": opening_balance,
        "created_at": now_utc,
        "created_by": created_by or "system",
    }
    await db.party_ledger_entries.insert_one(opening_entry)
    return ledger


# Helper functions
def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    # Integer exp is the most compatible across PyJWT versions and decoders
    to_encode.update({"exp": int(expire.timestamp())})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


# Approval Workflow Models
class PendingApproval(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str  # procurement_lot, production_order, etc
    entity_id: str
    entity_display: str  # Lot number, order number, etc
    change_type: str  # update, delete, status_change
    old_data: Dict[str, Any] = {}
    new_data: Dict[str, Any] = {}
    requested_by: str
    requester_name: str
    approval_status: ApprovalStatus = ApprovalStatus.pending
    approved_by: Optional[str] = None
    approval_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ApprovalAction(BaseModel):
    approval_id: str
    action: str  # approve, reject
    notes: Optional[str] = None

# Live Price Tracking Models
class LivePriceData(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str  # Vannamei 30/40, etc
    price_per_kg: float
    location: str = "Andhra Pradesh"
    market: str  # Nellore, Kakinada, etc
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    source: str = "Market Data"

class PhotoTracker(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str
    entity_id: str
    entity_display: str
    stage: str  # procurement, preprocessing, production, qc, cold_storage
    photo_url: str
    count_per_kg_visible: Optional[str] = None
    tray_count_visible: Optional[int] = None
    uploaded_by: str
    uploader_name: str
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PhotoUpload(BaseModel):
    entity_type: str
    entity_id: str
    entity_display: str
    stage: str
    photo_url: str
    count_per_kg_visible: Optional[str] = None
    tray_count_visible: Optional[int] = None
    notes: Optional[str] = None

# Traceability Models
class TraceabilityRecord(BaseModel):
    lot_number: str
    procurement_data: Optional[Dict[str, Any]] = None
    preprocessing_data: List[Dict[str, Any]] = []
    production_data: List[Dict[str, Any]] = []
    finished_goods_data: List[Dict[str, Any]] = []
    cold_storage_data: List[Dict[str, Any]] = []
    shipment_data: Optional[Dict[str, Any]] = None
    total_count_per_kg_changes: List[str] = []
    total_tray_count: int = 0
    photos_count: int = 0


# ============================================================================
# WASTAGE TRACKING MODELS (v4.0)
# ============================================================================

class YieldBenchmark(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    species: Species
    process_type: str  # 'gate_ice', 'heading', 'peeling', 'deveining', etc.
    min_yield_pct: Optional[float] = None
    optimal_yield_pct: Optional[float] = None
    max_yield_pct: Optional[float] = None
    tolerance_pct: Optional[float] = None  # For glazing/breading
    reference_rate_per_kg: Optional[float] = None
    description: Optional[str] = None
    is_active: bool = True
    set_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class YieldBenchmarkCreate(BaseModel):
    species: Species
    process_type: str
    min_yield_pct: Optional[float] = None
    optimal_yield_pct: Optional[float] = None
    max_yield_pct: Optional[float] = None
    tolerance_pct: Optional[float] = None
    reference_rate_per_kg: Optional[float] = None
    description: Optional[str] = None

class MarketRate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    species: Species
    product_form: Optional[str] = None  # NULL = raw/unprocessed
    size_value: Optional[str] = None  # NULL = all sizes
    rate_per_kg_inr: float
    rate_per_kg_usd: Optional[float] = None
    effective_from: date
    effective_to: Optional[date] = None  # NULL = currently active
    remarks: Optional[str] = None
    set_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MarketRateCreate(BaseModel):
    species: Species
    product_form: Optional[str] = None
    size_value: Optional[str] = None
    rate_per_kg_inr: float
    rate_per_kg_usd: Optional[float] = None
    effective_from: date
    effective_to: Optional[date] = None
    remarks: Optional[str] = None

class LotStageWastage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lot_id: str
    stage_sequence: int
    stage_name: str
    process_type: str
    source_entity_type: Optional[str] = None
    source_entity_id: Optional[str] = None
    input_weight_kg: float
    output_weight_kg: float
    wastage_kg: float = 0  # Calculated: input - output
    yield_pct: float = 0  # Calculated: (output / input) * 100
    min_yield_pct: Optional[float] = None
    optimal_yield_pct: Optional[float] = None
    threshold_status: str = "info"  # green | amber | red | info
    rate_per_kg_used: Optional[float] = None
    revenue_loss_inr: float = 0
    gradedown_kg: float = 0
    gradedown_rate_gap: float = 0
    gradedown_loss_inr: float = 0
    target_glaze_pct: Optional[float] = None
    actual_glaze_pct: Optional[float] = None
    glaze_variance_pct: Optional[float] = None
    glaze_revenue_gap_inr: float = 0
    byproduct_revenue_inr: float = 0
    net_loss_inr: float = 0
    is_alert: bool = False
    alert_acknowledged: bool = False
    alert_ack_by: Optional[str] = None
    alert_ack_at: Optional[datetime] = None
    recorded_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class LotStageWastageCreate(BaseModel):
    lot_id: str
    stage_sequence: int
    stage_name: str
    process_type: str
    source_entity_type: Optional[str] = None
    source_entity_id: Optional[str] = None
    input_weight_kg: float
    output_weight_kg: float

class WastageDashboardStats(BaseModel):
    today_wastage_kg: float = 0
    today_lots_count: int = 0
    month_revenue_loss_inr: float = 0
    month_lots_count: int = 0
    active_red_alerts: int = 0
    byproduct_revenue_inr: float = 0

class WastageBreachAlert(BaseModel):
    id: str
    lot_id: str
    lot_number: str
    stage_name: str
    species: str
    actual_yield_pct: float
    min_threshold_pct: float
    variance_pct: float
    loss_inr: float
    created_at: datetime


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    _t0 = time.perf_counter()
    _m: Dict[str, Any] = {
        "token_decode_ms": 0,
        "session_lookup_ms": 0,
        "user_lookup_ms": 0,
        "regex_fallback_used": False,
        "impersonation": False,
    }
    try:
        token = credentials.credentials
        _t_decode = time.perf_counter()
        payload = _decode_jwt_with_known_keys(token)
        _m["token_decode_ms"] = round((time.perf_counter() - _t_decode) * 1000, 1)
        email: str = payload.get("sub")
        token_type: str = payload.get("type", "regular")
        is_impersonation: bool = payload.get("is_impersonation", False)
        
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        # Handle impersonation tokens (check both formats)
        if token_type == "impersonation" or is_impersonation:
            _m["impersonation"] = True
            # Validate impersonation session
            session_id = payload.get("session_id")
            tenant_id = payload.get("tenant_id")
            impersonator = payload.get("impersonator")
            impersonator_name = payload.get("impersonator_name")
            
            # Check if session is still valid in MongoDB
            # Check in local DB first, then in super-admin DB collection
            _t_session = time.perf_counter()
            session = await db.impersonation_tokens.find_one({
                "session_id": session_id,
                "tenant_id": tenant_id
            })
            _m["session_lookup_ms"] = round((time.perf_counter() - _t_session) * 1000, 1)
            
            # If not found locally, the session might be valid from super-admin
            # We trust the JWT token if it's not expired
            if not session:
                # Check if the token was issued recently (within validity period)
                # Trust the JWT claims for impersonation from super-admin
                pass  # Allow impersonation to proceed
            
            # Find the user being impersonated
            _t_user = time.perf_counter()
            user_doc = await db.users.find_one({"email": email, "tenant_id": tenant_id}, {"_id": 0})
            _m["user_lookup_ms"] = round((time.perf_counter() - _t_user) * 1000, 1)
            
            if user_doc is None:
                # If user doesn't exist, create a temporary admin user object
                user_doc = {
                    "email": email,
                    "name": f"Admin ({payload.get('impersonator_name', 'Super Admin')})",
                    "role": "admin",
                    "tenant_id": tenant_id,
                    "is_impersonated": True,
                    "impersonator": impersonator,
                    "impersonator_name": impersonator_name,
                    "session_id": session_id,
                    "created_at": datetime.utcnow()
                }
            else:
                # Mark as impersonated
                user_doc["is_impersonated"] = True
                user_doc["impersonator"] = impersonator
                user_doc["impersonator_name"] = impersonator_name
                user_doc["session_id"] = session_id
            
            if isinstance(user_doc.get('created_at'), str):
                user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
            _m["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
            logger.debug("get_current_user metrics email=%s metrics=%s", email, _m)
            return User(**user_doc)
            
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    # Regular token flow
    email_lookup = (email or "").strip().lower()
    _t_user = time.perf_counter()
    user_doc = await db.users.find_one({"email": email_lookup}, {"_id": 0})
    if not user_doc:
        _m["regex_fallback_used"] = True
        user_doc = await db.users.find_one(
            {"email": {"$regex": f"^{re.escape((email or '').strip())}$", "$options": "i"}},
            {"_id": 0},
        )
    _m["user_lookup_ms"] = round((time.perf_counter() - _t_user) * 1000, 1)
    if user_doc is None:
        _m["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
        logger.debug("get_current_user metrics email=%s metrics=%s", email_lookup, _m)
        raise HTTPException(status_code=401, detail="User not found")
    if user_doc.get("is_active") is False:
        raise HTTPException(status_code=403, detail="Account is disabled")

    # Force logout if password/session version changed after token issue.
    token_pwd_ts = _normalize_pwd_timestamp(payload.get("pwd_ts"))
    current_pwd_ts_raw = user_doc.get("password_changed_at") or user_doc.get("updated_at") or user_doc.get("created_at")
    current_pwd_ts = _normalize_pwd_timestamp(current_pwd_ts_raw)
    if token_pwd_ts and current_pwd_ts and token_pwd_ts != current_pwd_ts:
        raise HTTPException(status_code=401, detail="Session expired. Please login again.")
    
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    # Same defensiveness as login() for legacy / migrated user documents
    if not user_doc.get('name') or str(user_doc.get('name')).strip() == "":
        user_doc['name'] = (user_doc.get('email') or email_lookup or "user").split("@")[0] or "User"
    role_raw = user_doc.get('role', UserRole.worker.value)
    if isinstance(role_raw, str):
        try:
            user_doc['role'] = UserRole(role_raw)
        except ValueError:
            user_doc['role'] = UserRole.worker
    try:
        _m["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
        logger.debug("get_current_user metrics email=%s metrics=%s", email_lookup, _m)
        return User(**{k: v for k, v in user_doc.items() if k not in ('password', 'password_hash')})
    except Exception:
        _m["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
        logger.debug("get_current_user metrics email=%s metrics=%s", email_lookup, _m)
        raise HTTPException(status_code=401, detail="Invalid user profile")

def generate_lot_number() -> str:
    """Generate unique lot number with configurable prefix per tenant"""
    prefix = tenant_context.lot_number_prefix
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    counter = now.strftime("%H%M%S")
    return f"{prefix}-{date_str}-{counter}"

def generate_batch_number() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"BATCH-{date_str}-{counter}"


# ============================================================================
# WASTAGE TRACKING HELPER FUNCTIONS
# ============================================================================

async def create_wastage_record(
    lot_id: str,
    stage_sequence: int,
    stage_name: str,
    process_type: str,
    input_weight_kg: float,
    output_weight_kg: float,
    source_entity_type: str,
    source_entity_id: str,
    recorded_by: str
) -> dict:
    """
    Create a wastage record and return it
    Auto-calculates yield, checks thresholds, creates alerts if needed
    """
    # Calculate derived fields
    wastage_kg = input_weight_kg - output_weight_kg
    yield_pct = (output_weight_kg / input_weight_kg * 100) if input_weight_kg > 0 else 0
    
    # Look up lot and benchmark
    lot = await db.procurement_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        return None
    
    benchmark = await db.yield_benchmarks.find_one({
        "species": lot['species'],
        "process_type": process_type,
        "is_active": True
    }, {"_id": 0})
    
    # Calculate threshold status
    threshold_status = "info"
    min_yield_pct = None
    optimal_yield_pct = None
    
    if benchmark:
        min_yield_pct = benchmark.get('min_yield_pct')
        optimal_yield_pct = benchmark.get('optimal_yield_pct')
        
        if min_yield_pct and optimal_yield_pct:
            if yield_pct >= optimal_yield_pct:
                threshold_status = "green"
            elif yield_pct >= min_yield_pct:
                threshold_status = "amber"
            else:
                threshold_status = "red"
    
    # Look up market rate
    rate_per_kg = None
    if benchmark and benchmark.get('reference_rate_per_kg'):
        rate_per_kg = benchmark['reference_rate_per_kg']
    else:
        rate_per_kg = lot.get('rate_per_kg', 0)
    
    # Calculate revenue loss
    revenue_loss_inr = wastage_kg * rate_per_kg if rate_per_kg else 0
    
    # Create wastage record
    wastage = {
        "id": str(uuid.uuid4()),
        "lot_id": lot_id,
        "stage_sequence": stage_sequence,
        "stage_name": stage_name,
        "process_type": process_type,
        "source_entity_type": source_entity_type,
        "source_entity_id": source_entity_id,
        "input_weight_kg": input_weight_kg,
        "output_weight_kg": output_weight_kg,
        "wastage_kg": wastage_kg,
        "yield_pct": round(yield_pct, 2),
        "min_yield_pct": min_yield_pct,
        "optimal_yield_pct": optimal_yield_pct,
        "threshold_status": threshold_status,
        "rate_per_kg_used": rate_per_kg,
        "revenue_loss_inr": revenue_loss_inr,
        "gradedown_kg": 0,
        "gradedown_rate_gap": 0,
        "gradedown_loss_inr": 0,
        "target_glaze_pct": None,
        "actual_glaze_pct": None,
        "glaze_variance_pct": None,
        "glaze_revenue_gap_inr": 0,
        "byproduct_revenue_inr": 0,
        "net_loss_inr": revenue_loss_inr,
        "is_alert": (threshold_status == "red"),
        "alert_acknowledged": False,
        "alert_ack_by": None,
        "alert_ack_at": None,
        "recorded_by": recorded_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.lot_stage_wastage.insert_one(wastage)
    
    # Create alert notification if red
    if wastage["is_alert"]:
        notification = {
            "id": str(uuid.uuid4()),
            "title": f"🔴 Yield Alert: {stage_name} below minimum threshold",
            "message": f"Lot {lot['lot_number']} {stage_name}: {yield_pct:.2f}% yield (min: {min_yield_pct}%)\nWastage: {wastage_kg:.2f} kg | Revenue loss: ₹{revenue_loss_inr:.2f}",
            "type": "alert",
            "priority": "urgent",
            "target_roles": ['admin', 'owner', 'production_supervisor'],
            "link": f"/lot/{lot_id}/wastage",
            "is_read": False,
            "created_by": recorded_by,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notification)
    
    return wastage


def generate_order_number() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"PO-{date_str}-{counter}"

def generate_fg_code() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"FG-{date_str}-{counter}"

async def create_audit_log(user_id: str, action: str, module: str, details: dict):
    log = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "action": action,
        "module": module,
        "details": details,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log)


def generate_inspection_code() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"QC-{date_str}-{counter}"

def generate_order_num() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"SO-{date_str}-{counter}"

def generate_shipment_number() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"SHIP-{date_str}-{counter}"

def generate_bill_number() -> str:
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    counter = now.strftime("%H%M%S")
    return f"BILL-{date_str}-{counter}"


def generate_procurement_receipt_pdf(lot: ProcurementLot, agent: Agent) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("PROCUREMENT RECEIPT", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    info_data = [
        ['Lot Number:', lot.lot_number, 'Date:', lot.arrival_time.strftime('%d-%m-%Y %H:%M')],
        ['Agent:', agent.name, 'Agent Code:', agent.agent_code],
        ['Vehicle:', lot.vehicle_number, 'Driver:', lot.driver_name],
        ['Species:', lot.species.value, 'Count/KG:', lot.count_per_kg],
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    weight_data = [
        ['Description', 'Value'],
        ['Boxes/Crates', str(lot.boxes_count)],
        ['Gross Weight (KG)', f"{lot.gross_weight_kg:.2f}"],
        ['Ice Weight (KG)', f"{lot.ice_weight_kg:.2f}"],
        ['Net Weight (KG)', f"{lot.net_weight_kg:.2f}"],
        ['Tons', f"{lot.no_of_tons:.3f}"],
        ['Freshness Grade', lot.freshness_grade.value],
    ]
    
    weight_table = Table(weight_data, colWidths=[3*inch, 3*inch])
    weight_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(weight_table)
    elements.append(Spacer(1, 0.3*inch))
    
    payment_data = [
        ['Payment Details', ''],
        ['Rate per KG', f"₹{lot.rate_per_kg:.2f}"],
        ['Total Amount', f"₹{lot.total_amount:.2f}"],
        ['Advance Paid', f"₹{lot.advance_paid:.2f}"],
        ['Balance Due', f"₹{lot.balance_due:.2f}"],
        ['Payment Status', lot.payment_status.value.upper()],
    ]
    
    payment_table = Table(payment_data, colWidths=[3*inch, 3*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(payment_table)
    
    if lot.notes:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph(f"<b>Notes:</b> {lot.notes}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_wage_bill_pdf(bill: WageBill) -> bytes:
    """Generate PDF for wage bill with worker line items"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("WAGE BILL", title_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Bill Info

# ══════════════════════════════════════════════════════════════════════════════
# Purchase Invoice Business Logic (Amendment A4)
# ══════════════════════════════════════════════════════════════════════════════

async def generate_invoice_number() -> str:
    """Generate invoice number in format: sequential_no/FY_short (e.g., 2376/25-26)"""
    # Calculate financial year
    now = datetime.now(timezone.utc)
    month = now.month
    year = now.year
    
    if month >= 4:  # April onwards = new FY
        fy_short = f"{str(year)[-2:]}-{str(year+1)[-2:]}"
    else:  # Jan-Mar = previous FY
        fy_short = f"{str(year-1)[-2:]}-{str(year)[-2:]}"
    
    # Get max sequential number
    existing = await db.purchase_invoices.find({}, {"_id": 0, "invoice_no": 1}).to_list(10000)
    max_seq = 0
    for inv in existing:
        if inv.get('invoice_no'):
            try:
                seq_part = inv['invoice_no'].split('/')[0]
                seq_num = int(seq_part)
                if seq_num > max_seq:
                    max_seq = seq_num
            except:
                pass
    
    next_seq = max_seq + 1
    return f"{next_seq}/{fy_short}"

def calculate_invoice_totals(line_items: List[PurchaseInvoiceLineCreate], tds_rate_pct: float) -> dict:
    """Calculate all invoice totals following the exact formula from reference images"""
    # Calculate line amounts and subtotal
    total_qty = 0.0
    subtotal = 0.0
    
    for line in line_items:
        line_amount = round(line.quantity_kg * line.rate, 2)
        subtotal += line_amount
        total_qty += line.quantity_kg
    
    subtotal = round(subtotal, 2)
    
    # TDS calculation: 0.1% = 0.001
    tds_amount = round(subtotal * (tds_rate_pct / 100), 2)
    
    # Pre-round amount
    pre_round = subtotal - tds_amount
    
    # Rounded off adjustment
    grand_total = round(pre_round)
    rounded_off = round(grand_total - pre_round, 2)
    
    return {
        "total_quantity_kg": round(total_qty, 3),
        "subtotal": subtotal,
        "tds_amount": tds_amount,
        "rounded_off": rounded_off,
        "grand_total": grand_total
    }


def _pdf_safe_float(value: Any, default: float = 0.0) -> float:
    """Coerce DB/JSON values to float for ReportLab formatting (None, '', strings with commas)."""
    if value is None:
        return default
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    try:
        s = str(value).strip().replace(",", "")
        if s == "":
            return default
        return float(s)
    except (ValueError, TypeError):
        return default


def _normalize_pwd_timestamp(value: Any) -> str:
    """Normalize password/session timestamp for stable JWT comparison."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return ""
        try:
            # Canonicalize equivalent variants like " " vs "T".
            return datetime.fromisoformat(cleaned).isoformat()
        except Exception:
            return cleaned
    return str(value)


def amount_to_words_indian(amount: Any) -> str:
    """Convert amount to Indian rupee words (Lakh/Crore system)"""
    amount = int(round(_pdf_safe_float(amount, 0.0)))
    
    if amount == 0:
        return "Zero Rupees Only"
    
    # Indian number system
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    
    def convert_two_digit(n):
        if n < 10:
            return ones[n]
        elif n < 20:
            return teens[n-10]
        else:
            return tens[n//10] + (" " + ones[n%10] if n%10 != 0 else "")
    
    def convert_three_digit(n):
        if n >= 100:
            result = ones[n//100] + " Hundred"
            if n % 100 != 0:
                result += " " + convert_two_digit(n % 100)
            return result
        else:
            return convert_two_digit(n)
    
    crore = amount // 10000000
    amount %= 10000000
    lakh = amount // 100000
    amount %= 100000
    thousand = amount // 1000
    amount %= 1000
    hundred = amount
    
    result = []
    
    if crore > 0:
        result.append(convert_two_digit(crore) + " Crore")
    if lakh > 0:
        result.append(convert_two_digit(lakh) + " Lakh")
    if thousand > 0:
        result.append(convert_two_digit(thousand) + " Thousand")
    if hundred > 0:
        result.append(convert_three_digit(hundred))
    
    return " ".join(result) + " Rupees Only"

def generate_wage_bill_pdf(bill: WageBill) -> bytes:
    """Generate PDF for wage bill with worker line items"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("WAGE BILL", title_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Bill Info
    info_data = [
        ['Bill Number:', bill.bill_number, 'Bill Type:', bill.bill_type.upper()],
        ['Department:', bill.department, 'Status:', bill.payment_status.value.upper()],
        ['Period From:', bill.period_from.strftime('%d-%m-%Y'), 'Period To:', bill.period_to.strftime('%d-%m-%Y')],
        ['Created:', bill.created_at.strftime('%d-%m-%Y %H:%M'), '', ''],
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Worker Line Items Table
    if bill.line_items and len(bill.line_items) > 0:
        elements.append(Paragraph("<b>Worker Details</b>", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        worker_data = [['Code', 'Name', 'Days', 'Rate', 'Basic', 'VA', 'TDS', 'Net']]
        
        for item in bill.line_items:
            worker_data.append([
                item.get('worker_code', ''),
                item.get('worker_name', ''),
                str(item.get('days_worked', 0)),
                f"₹{item.get('rate_per_day', 0):.0f}",
                f"₹{item.get('basic_amount', 0):.2f}",
                f"₹{item.get('va_allowance', 0):.2f}",
                f"₹{item.get('tds_deducted', 0):.2f}",
                f"₹{item.get('net_amount', 0):.2f}",
            ])
        
        # Calculate totals
        total_basic = sum(item.get('basic_amount', 0) for item in bill.line_items)
        total_va = sum(item.get('va_allowance', 0) for item in bill.line_items)
        total_tds = sum(item.get('tds_deducted', 0) for item in bill.line_items)
        total_net = sum(item.get('net_amount', 0) for item in bill.line_items)
        
        worker_data.append([
            '', 'TOTAL', '', '',
            f"₹{total_basic:.2f}",
            f"₹{total_va:.2f}",
            f"₹{total_tds:.2f}",
            f"₹{total_net:.2f}"
        ])
        
        worker_table = Table(worker_data, colWidths=[0.6*inch, 1.8*inch, 0.5*inch, 0.7*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1*inch])
        worker_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(worker_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Summary Table
    elements.append(Paragraph("<b>Bill Summary</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    summary_data = [
        ['Gross Amount', f"₹{bill.gross_amount:,.2f}"],
        ['TDS Deduction', f"₹{bill.tds_deduction:,.2f}"],
        ['Net Payable', f"₹{bill.net_payable:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#10b981')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    
    # Notes
    if bill.notes:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph(f"<b>Notes:</b> {bill.notes}", styles['Normal']))
    
    # Payment Status
    if bill.payment_date:
        elements.append(Spacer(1, 0.2*inch))
        payment_text = f"<b>Payment Date:</b> {bill.payment_date.strftime('%d-%m-%Y %H:%M')}"
        elements.append(Paragraph(payment_text, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# Auth endpoints
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    email_reg = user_data.email.strip().lower()
    existing = await db.users.find_one({"email": email_reg}, {"_id": 0})
    if not existing:
        existing = await db.users.find_one(
            {"email": {"$regex": f"^{re.escape(user_data.email.strip())}$", "$options": "i"}},
            {"_id": 0},
        )
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = get_password_hash(user_data.password)
    user = User(
        email=email_reg,
        name=user_data.name,
        role=user_data.role,
        phone=user_data.phone
    )
    
    user_dict = user.model_dump()
    user_dict['password'] = hashed_password
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    _t0 = time.perf_counter()
    _metrics: Dict[str, Any] = {
        "tenant_hint": None,
        "shared_lookup_ms": 0,
        "cache_lookup_ms": 0,
        "hinted_tenant_lookup_ms": 0,
        "full_scan_ms": 0,
        "full_scan_tenants_checked": 0,
        "password_verify_ms": 0,
        "feature_flags_ms": 0,
        "cross_tenant_cache_hit": False,
        "fallback_used": False,
    }
    email_key = credentials.email.strip().lower()
    try:
        hinted_tenant_id = tenant_context.get_tenant()
    except Exception:
        hinted_tenant_id = None
    _metrics["tenant_hint"] = hinted_tenant_id

    user_doc = None
    # Prefer tenant-scoped lookup first to avoid cross-tenant collisions
    # when the same email exists in multiple client tenants.
    _t_shared = time.perf_counter()
    if hinted_tenant_id:
        user_doc = await db.users.find_one({"email": email_key, "tenant_id": hinted_tenant_id}, {"_id": 0})
        if not user_doc:
            user_doc = await db.users.find_one(
                {
                    "tenant_id": hinted_tenant_id,
                    "email": {"$regex": f"^{re.escape(credentials.email.strip())}$", "$options": "i"},
                },
                {"_id": 0},
            )
    if not user_doc:
        user_doc = await db.users.find_one({"email": email_key}, {"_id": 0})
    if not user_doc:
        # Legacy rows may use different email casing
        user_doc = await db.users.find_one(
            {"email": {"$regex": f"^{re.escape(credentials.email.strip())}$", "$options": "i"}},
            {"_id": 0},
        )
    _metrics["shared_lookup_ms"] = round((time.perf_counter() - _t_shared) * 1000, 1)
    if not user_doc and ENABLE_MULTI_DB_ROUTING:
        # Fast path: use cached tenant resolution for this email if available.
        _t_cache = time.perf_counter()
        cached = _login_tenant_cache.get(email_key)
        if cached and cached.get("expires_at", 0) > time.time():
            try:
                tid = cached.get("tenant_id")
                db_name = cached.get("db_name")
                if db_name:
                    tdb = client[db_name]
                    user_doc = await tdb.users.find_one({"email": email_key}, {"_id": 0})
                    if not user_doc:
                        user_doc = await tdb.users.find_one(
                            {"email": {"$regex": f"^{re.escape(credentials.email.strip())}$", "$options": "i"}},
                            {"_id": 0},
                        )
                    if user_doc and tid:
                        _metrics["cross_tenant_cache_hit"] = True
                        if not user_doc.get("tenant_id"):
                            user_doc["tenant_id"] = tid
                        tenant_context.set_tenant(tid, "PRW")
            except Exception as e:
                logger.debug("Cached cross-tenant login lookup miss for %s: %s", email_key, e)
                _login_tenant_cache.pop(email_key, None)
        _metrics["cache_lookup_ms"] = round((time.perf_counter() - _t_cache) * 1000, 1)
    if not user_doc and ENABLE_MULTI_DB_ROUTING:
        # Fallback: search other tenant databases (for DB-per-tenant clients) using super-admin client mappings.
        _metrics["fallback_used"] = True
        try:
            email_regex = {"$regex": f"^{re.escape(credentials.email.strip())}$", "$options": "i"}
            clients_coll = client[SUPER_ADMIN_DB_NAME].clients

            # First try hinted tenant only (cheap and usually correct).
            hinted_checked = False
            if hinted_tenant_id:
                _t_hinted = time.perf_counter()
                hinted_checked = True
                hinted_doc = await clients_coll.find_one(
                    {"tenant_id": hinted_tenant_id},
                    {"_id": 0, "tenant_id": 1, "client_db_name": 1},
                )
                if hinted_doc:
                    tid = hinted_doc.get("tenant_id")
                    db_name = hinted_doc.get("client_db_name") or _derive_tenant_db_name(tid)
                    if db_name and db_name != DEFAULT_DB_NAME:
                        tdb = client[db_name]
                        cand = await tdb.users.find_one({"email": email_key}, {"_id": 0})
                        if not cand:
                            cand = await tdb.users.find_one({"email": email_regex}, {"_id": 0})
                        if cand:
                            if not cand.get("tenant_id"):
                                cand["tenant_id"] = tid
                            user_doc = cand
                            if tid:
                                tenant_context.set_tenant(tid, "PRW")
                                _login_tenant_cache[email_key] = {
                                    "tenant_id": tid,
                                    "db_name": db_name,
                                    "expires_at": time.time() + _LOGIN_TENANT_CACHE_TTL_SEC,
                                }
                _metrics["hinted_tenant_lookup_ms"] = round((time.perf_counter() - _t_hinted) * 1000, 1)

            if not user_doc:
                _t_scan = time.perf_counter()
                clients_query = {} if not hinted_checked else {"tenant_id": {"$ne": hinted_tenant_id}}
                clients_rows = await clients_coll.find(
                    clients_query,
                    {"_id": 0, "tenant_id": 1, "client_db_name": 1},
                ).to_list(2000)

                for cdoc in clients_rows:
                    _metrics["full_scan_tenants_checked"] += 1
                    tid = cdoc.get("tenant_id")
                    db_name = cdoc.get("client_db_name") or _derive_tenant_db_name(tid)
                    if not db_name or db_name == DEFAULT_DB_NAME:
                        continue
                    tdb = client[db_name]
                    cand = await tdb.users.find_one({"email": email_key}, {"_id": 0})
                    if not cand:
                        cand = await tdb.users.find_one({"email": email_regex}, {"_id": 0})
                    if cand:
                        if not cand.get("tenant_id"):
                            cand["tenant_id"] = tid
                        user_doc = cand
                        # Ensure request-local routing now points to the resolved tenant DB.
                        if tid:
                            tenant_context.set_tenant(tid, "PRW")
                            _login_tenant_cache[email_key] = {
                                "tenant_id": tid,
                                "db_name": db_name,
                                "expires_at": time.time() + _LOGIN_TENANT_CACHE_TTL_SEC,
                            }
                        break
                _metrics["full_scan_ms"] = round((time.perf_counter() - _t_scan) * 1000, 1)
        except Exception as e:
            logger.warning("Cross-tenant login lookup failed for %s: %s", email_key, e)
    if not user_doc:
        _metrics["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
        logger.info("login failed metrics email=%s metrics=%s", email_key, _metrics)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Support both 'password' and 'password_hash' fields for compatibility
    _t_pwd = time.perf_counter()
    stored_password = user_doc.get('password') or user_doc.get('password_hash')
    if not stored_password or not verify_password(credentials.password, stored_password):
        _metrics["password_verify_ms"] = round((time.perf_counter() - _t_pwd) * 1000, 1)
        _metrics["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
        logger.info("login failed metrics email=%s metrics=%s", email_key, _metrics)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    _metrics["password_verify_ms"] = round((time.perf_counter() - _t_pwd) * 1000, 1)

    if user_doc.get("is_active") is False:
        raise HTTPException(status_code=403, detail="Account is disabled")

    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])

    # Remove both password fields before creating user object
    user_data = {k: v for k, v in user_doc.items() if k not in ['password', 'password_hash']}
    # Defensive defaults for migrated / partial user documents
    if not user_data.get('name') or str(user_data.get('name')).strip() == "":
        user_data['name'] = (user_data.get('email') or email_key or "user").split("@")[0] or "User"
    role_raw = user_data.get('role', UserRole.worker.value)
    if isinstance(role_raw, str):
        try:
            user_data['role'] = UserRole(role_raw)
        except ValueError:
            user_data['role'] = UserRole.worker

    try:
        user = User(**user_data)
    except Exception as e:
        logger.warning("Login: invalid user profile for %s: %s", email_key, e)
        raise HTTPException(
            status_code=500,
            detail="User account data is invalid. Ask an administrator to fix the user record.",
        )

    # Include tenant_id in JWT token for proper multi-tenant support
    tenant_id = user_doc.get('tenant_id', 'cli_001')
    pwd_ts = user_doc.get("password_changed_at") or user_doc.get("updated_at") or user_doc.get("created_at")
    access_token = create_access_token(
        data={"sub": str(user.email), "tenant_id": tenant_id, "pwd_ts": _normalize_pwd_timestamp(pwd_ts)}
    )
    # Return features in login response so client can skip /auth/me (faster load)
    _t_flags = time.perf_counter()
    try:
        features = await feature_service.get_all_flags(tenant_id)
    except Exception as e:
        logger.warning("get_all_flags failed during login for tenant %s: %s", tenant_id, e)
        features = {}
    _metrics["feature_flags_ms"] = round((time.perf_counter() - _t_flags) * 1000, 1)
    _metrics["resolved_tenant"] = tenant_id
    _metrics["total_ms"] = round((time.perf_counter() - _t0) * 1000, 1)
    logger.info("login success metrics email=%s metrics=%s", email_key, _metrics)
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=user,
        features=features,
        tenant_id=tenant_id,
        lot_number_prefix="PRW"
    )

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    # Add tenant info and feature flags to user response
    try:
        tenant_id = tenant_context.get_tenant()
    except Exception:
        tenant_id = getattr(current_user, "tenant_id", None) or "cli_001"
    features = await feature_service.get_all_flags(tenant_id)
    
    response = {
        **current_user.model_dump(),
        "tenant_id": tenant_id,
        "lot_number_prefix": tenant_context.lot_number_prefix,
        "features": features
    }
    
    # Add impersonation info if present
    if current_user.is_impersonated:
        response["is_impersonated"] = True
        response["impersonator"] = current_user.impersonator
        response["impersonator_name"] = current_user.impersonator_name
        response["session_id"] = current_user.session_id
    
    return response


@api_router.get("/public-config")
async def get_public_config():
    """Get public branding/config for the client ERP (no auth required). Cached 5 min."""
    try:
        tenant_id = tenant_context.get_tenant()
    except Exception:
        tenant_id = "default"
    cache_key = f"public_config:{tenant_id}"
    cached = _cache_get(cache_key, ttl_sec=300)
    if cached is not None:
        return cached
    config_keys = ["company_name", "sidebar_label", "primary_color", "login_bg_color", "logo_url", "favicon_url"]
    config_values = {}
    async for doc in db.tenant_config.find({"key": {"$in": config_keys}}, {"_id": 0}):
        config_values[doc["key"]] = doc.get("value", "")
    out = {
        "company_name": config_values.get("company_name") or "Prawn ERP",
        "sidebar_label": config_values.get("sidebar_label") or config_values.get("company_name") or "Prawn ERP",
        "primary_color": config_values.get("primary_color") or "#1e40af",
        "login_bg_color": config_values.get("login_bg_color") or "#0f1117",
        "logo_url": config_values.get("logo_url") or "",
        "favicon_url": config_values.get("favicon_url") or ""
    }
    _cache_set(cache_key, out, ttl_sec=300)
    return out

@api_router.post("/auth/impersonation/validate")
async def validate_impersonation(token: str):
    """Validate an impersonation token and return session info"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        
        if payload.get("type") != "impersonation":
            raise HTTPException(status_code=400, detail="Not an impersonation token")
        
        session_id = payload.get("session_id")
        tenant_id = payload.get("tenant_id")
        
        # Check if session is still valid
        session = await db.impersonation_tokens.find_one({
            "session_id": session_id,
            "tenant_id": tenant_id
        })
        
        if not session:
            raise HTTPException(status_code=401, detail="Impersonation session not found or expired")
        
        return {
            "valid": True,
            "session_id": session_id,
            "tenant_id": tenant_id,
            "email": payload.get("sub"),
            "impersonator": payload.get("impersonator"),
            "impersonator_name": payload.get("impersonator_name"),
            "expires_at": session.get("expires_at")
        }
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Impersonation token has expired")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid impersonation token")

# Agents endpoints
@api_router.post("/agents", response_model=Agent)
async def create_agent(agent_data: AgentCreate, current_user: User = Depends(get_current_user)):
    agent = Agent(**agent_data.model_dump())
    agent_dict = agent.model_dump()
    agent_dict['created_at'] = agent_dict['created_at'].isoformat()
    
    await db.agents.insert_one(agent_dict)
    _invalidate_agents_cache()
    await create_audit_log(current_user.id, "CREATE_AGENT", "procurement", {"agent_id": agent.id})
    return agent

def _invalidate_agents_cache():
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"agents_list:{tid}", None)
    except Exception:
        pass


@api_router.get("/agents", response_model=List[Agent])
async def get_agents(current_user: User = Depends(get_current_user)):
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"agents_list:{tid}"
    cached = _cache_get(cache_key, ttl_sec=30)
    if cached is not None:
        return cached
    agents = await db.agents.find({}, {"_id": 0}).to_list(1000)
    for agent in agents:
        created_at = agent.get("created_at")
        parsed = _safe_parse_datetime(created_at)
        if parsed is not None:
            agent["created_at"] = parsed
    _cache_set(cache_key, agents, ttl_sec=30)
    return agents

@api_router.get("/agents/{agent_id}", response_model=Agent)
async def get_agent(agent_id: str, current_user: User = Depends(get_current_user)):
    agent = await db.agents.find_one({"id": agent_id}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    created_at = agent.get("created_at")
    parsed = _safe_parse_datetime(created_at)
    if parsed is not None:
        agent["created_at"] = parsed
    return Agent(**agent)

# Procurement endpoints
@api_router.post("/procurement/lots", response_model=ProcurementLot)
async def create_procurement_lot(lot_data: ProcurementLotCreate, current_user: User = Depends(get_current_user)):
    agent = await db.agents.find_one({"id": lot_data.agent_id}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    
    net_weight = lot_data.gross_weight_kg - lot_data.ice_weight_kg
    total_amount = net_weight * lot_data.rate_per_kg
    balance = total_amount - lot_data.advance_paid
    
    payment_status = PaymentStatus.paid if balance <= 0 else (PaymentStatus.partial if lot_data.advance_paid > 0 else PaymentStatus.pending)
    
    lot = ProcurementLot(
        lot_number=generate_lot_number(),
        agent_id=lot_data.agent_id,
        agent_name=agent['name'],
        vehicle_number=lot_data.vehicle_number,
        driver_name=lot_data.driver_name,
        arrival_time=lot_data.arrival_time,
        species=lot_data.species,
        count_per_kg=lot_data.count_per_kg,
        boxes_count=lot_data.boxes_count,
        no_of_trays=lot_data.no_of_trays,
        gross_weight_kg=lot_data.gross_weight_kg,
        ice_weight_kg=lot_data.ice_weight_kg,
        net_weight_kg=net_weight,
        no_of_tons=net_weight / 1000,
        rate_per_kg=lot_data.rate_per_kg,
        total_amount=total_amount,
        advance_paid=lot_data.advance_paid,
        balance_due=balance,
        ice_ratio_pct=lot_data.ice_ratio_pct,
        freshness_grade=lot_data.freshness_grade,
        is_rejected=lot_data.is_rejected,
        rejection_reason=lot_data.rejection_reason,
        payment_status=payment_status,
        notes=lot_data.notes,
        created_by=current_user.id
    )
    
    lot_dict = lot.model_dump()
    lot_dict['created_at'] = lot_dict['created_at'].isoformat()
    lot_dict['arrival_time'] = lot_dict['arrival_time'].isoformat()
    
    await db.procurement_lots.insert_one(lot_dict)
    _invalidate_procurement_lots_cache()
    await create_audit_log(current_user.id, "CREATE_LOT", "procurement", {"lot_id": lot.id})
    # Auto-create gate_ice wastage record
    if lot.gross_weight_kg > 0:
        await create_wastage_record(
            lot_id=lot.id,
            stage_sequence=0,
            stage_name="Gate Ice Deduction",
            process_type="gate_ice",
            input_weight_kg=lot.gross_weight_kg,
            output_weight_kg=lot.net_weight_kg,
            source_entity_type="procurement_lot",
            source_entity_id=lot.id,
            recorded_by=current_user.id
        )
    
    return lot

def _invalidate_procurement_lots_cache():
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"procurement_lots:{tid}", None)
    except Exception:
        pass


@api_router.get("/procurement/lots", response_model=List[ProcurementLot])
async def get_procurement_lots(current_user: User = Depends(get_current_user)):
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"procurement_lots:{tid}"
    cached = _cache_get(cache_key, ttl_sec=30)
    if cached is not None:
        return cached
    lots = await db.procurement_lots.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for lot in lots:
        if isinstance(lot.get('created_at'), str):
            lot['created_at'] = datetime.fromisoformat(lot['created_at'])
        if isinstance(lot.get('arrival_time'), str):
            lot['arrival_time'] = datetime.fromisoformat(lot['arrival_time'])
        for payment in lot.get('payments', []):
            if isinstance(payment.get('payment_date'), str):
                payment['payment_date'] = datetime.fromisoformat(payment['payment_date'])
    _cache_set(cache_key, lots, ttl_sec=30)
    return lots


@api_router.get("/procurement/lots/agent-wise-count")
async def procurement_lots_agent_wise_count(
    limit: int = 10,
    current_user: User = Depends(get_current_user),
):
    """Return number of procurement lots grouped by agent."""
    limit = max(1, min(int(limit), 100))
    rows = await db.procurement_lots.aggregate([
        {
            "$group": {
                "_id": {"agent_id": "$agent_id", "agent_name": "$agent_name"},
                "lot_count": {"$sum": 1},
            }
        },
        {"$sort": {"lot_count": -1}},
        {"$limit": limit},
        {
            "$project": {
                "_id": 0,
                "agent_id": "$_id.agent_id",
                "agent_name": "$_id.agent_name",
                "lot_count": 1,
            }
        },
    ]).to_list(limit)
    return rows


@api_router.get("/procurement/lots/party-wise-count")
async def procurement_lots_party_wise_count(
    limit: int = 10,
    current_user: User = Depends(get_current_user),
):
    """Return number of procurement lots grouped by party (via purchase invoice)."""
    limit = max(1, min(int(limit), 100))
    rows = await db.procurement_lots.aggregate([
        {
            "$match": {
                "purchase_invoice_id": {"$exists": True, "$ne": None, "$ne": ""},
            }
        },
        {
            "$lookup": {
                "from": "purchase_invoices",
                "localField": "purchase_invoice_id",
                "foreignField": "id",
                "as": "invoice",
            }
        },
        {"$unwind": {"path": "$invoice", "preserveNullAndEmptyArrays": False}},
        {
            "$group": {
                "_id": {
                    "party_id": {"$ifNull": ["$invoice.party_id", ""]},
                    "party_name_text": {"$ifNull": ["$invoice.party_name_text", "Unknown"]},
                },
                "lot_count": {"$sum": 1},
            }
        },
        {"$sort": {"lot_count": -1}},
        {"$limit": limit},
        {
            "$project": {
                "_id": 0,
                "party_id": "$_id.party_id",
                "party_name_text": "$_id.party_name_text",
                "lot_count": 1,
            }
        },
    ]).to_list(limit)
    return rows

@api_router.get("/procurement/lots/{lot_id}", response_model=ProcurementLot)
async def get_procurement_lot(lot_id: str, current_user: User = Depends(get_current_user)):
    lot = await db.procurement_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    if isinstance(lot.get('created_at'), str):
        lot['created_at'] = datetime.fromisoformat(lot['created_at'])
    if isinstance(lot.get('arrival_time'), str):
        lot['arrival_time'] = datetime.fromisoformat(lot['arrival_time'])
    return ProcurementLot(**lot)

@api_router.get("/procurement/lots/{lot_id}/receipt")
async def download_receipt(lot_id: str, current_user: User = Depends(get_current_user)):
    lot = await db.procurement_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    if isinstance(lot.get('created_at'), str):
        lot['created_at'] = datetime.fromisoformat(lot['created_at'])
    if isinstance(lot.get('arrival_time'), str):
        lot['arrival_time'] = datetime.fromisoformat(lot['arrival_time'])
    
    agent = await db.agents.find_one({"id": lot['agent_id']}, {"_id": 0})
    if isinstance(agent.get('created_at'), str):
        agent['created_at'] = datetime.fromisoformat(agent['created_at'])
    
    lot_obj = ProcurementLot(**lot)
    agent_obj = Agent(**agent)
    
    pdf_bytes = generate_procurement_receipt_pdf(lot_obj, agent_obj)
    
    pdf_path = UPLOADS_DIR / f"receipt_{lot_id}.pdf"
    with open(pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    
    return FileResponse(
        path=pdf_path,
        media_type='application/pdf',
        filename=f"receipt_{lot_obj.lot_number}.pdf"
    )

# Preprocessing endpoints
@api_router.post("/preprocessing/batches", response_model=PreprocessingBatch)
async def create_preprocessing_batch(batch_data: PreprocessingBatchCreate, current_user: User = Depends(get_current_user)):
    waste_weight = batch_data.input_weight_kg - batch_data.output_weight_kg
    yield_pct = (batch_data.output_weight_kg / batch_data.input_weight_kg) * 100
    yield_alert = yield_pct < 75.0
    
    duration_mins = None
    if batch_data.end_time:
        duration = batch_data.end_time - batch_data.start_time
        duration_mins = duration.total_seconds() / 60
    
    batch = PreprocessingBatch(
        batch_number=generate_batch_number(),
        procurement_lot_id=batch_data.procurement_lot_id,
        process_type=batch_data.process_type,
        input_weight_kg=batch_data.input_weight_kg,
        output_weight_kg=batch_data.output_weight_kg,
        waste_weight_kg=waste_weight,
        yield_pct=yield_pct,
        yield_alert=yield_alert,
        start_time=batch_data.start_time,
        end_time=batch_data.end_time,
        duration_mins=duration_mins,
        workers=batch_data.workers,
        supervisor=batch_data.supervisor,
        notes=batch_data.notes,
        created_by=current_user.id
    )
    
    batch_dict = batch.model_dump()
    batch_dict['created_at'] = batch_dict['created_at'].isoformat()
    batch_dict['start_time'] = batch_dict['start_time'].isoformat()
    if batch_dict.get('end_time'):
        batch_dict['end_time'] = batch_dict['end_time'].isoformat()
    
    await db.preprocessing_batches.insert_one(batch_dict)
    await create_audit_log(current_user.id, "CREATE_BATCH", "preprocessing", {"batch_id": batch.id})
    
    # Auto-create wastage record for this preprocessing stage
    stage_sequence_map = {"heading": 1, "peeling": 2, "deveining": 3, "grading": 4}
    stage_sequence = stage_sequence_map.get(batch.process_type, 5)
    
    await create_wastage_record(
        lot_id=batch.procurement_lot_id,
        stage_sequence=stage_sequence,
        stage_name=batch.process_type.capitalize(),
        process_type=batch.process_type,
        input_weight_kg=batch.input_weight_kg,
        output_weight_kg=batch.output_weight_kg,
        source_entity_type="preprocessing_batch",
        source_entity_id=batch.id,
        recorded_by=current_user.id
    )
    
    return batch

@api_router.get("/preprocessing/batches", response_model=List[PreprocessingBatch])
async def get_preprocessing_batches(current_user: User = Depends(get_current_user)):
    batches = await db.preprocessing_batches.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for batch in batches:
        if isinstance(batch.get('created_at'), str):
            batch['created_at'] = datetime.fromisoformat(batch['created_at'])
        if isinstance(batch.get('start_time'), str):
            batch['start_time'] = datetime.fromisoformat(batch['start_time'])
        if batch.get('end_time') and isinstance(batch['end_time'], str):
            batch['end_time'] = datetime.fromisoformat(batch['end_time'])
    return batches

# Production endpoints
@api_router.post("/production/orders", response_model=ProductionOrder)
async def create_production_order(order_data: ProductionOrderCreate, current_user: User = Depends(get_current_user)):
    conversion_rate = (order_data.output_weight_kg / order_data.input_weight_kg) * 100
    
    order = ProductionOrder(
        order_number=generate_order_number(),
        preprocessing_batch_ids=order_data.preprocessing_batch_ids,
        product_form=order_data.product_form,
        target_size_count=order_data.target_size_count,
        glazing_pct=order_data.glazing_pct,
        block_weight_kg=order_data.block_weight_kg,
        no_of_blocks=order_data.no_of_blocks,
        input_weight_kg=order_data.input_weight_kg,
        output_weight_kg=order_data.output_weight_kg,
        conversion_rate_pct=conversion_rate,
        notes=order_data.notes,
        created_by=current_user.id
    )
    
    order_dict = order.model_dump()
    order_dict['created_at'] = order_dict['created_at'].isoformat()
    
    await db.production_orders.insert_one(order_dict)
    await create_audit_log(current_user.id, "CREATE_PRODUCTION_ORDER", "production", {"order_id": order.id})
    
    # Auto-create wastage record for production stage
    # Get the preprocessing batch to find the lot_id
    batch = await db.preprocessing_batches.find_one({"id": order.preprocessing_batch_id}, {"_id": 0})
    if batch:
        lot_id = batch.get("procurement_lot_id")
        
        # Determine stage sequence based on process_type
        stage_sequence_map = {"cooking": 5, "blanching": 5, "iqf_freezing": 6, "glazing": 7, "breading": 8}
        stage_sequence = stage_sequence_map.get(order.process_type, 9)
        
        await create_wastage_record(
            lot_id=lot_id,
            stage_sequence=stage_sequence,
            stage_name=order.process_type.capitalize(),
            process_type=order.process_type,
            input_weight_kg=order.input_weight_kg,
            output_weight_kg=order.output_weight_kg,
            source_entity_type="production_order",
            source_entity_id=order.id,
            recorded_by=current_user.id
        )
    
    return order

@api_router.get("/production/orders", response_model=List[ProductionOrder])
async def get_production_orders(current_user: User = Depends(get_current_user)):
    orders = await db.production_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for order in orders:
        if isinstance(order.get('created_at'), str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
    return orders

# Finished Goods endpoints
@api_router.post("/finished-goods", response_model=FinishedGood)
async def create_finished_good(fg_data: FinishedGoodCreate, current_user: User = Depends(get_current_user)):
    fg = FinishedGood(
        fg_code=generate_fg_code(),
        production_order_id=fg_data.production_order_id,
        product_form=fg_data.product_form,
        size_count=fg_data.size_count,
        weight_kg=fg_data.weight_kg,
        storage_location=fg_data.storage_location,
        temperature_c=fg_data.temperature_c,
        expiry_date=fg_data.expiry_date
    )
    
    fg_dict = fg.model_dump()
    fg_dict['created_at'] = fg_dict['created_at'].isoformat()
    fg_dict['manufactured_date'] = fg_dict['manufactured_date'].isoformat()
    if fg_dict.get('expiry_date'):
        fg_dict['expiry_date'] = fg_dict['expiry_date'].isoformat()
    
    await db.finished_goods.insert_one(fg_dict)
    _invalidate_finished_goods_cache()
    await create_audit_log(current_user.id, "CREATE_FINISHED_GOOD", "production", {"fg_id": fg.id})
    return fg

def _invalidate_finished_goods_cache():
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"finished_goods:{tid}", None)
    except Exception:
        pass


@api_router.get("/finished-goods", response_model=List[FinishedGood])
async def get_finished_goods(current_user: User = Depends(get_current_user)):
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"finished_goods:{tid}"
    cached = _cache_get(cache_key, ttl_sec=30)
    if cached is not None:
        return cached
    fgs = await db.finished_goods.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for fg in fgs:
        if isinstance(fg.get('created_at'), str):
            fg['created_at'] = datetime.fromisoformat(fg['created_at'])
        if isinstance(fg.get('manufactured_date'), str):
            fg['manufactured_date'] = datetime.fromisoformat(fg['manufactured_date'])
        if fg.get('expiry_date') and isinstance(fg['expiry_date'], str):
            fg['expiry_date'] = datetime.fromisoformat(fg['expiry_date'])
    _cache_set(cache_key, fgs, ttl_sec=30)
    return fgs

# Notifications
@api_router.post("/notifications", response_model=Notification)
async def create_notification(notif_data: NotificationCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.admin, UserRole.owner]:
        raise HTTPException(status_code=403, detail="Only admins can create notifications")
    
    notif = Notification(
        title=notif_data.title,
        message=notif_data.message,
        module=notif_data.module,
        target_roles=notif_data.target_roles,
        created_by=current_user.id
    )
    
    notif_dict = notif.model_dump()
    notif_dict['created_at'] = notif_dict['created_at'].isoformat()
    
    await db.notifications.insert_one(notif_dict)
    return notif

@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(current_user: User = Depends(get_current_user)):
    notifs = await db.notifications.find(
        {"target_roles": current_user.role.value},
        {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)
    
    for notif in notifs:
        if isinstance(notif.get('created_at'), str):
            notif['created_at'] = datetime.fromisoformat(notif['created_at'])
    return notifs

# Dashboard (optimized: counts + aggregation instead of loading all docs)
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    # Use count_documents and aggregation for speed (no full collection scan)
    lot_count = await db.procurement_lots.count_documents({})
    lot_agg = await db.procurement_lots.aggregate([
        {"$group": {"_id": None, "weight": {"$sum": {"$ifNull": ["$net_weight_kg", 0]}}, "value": {"$sum": {"$ifNull": ["$total_amount", 0]}}}}
    ]).to_list(1)
    total_value = lot_agg[0]["value"] if lot_agg else 0
    inv_agg = await db.purchase_invoices.aggregate([
        {"$group": {"_id": None, "total_quantity_kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}}}}
    ]).to_list(1)
    total_weight = (inv_agg[0]["total_quantity_kg"] if inv_agg else 0) or 0

    active_batches = await db.preprocessing_batches.count_documents({"end_time": None})
    active_orders = await db.production_orders.count_documents({"qc_status": "pending"})
    fg_agg = await db.finished_goods.aggregate([
        {"$group": {"_id": None, "inventory": {"$sum": "$weight_kg"}}}
    ]).to_list(1)
    fg_inventory = fg_agg[0]["inventory"] if fg_agg else 0
    pending_qc = await db.finished_goods.count_documents({"qc_status": "pending"})

    recent_lots = await db.procurement_lots.find({}, {"_id": 0, "lot_number": 1, "net_weight_kg": 1, "created_at": 1}).sort("created_at", -1).limit(5).to_list(5)
    recent_activities = [
        {"type": "procurement", "description": f"New lot {lot.get('lot_number')} - {(lot.get('net_weight_kg') or 0):.2f} KG", "timestamp": lot.get("created_at")}
        for lot in recent_lots
    ]

    return DashboardStats(
        total_procurement_lots=lot_count,
        total_weight_procured_kg=total_weight,
        total_procurement_value=total_value,
        active_preprocessing_batches=active_batches,
        active_production_orders=active_orders,
        finished_goods_inventory_kg=fg_inventory,
        pending_qc_items=pending_qc,
        recent_activities=recent_activities
    )


@api_router.get("/dashboard/overview")
async def get_dashboard_overview(current_user: User = Depends(get_current_user)):
    """Single endpoint for dashboard: stats + recent lots + batches + live prices. Cached 30s per tenant. Queries run in parallel for speed."""
    _t0 = time.perf_counter()
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"dashboard_overview:{tid}"
    _t_cache = time.perf_counter()
    cached = _cache_get(cache_key, ttl_sec=30)
    cache_lookup_ms = round((time.perf_counter() - _t_cache) * 1000, 1)
    if cached is not None:
        logger.debug(
            "dashboard_overview metrics tenant=%s cache_hit=%s cache_lookup_ms=%s total_ms=%s",
            tid,
            True,
            cache_lookup_ms,
            round((time.perf_counter() - _t0) * 1000, 1),
        )
        return cached

    # Run all independent MongoDB operations in parallel (biggest win after login)
    _t_queries = time.perf_counter()
    (
        lot_count,
        lot_agg_list,
        purchase_inv_agg_list,
        active_batches,
        active_orders,
        fg_agg_list,
        pending_qc,
        recent_lots_list,
        lots,
        batches,
    ) = await asyncio.gather(
        db.procurement_lots.count_documents({}),
        db.procurement_lots.aggregate([{"$group": {"_id": None, "weight": {"$sum": {"$ifNull": ["$net_weight_kg", 0]}}, "value": {"$sum": {"$ifNull": ["$total_amount", 0]}}}}]).to_list(1),
        db.purchase_invoices.aggregate([{"$group": {"_id": None, "total_quantity_kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}}}}]).to_list(1),
        db.preprocessing_batches.count_documents({"end_time": None}),
        db.production_orders.count_documents({"qc_status": "pending"}),
        db.finished_goods.aggregate([{"$group": {"_id": None, "inventory": {"$sum": "$weight_kg"}}}]).to_list(1),
        db.finished_goods.count_documents({"qc_status": "pending"}),
        db.procurement_lots.find({}, {"_id": 0, "lot_number": 1, "net_weight_kg": 1, "created_at": 1}).sort("created_at", -1).limit(5).to_list(5),
        db.procurement_lots.find({}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50),
        db.preprocessing_batches.find({}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50),
    )
    queries_ms = round((time.perf_counter() - _t_queries) * 1000, 1)

    lot_agg = lot_agg_list
    total_value = lot_agg[0]["value"] if lot_agg else 0
    # Weight procured = sum of kg from purchase invoices only
    total_weight = (purchase_inv_agg_list[0]["total_quantity_kg"] if purchase_inv_agg_list else 0) or 0
    fg_agg = fg_agg_list
    fg_inventory = fg_agg[0]["inventory"] if fg_agg else 0
    recent_activities = [
        {"type": "procurement", "description": f"New lot {lot.get('lot_number')} - {(lot.get('net_weight_kg') or 0):.2f} KG", "timestamp": lot.get("created_at")}
        for lot in recent_lots_list
    ]
    stats = {
        "total_procurement_lots": lot_count,
        "total_weight_procured_kg": total_weight,
        "total_procurement_value": total_value,
        "active_preprocessing_batches": active_batches,
        "active_production_orders": active_orders,
        "finished_goods_inventory_kg": fg_inventory,
        "pending_qc_items": pending_qc,
        "recent_activities": recent_activities,
    }
    for lot in lots:
        lot["created_at"] = _safe_parse_datetime(lot.get("created_at"))
        lot["arrival_time"] = _safe_parse_datetime(lot.get("arrival_time"))
        for p in lot.get("payments", []):
            p["payment_date"] = _safe_parse_datetime(p.get("payment_date"))
    for batch in batches:
        batch["created_at"] = _safe_parse_datetime(batch.get("created_at"))
        batch["start_time"] = _safe_parse_datetime(batch.get("start_time"))
        batch["end_time"] = _safe_parse_datetime(batch.get("end_time"))
    live_prices = [
        {"id": str(uuid.uuid4()), "category": "Vannamei 30/40", "price_per_kg": 420.0, "location": "Andhra Pradesh", "market": "Nellore", "date": datetime.now(timezone.utc), "source": "Market Data"},
        {"id": str(uuid.uuid4()), "category": "Vannamei 40/60", "price_per_kg": 380.0, "location": "Andhra Pradesh", "market": "Kakinada", "date": datetime.now(timezone.utc), "source": "Market Data"},
        {"id": str(uuid.uuid4()), "category": "Vannamei 60/80", "price_per_kg": 340.0, "location": "Andhra Pradesh", "market": "Bhimavaram", "date": datetime.now(timezone.utc), "source": "Market Data"},
        {"id": str(uuid.uuid4()), "category": "Black Tiger 20/30", "price_per_kg": 650.0, "location": "Andhra Pradesh", "market": "Nellore", "date": datetime.now(timezone.utc), "source": "Market Data"},
    ]
    out = {"stats": stats, "lots": lots, "batches": batches, "live_prices": live_prices}
    _t_cache_set = time.perf_counter()
    _cache_set(cache_key, out, ttl_sec=30)
    cache_set_ms = round((time.perf_counter() - _t_cache_set) * 1000, 1)
    logger.info(
        "dashboard_overview metrics tenant=%s cache_hit=%s cache_lookup_ms=%s queries_ms=%s cache_set_ms=%s total_ms=%s rows_lots=%s rows_batches=%s",
        tid,
        False,
        cache_lookup_ms,
        queries_ms,
        cache_set_ms,
        round((time.perf_counter() - _t0) * 1000, 1),
        len(lots),
        len(batches),
    )
    return out



# QC Module endpoints
@api_router.post("/qc/inspections", response_model=QCInspection)
async def create_qc_inspection(inspection_data: QCInspectionCreate, current_user: User = Depends(get_current_user)):
    inspection = QCInspection(
        inspection_code=generate_inspection_code(),
        entity_type=inspection_data.entity_type,
        entity_id=inspection_data.entity_id,
        qc_officer=inspection_data.qc_officer,
        parameters=inspection_data.parameters,
        overall_grade=inspection_data.overall_grade,
        pass_fail=inspection_data.pass_fail,
        failure_reason=inspection_data.failure_reason,
        lab_report_ref=inspection_data.lab_report_ref,
        notes=inspection_data.notes,
        created_by=current_user.id
    )
    
    inspection_dict = inspection.model_dump()
    inspection_dict['created_at'] = inspection_dict['created_at'].isoformat()
    inspection_dict['inspection_date'] = inspection_dict['inspection_date'].isoformat()
    
    await db.qc_inspections.insert_one(inspection_dict)
    await create_audit_log(current_user.id, "CREATE_QC_INSPECTION", "qc", {"inspection_id": inspection.id})
    return inspection

@api_router.get("/qc/inspections", response_model=List[QCInspection])
async def get_qc_inspections(current_user: User = Depends(get_current_user)):
    inspections = await db.qc_inspections.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for insp in inspections:
        if isinstance(insp.get('created_at'), str):
            insp['created_at'] = datetime.fromisoformat(insp['created_at'])
        if isinstance(insp.get('inspection_date'), str):
            insp['inspection_date'] = datetime.fromisoformat(insp['inspection_date'])
    return inspections

# Cold Storage endpoints
@api_router.post("/cold-storage/chambers", response_model=ColdStorageChamber)
async def create_chamber(chamber_data: ColdStorageChamberCreate, current_user: User = Depends(get_current_user)):
    chamber = ColdStorageChamber(**chamber_data.model_dump())
    chamber_dict = chamber.model_dump()
    chamber_dict['created_at'] = chamber_dict['created_at'].isoformat()
    
    await db.cold_storage_chambers.insert_one(chamber_dict)
    await create_audit_log(current_user.id, "CREATE_CHAMBER", "cold_storage", {"chamber_id": chamber.id})
    return chamber

@api_router.get("/cold-storage/chambers", response_model=List[ColdStorageChamber])
async def get_chambers(current_user: User = Depends(get_current_user)):
    chambers = await db.cold_storage_chambers.find({}, {"_id": 0}).to_list(1000)
    for chamber in chambers:
        if isinstance(chamber.get('created_at'), str):
            chamber['created_at'] = datetime.fromisoformat(chamber['created_at'])
    return chambers

@api_router.post("/cold-storage/slots", response_model=ColdStorageSlot)
async def create_slot(slot_data: ColdStorageSlotCreate, current_user: User = Depends(get_current_user)):
    slot = ColdStorageSlot(**slot_data.model_dump())
    slot_dict = slot.model_dump()
    slot_dict['created_at'] = slot_dict['created_at'].isoformat()
    if slot_dict.get('intake_date'):
        slot_dict['intake_date'] = slot_dict['intake_date'].isoformat()
    
    await db.cold_storage_slots.insert_one(slot_dict)
    return slot

@api_router.get("/cold-storage/slots", response_model=List[ColdStorageSlot])
async def get_slots(current_user: User = Depends(get_current_user)):
    slots = await db.cold_storage_slots.find({}, {"_id": 0}).to_list(1000)
    for slot in slots:
        if isinstance(slot.get('created_at'), str):
            slot['created_at'] = datetime.fromisoformat(slot['created_at'])
        if slot.get('intake_date') and isinstance(slot['intake_date'], str):
            slot['intake_date'] = datetime.fromisoformat(slot['intake_date'])
    return slots

@api_router.post("/cold-storage/inventory", response_model=ColdStorageInventory)
async def add_to_inventory(inventory_data: ColdStorageInventoryCreate, current_user: User = Depends(get_current_user)):
    intake_date = datetime.now(timezone.utc)
    days_in_storage = 0
    
    inventory = ColdStorageInventory(
        slot_id=inventory_data.slot_id,
        fg_id=inventory_data.fg_id,
        quantity_kg=inventory_data.quantity_kg,
        carton_count=inventory_data.carton_count,
        intake_date=intake_date,
        days_in_storage=days_in_storage
    )
    
    inventory_dict = inventory.model_dump()
    inventory_dict['created_at'] = inventory_dict['created_at'].isoformat()
    inventory_dict['intake_date'] = inventory_dict['intake_date'].isoformat()
    
    await db.cold_storage_inventory.insert_one(inventory_dict)
    
    # Update slot status
    await db.cold_storage_slots.update_one(
        {"id": inventory_data.slot_id},
        {"$set": {
            "status": "occupied",
            "occupied_weight_kg": inventory_data.quantity_kg,
            "fg_id": inventory_data.fg_id,
            "intake_date": intake_date.isoformat()
        }}
    )
    
    return inventory

@api_router.get("/cold-storage/inventory", response_model=List[ColdStorageInventory])
async def get_inventory(current_user: User = Depends(get_current_user)):
    inventory = await db.cold_storage_inventory.find({}, {"_id": 0}).to_list(1000)
    for item in inventory:
        if isinstance(item.get('created_at'), str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
        if isinstance(item.get('intake_date'), str):
            item['intake_date'] = datetime.fromisoformat(item['intake_date'])
        # Calculate days in storage
        if item.get('intake_date'):
            days = (datetime.now(timezone.utc) - item['intake_date']).days
            item['days_in_storage'] = days
    return inventory

@api_router.post("/cold-storage/temperature-logs", response_model=TemperatureLog)
async def log_temperature(log_data: TemperatureLogCreate, current_user: User = Depends(get_current_user)):
    chamber = await db.cold_storage_chambers.find_one({"id": log_data.chamber_id}, {"_id": 0})
    if not chamber:
        raise HTTPException(status_code=404, detail="Chamber not found")
    
    setpoint = chamber.get('setpoint_temperature_c', -18.0)
    temp_diff = abs(log_data.temperature_c - setpoint)
    alert = temp_diff > 2.0
    alert_reason = f"Temperature deviation: {temp_diff:.1f}°C from setpoint" if alert else None
    
    temp_log = TemperatureLog(
        chamber_id=log_data.chamber_id,
        temperature_c=log_data.temperature_c,
        alert=alert,
        alert_reason=alert_reason
    )
    
    temp_log_dict = temp_log.model_dump()
    temp_log_dict['recorded_at'] = temp_log_dict['recorded_at'].isoformat()
    
    await db.temperature_logs.insert_one(temp_log_dict)
    return temp_log

@api_router.get("/cold-storage/temperature-logs", response_model=List[TemperatureLog])
async def get_temperature_logs(chamber_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"chamber_id": chamber_id} if chamber_id else {}
    logs = await db.temperature_logs.find(query, {"_id": 0}).sort("recorded_at", -1).limit(100).to_list(100)
    for log in logs:
        if isinstance(log.get('recorded_at'), str):
            log['recorded_at'] = datetime.fromisoformat(log['recorded_at'])
    return logs

# Sales & Dispatch endpoints
@api_router.post("/buyers", response_model=Buyer)
async def create_buyer(buyer_data: BuyerCreate, current_user: User = Depends(get_current_user)):
    buyer = Buyer(**buyer_data.model_dump())
    buyer_dict = buyer.model_dump()
    buyer_dict['created_at'] = buyer_dict['created_at'].isoformat()
    
    await db.buyers.insert_one(buyer_dict)
    await create_audit_log(current_user.id, "CREATE_BUYER", "sales", {"buyer_id": buyer.id})
    return buyer

@api_router.get("/buyers", response_model=List[Buyer])
async def get_buyers(current_user: User = Depends(get_current_user)):
    buyers = await db.buyers.find({}, {"_id": 0}).to_list(1000)
    for buyer in buyers:
        if isinstance(buyer.get('created_at'), str):
            buyer['created_at'] = datetime.fromisoformat(buyer['created_at'])
    return buyers

@api_router.post("/sales/orders", response_model=SalesOrder)
async def create_sales_order(order_data: SalesOrderCreate, current_user: User = Depends(get_current_user)):
    buyer = await db.buyers.find_one({"id": order_data.buyer_id}, {"_id": 0})
    if not buyer:
        raise HTTPException(status_code=404, detail="Buyer not found")
    
    total_value = order_data.quantity_kg * order_data.rate_per_kg_usd
    
    sales_order = SalesOrder(
        order_number=generate_order_num(),
        buyer_id=order_data.buyer_id,
        buyer_name=buyer['company_name'],
        quantity_kg=order_data.quantity_kg,
        rate_per_kg_usd=order_data.rate_per_kg_usd,
        currency=order_data.currency,
        total_value_usd=total_value,
        delivery_date=order_data.delivery_date,
        notes=order_data.notes,
        created_by=current_user.id
    )
    
    order_dict = sales_order.model_dump()
    order_dict['created_at'] = order_dict['created_at'].isoformat()
    order_dict['delivery_date'] = order_dict['delivery_date'].isoformat()
    
    await db.sales_orders.insert_one(order_dict)
    await create_audit_log(current_user.id, "CREATE_SALES_ORDER", "sales", {"order_id": sales_order.id})
    return sales_order

@api_router.get("/sales/orders", response_model=List[SalesOrder])
async def get_sales_orders(current_user: User = Depends(get_current_user)):
    orders = await db.sales_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for order in orders:
        if isinstance(order.get('created_at'), str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
        if isinstance(order.get('delivery_date'), str):
            order['delivery_date'] = datetime.fromisoformat(order['delivery_date'])
    return orders

@api_router.post("/shipments", response_model=Shipment)
async def create_shipment(shipment_data: ShipmentCreate, current_user: User = Depends(get_current_user)):
    shipment = Shipment(
        shipment_number=generate_shipment_number(),
        sales_order_id=shipment_data.sales_order_id,
        container_no=shipment_data.container_no,
        seal_no=shipment_data.seal_no,
        shipping_line=shipment_data.shipping_line,
        vessel_name=shipment_data.vessel_name,
        port_of_loading=shipment_data.port_of_loading,
        port_of_discharge=shipment_data.port_of_discharge,
        destination_country=shipment_data.destination_country,
        etd=shipment_data.etd,
        eta=shipment_data.eta,
        bill_of_lading=shipment_data.bill_of_lading,
        created_by=current_user.id
    )
    
    shipment_dict = shipment.model_dump()
    shipment_dict['created_at'] = shipment_dict['created_at'].isoformat()
    shipment_dict['etd'] = shipment_dict['etd'].isoformat()
    shipment_dict['eta'] = shipment_dict['eta'].isoformat()
    
    await db.shipments.insert_one(shipment_dict)
    await create_audit_log(current_user.id, "CREATE_SHIPMENT", "sales", {"shipment_id": shipment.id})
    return shipment

@api_router.get("/shipments", response_model=List[Shipment])
async def get_shipments(current_user: User = Depends(get_current_user)):
    shipments = await db.shipments.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for shipment in shipments:
        if isinstance(shipment.get('created_at'), str):
            shipment['created_at'] = datetime.fromisoformat(shipment['created_at'])
        if isinstance(shipment.get('etd'), str):
            shipment['etd'] = datetime.fromisoformat(shipment['etd'])
        if isinstance(shipment.get('eta'), str):
            shipment['eta'] = datetime.fromisoformat(shipment['eta'])
    return shipments

# Wage & Billing endpoints
@api_router.post("/wage-bills", response_model=WageBill)
async def create_wage_bill(bill_data: WageBillCreate, current_user: User = Depends(get_current_user)):
    net_payable = bill_data.gross_amount - bill_data.tds_deduction
    
    wage_bill = WageBill(
        bill_number=generate_bill_number(),
        bill_type=bill_data.bill_type,
        period_from=bill_data.period_from,
        period_to=bill_data.period_to,
        department=bill_data.department,
        gross_amount=bill_data.gross_amount,
        tds_deduction=bill_data.tds_deduction,
        net_payable=net_payable,
        line_items=bill_data.line_items,
        notes=bill_data.notes,
        created_by=current_user.id
    )
    
    bill_dict = wage_bill.model_dump()
    bill_dict['created_at'] = bill_dict['created_at'].isoformat()
    bill_dict['period_from'] = bill_dict['period_from'].isoformat()
    bill_dict['period_to'] = bill_dict['period_to'].isoformat()
    if bill_dict.get('payment_date'):
        bill_dict['payment_date'] = bill_dict['payment_date'].isoformat()
    
    await db.wage_bills.insert_one(bill_dict)
    await create_audit_log(current_user.id, "CREATE_WAGE_BILL", "accounts", {"bill_id": wage_bill.id})
    return wage_bill

@api_router.get("/wage-bills", response_model=List[WageBill])
async def get_wage_bills(
    bill_type: Optional[str] = None,
    department: Optional[str] = None,
    payment_status: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all wage bills with optional filters"""
    query = {}
    
    if bill_type:
        query['bill_type'] = bill_type
    if department:
        query['department'] = department
    if payment_status:
        query['payment_status'] = payment_status
    if period_from:
        query['period_from'] = {"$gte": datetime.fromisoformat(period_from)}
    if period_to:
        if 'period_to' not in query:
            query['period_to'] = {}
        query['period_to']['$lte'] = datetime.fromisoformat(period_to)
    
    bills = await db.wage_bills.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for bill in bills:
        if isinstance(bill.get('created_at'), str):
            bill['created_at'] = datetime.fromisoformat(bill['created_at'])
        if isinstance(bill.get('period_from'), str):
            bill['period_from'] = datetime.fromisoformat(bill['period_from'])
        if isinstance(bill.get('period_to'), str):
            bill['period_to'] = datetime.fromisoformat(bill['period_to'])
        if bill.get('payment_date') and isinstance(bill['payment_date'], str):
            bill['payment_date'] = datetime.fromisoformat(bill['payment_date'])
    return bills

@api_router.get("/wage-bills/{bill_id}", response_model=WageBill)
async def get_wage_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    """Get a single wage bill by ID"""
    bill = await db.wage_bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Wage bill not found")
    
    if isinstance(bill.get('created_at'), str):
        bill['created_at'] = datetime.fromisoformat(bill['created_at'])
    if isinstance(bill.get('period_from'), str):
        bill['period_from'] = datetime.fromisoformat(bill['period_from'])
    if isinstance(bill.get('period_to'), str):
        bill['period_to'] = datetime.fromisoformat(bill['period_to'])
    if bill.get('payment_date') and isinstance(bill['payment_date'], str):
        bill['payment_date'] = datetime.fromisoformat(bill['payment_date'])
    
    return bill

@api_router.put("/wage-bills/{bill_id}", response_model=WageBill)
async def update_wage_bill(bill_id: str, bill_data: WageBillCreate, current_user: User = Depends(get_current_user)):
    """Update an existing wage bill"""
    existing_bill = await db.wage_bills.find_one({"id": bill_id}, {"_id": 0})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Wage bill not found")
    
    net_payable = bill_data.gross_amount - bill_data.tds_deduction
    
    update_data = {
        "bill_type": bill_data.bill_type,
        "period_from": bill_data.period_from.isoformat(),
        "period_to": bill_data.period_to.isoformat(),
        "department": bill_data.department,
        "gross_amount": bill_data.gross_amount,
        "tds_deduction": bill_data.tds_deduction,
        "net_payable": net_payable,
        "line_items": bill_data.line_items,
        "notes": bill_data.notes
    }
    
    await db.wage_bills.update_one({"id": bill_id}, {"$set": update_data})
    await create_audit_log(current_user.id, "UPDATE_WAGE_BILL", "accounts", {"bill_id": bill_id})
    
    # Fetch and return updated bill
    updated_bill = await db.wage_bills.find_one({"id": bill_id}, {"_id": 0})
    if isinstance(updated_bill.get('created_at'), str):
        updated_bill['created_at'] = datetime.fromisoformat(updated_bill['created_at'])
    if isinstance(updated_bill.get('period_from'), str):
        updated_bill['period_from'] = datetime.fromisoformat(updated_bill['period_from'])
    if isinstance(updated_bill.get('period_to'), str):
        updated_bill['period_to'] = datetime.fromisoformat(updated_bill['period_to'])
    if updated_bill.get('payment_date') and isinstance(updated_bill['payment_date'], str):
        updated_bill['payment_date'] = datetime.fromisoformat(updated_bill['payment_date'])
    
    return updated_bill

@api_router.post("/wage-bills/{bill_id}/mark-paid")
async def mark_wage_bill_paid(bill_id: str, current_user: User = Depends(get_current_user)):
    """Mark a wage bill as paid"""
    bill = await db.wage_bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Wage bill not found")
    
    payment_date = datetime.now(timezone.utc)
    await db.wage_bills.update_one(
        {"id": bill_id},
        {"$set": {
            "payment_status": "paid",
            "payment_date": payment_date.isoformat()
        }}
    )
    
    await create_audit_log(current_user.id, "MARK_PAID_WAGE_BILL", "accounts", {"bill_id": bill_id})
    
    return {"status": "success", "message": "Wage bill marked as paid", "payment_date": payment_date.isoformat()}

@api_router.delete("/wage-bills/{bill_id}")
async def delete_wage_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    """Delete a wage bill"""
    result = await db.wage_bills.delete_one({"id": bill_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Wage bill not found")
    
    await create_audit_log(current_user.id, "DELETE_WAGE_BILL", "accounts", {"bill_id": bill_id})
    
    return {"status": "success", "message": "Wage bill deleted"}

@api_router.get("/wage-bills/{bill_id}/pdf")
async def download_wage_bill_pdf(bill_id: str, current_user: User = Depends(get_current_user)):
    """Download wage bill as PDF"""
    bill = await db.wage_bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Wage bill not found")
    
    # Convert datetime strings back to datetime objects
    if isinstance(bill.get('created_at'), str):
        bill['created_at'] = datetime.fromisoformat(bill['created_at'])
    if isinstance(bill.get('period_from'), str):
        bill['period_from'] = datetime.fromisoformat(bill['period_from'])
    if isinstance(bill.get('period_to'), str):
        bill['period_to'] = datetime.fromisoformat(bill['period_to'])
    if bill.get('payment_date') and isinstance(bill['payment_date'], str):
        bill['payment_date'] = datetime.fromisoformat(bill['payment_date'])
    
    # Create WageBill object
    wage_bill = WageBill(**bill)
    
    # Generate PDF
    pdf_bytes = generate_wage_bill_pdf(wage_bill)
    
    # Return as downloadable file
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=wage_bill_{bill["bill_number"]}.pdf'
        }
    )

# ══════════════════════════════════════════════════════════════════════════════
# Purchase Invoice Endpoints (Amendment A4)
# ══════════════════════════════════════════════════════════════════════════════

def normalize_invoice_advance_balance(invoice: dict) -> None:
    """Ensure advance_paid and balance_due are set on invoice dict (for list, single, PDF, ledger).
    Always derive balance_due from grand_total - advance_paid so stored zeros don't hide real amounts.
    """
    if not invoice:
        return
    advance = _pdf_safe_float(invoice.get("advance_paid"), 0.0)
    invoice["advance_paid"] = advance
    grand = _pdf_safe_float(invoice.get("grand_total"), 0.0)
    # Always compute from source of truth so metrics/list never show wrong zeros
    invoice["balance_due"] = round(grand - advance, 2)


def _normalize_purchase_invoice_list_sub_tab(raw: Optional[str]) -> Optional[str]:
    if not raw or not str(raw).strip():
        return None
    v = str(raw).strip().lower()
    if v in ("pending", "pushed", "audit"):
        return v
    return None


def _build_purchase_invoice_list_query(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    payment_status: Optional[str] = None,
    invoice_status: Optional[str] = None,
    agent_name: Optional[str] = None,
    party_name: Optional[str] = None,
    search: Optional[str] = None,
    list_sub_tab: Optional[str] = None,
) -> dict:
    """Build MongoDB query for purchase invoice list/metrics.
    Same logic for both so stats cards match filters.
    list_sub_tab: pending (status != pushed), pushed, audit (is_manually_recorded).
    """
    query: dict = {}
    if from_date or to_date:
        query["invoice_date"] = {}
        if from_date:
            query["invoice_date"]["$gte"] = from_date
        if to_date:
            query["invoice_date"]["$lte"] = to_date
    if payment_status and payment_status.strip():
        statuses = [s.strip() for s in payment_status.split(",") if s.strip()]
        if statuses:
            query["payment_status"] = {"$in": statuses}
    if invoice_status and invoice_status.strip():
        statuses = [s.strip() for s in invoice_status.split(",") if s.strip()]
        if statuses:
            query["status"] = {"$in": statuses}

    if agent_name and agent_name.strip():
        agent_term = agent_name.strip()
        query["agent_ref_name"] = {"$regex": agent_term, "$options": "i"}

    if party_name and party_name.strip():
        party_term = party_name.strip()
        query["party_name_text"] = {"$regex": party_term, "$options": "i"}

    if search and search.strip():
        search_term = search.strip()
        query["$or"] = [
            {"farmer_name": {"$regex": search_term, "$options": "i"}},
            {"invoice_no": {"$regex": search_term, "$options": "i"}},
            {"agent_ref_name": {"$regex": search_term, "$options": "i"}},
            {"party_name_text": {"$regex": search_term, "$options": "i"}},
        ]

    tab = _normalize_purchase_invoice_list_sub_tab(list_sub_tab)
    sub_tab_filter: Optional[dict] = None
    if tab == "pending":
        sub_tab_filter = {"status": {"$ne": "pushed"}}
    elif tab == "pushed":
        sub_tab_filter = {"status": "pushed"}
    elif tab == "audit":
        sub_tab_filter = {"is_manually_recorded": True}

    if sub_tab_filter:
        if not query:
            return sub_tab_filter
        return {"$and": [dict(query), sub_tab_filter]}
    return query


async def _purchase_invoice_metrics_aggregation(query: dict):
    """Run aggregation to get count + metrics. Total kg = sum of total_quantity_kg on each matched invoice
    (same field as the list/grid), so stats match the active filters including list_sub_tab."""
    pipeline = [
        {"$match": query},
        {"$addFields": {
            "balance_due": {"$round": [{"$subtract": [{"$ifNull": ["$grand_total", 0]}, {"$ifNull": ["$advance_paid", 0]}]}, 2]},
        }},
        {"$group": {
            "_id": None,
            "total_count": {"$sum": 1},
            "total_value": {"$sum": {"$ifNull": ["$grand_total", 0]}},
            "total_quantity_kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}},
            "advances_paid_total": {"$sum": {"$ifNull": ["$advance_paid", 0]}},
            "advances_paid_count": {"$sum": {"$cond": [{"$gt": [{"$ifNull": ["$advance_paid", 0]}, 0]}, 1, 0]}},
            "partial_count": {"$sum": {"$cond": [{"$eq": ["$payment_status", "partial"]}, 1, 0]}},
            "partial_total": {"$sum": {"$cond": [{"$eq": ["$payment_status", "partial"]}, "$balance_due", 0]}},
        }},
    ]
    cursor = db.purchase_invoices.aggregate(pipeline)
    row = await cursor.to_list(1)
    if not row:
        return {
            "total_count": 0,
            "total_value": 0,
            "total_quantity_kg": 0,
            "partial_count": 0,
            "partial_total": 0,
            "advances_paid_count": 0,
            "advances_paid_total": 0,
        }
    r = row[0]
    return {
        "total_count": r["total_count"],
        "total_value": round(float(r["total_value"]), 2),
        "total_quantity_kg": round(float(r.get("total_quantity_kg") or 0), 3),
        "partial_count": r["partial_count"],
        "partial_total": round(float(r["partial_total"]), 2),
        "advances_paid_count": r["advances_paid_count"],
        "advances_paid_total": round(float(r["advances_paid_total"]), 2),
    }


async def _purchase_invoice_sub_tab_counts(query: dict) -> dict:
    """Counts per list sub-tab for the same filter set (excluding list_sub_tab). Overlap allowed (e.g. pushed + audit)."""
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "pending": {"$sum": {"$cond": [{"$ne": ["$status", "pushed"]}, 1, 0]}},
            "pushed": {"$sum": {"$cond": [{"$eq": ["$status", "pushed"]}, 1, 0]}},
            "audit": {"$sum": {"$cond": [{"$eq": ["$is_manually_recorded", True]}, 1, 0]}},
        }},
    ]
    cursor = db.purchase_invoices.aggregate(pipeline)
    row = await cursor.to_list(1)
    if not row:
        return {"pending": 0, "pushed": 0, "audit": 0}
    r = row[0]
    return {
        "pending": int(r.get("pending") or 0),
        "pushed": int(r.get("pushed") or 0),
        "audit": int(r.get("audit") or 0),
    }


async def _purchase_invoice_total_kg_all() -> float:
    """Sum of quantity_kg from line items across ALL purchase invoices. Aggregates directly on lines collection (no lookup)."""
    cursor = db.purchase_invoice_lines.aggregate([
        {"$group": {"_id": None, "total_quantity_kg": {"$sum": {"$ifNull": ["$quantity_kg", 0]}}}},
    ])
    row = await cursor.to_list(1)
    if not row:
        return 0.0
    return float(row[0].get("total_quantity_kg") or 0)


@api_router.post("/purchase-invoices", response_model=PurchaseInvoice)
async def create_purchase_invoice(
    invoice_data: PurchaseInvoiceCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new purchase invoice (draft status)"""
    # Generate invoice number
    invoice_no = await generate_invoice_number()
    
    # Calculate totals
    totals = calculate_invoice_totals(invoice_data.line_items, invoice_data.tds_rate_pct)
    
    # Create invoice
    invoice = PurchaseInvoice(
        invoice_no=invoice_no,
        invoice_date=invoice_data.invoice_date,
        farmer_name=invoice_data.farmer_name,
        farmer_mobile=invoice_data.farmer_mobile,
        farmer_location=invoice_data.farmer_location,
        agent_ref_name=invoice_data.agent_ref_name,
        weighment_slip_no=invoice_data.weighment_slip_no,
        weighment_slip_file_url=invoice_data.weighment_slip_file_url,
        weighment_slip_mime_type=invoice_data.weighment_slip_mime_type,
        custom_field_1_label=invoice_data.custom_field_1_label,
        custom_field_1_value=invoice_data.custom_field_1_value,
        custom_field_2_label=invoice_data.custom_field_2_label,
        custom_field_2_value=invoice_data.custom_field_2_value,
        tds_rate_pct=invoice_data.tds_rate_pct,
        advance_paid=invoice_data.advance_paid,
        notes=invoice_data.notes,
        created_by=current_user.id,
        # A5: Party fields
        party_id=invoice_data.party_id,
        party_name_text=invoice_data.party_name_text if not invoice_data.same_as_farmer else invoice_data.farmer_name,
        **totals
    )
    
    # Calculate balance due
    invoice.balance_due = invoice.grand_total - invoice.advance_paid
    
    # Update payment status
    if invoice.balance_due <= 0:
        invoice.payment_status = PaymentStatus.paid
    elif invoice.advance_paid > 0:
        invoice.payment_status = PaymentStatus.partial
    else:
        invoice.payment_status = PaymentStatus.pending
    
    # Save invoice
    invoice_dict = invoice.model_dump()
    invoice_dict['invoice_date'] = invoice_dict['invoice_date'].isoformat()
    invoice_dict['created_at'] = invoice_dict['created_at'].isoformat()
    invoice_dict['updated_at'] = invoice_dict['updated_at'].isoformat()
    if invoice_dict.get('approved_at'):
        invoice_dict['approved_at'] = invoice_dict['approved_at'].isoformat()
    if invoice_dict.get('pushed_at'):
        invoice_dict['pushed_at'] = invoice_dict['pushed_at'].isoformat()
    
    # Remove line_items from main doc (save separately)
    line_items_data = invoice_dict.pop('line_items', [])
    
    await db.purchase_invoices.insert_one(invoice_dict)
    
    # Save line items
    for line_data in invoice_data.line_items:
        line = PurchaseInvoiceLine(
            invoice_id=invoice.id,
            line_no=line_data.line_no,
            variety=line_data.variety,
            count_value=line_data.count_value,
            custom_variety_notes=line_data.custom_variety_notes,
            custom_count_notes=line_data.custom_count_notes,
            quantity_kg=line_data.quantity_kg,
            rate=line_data.rate,
            amount=round(line_data.quantity_kg * line_data.rate, 2)
        )
        line_dict = line.model_dump()
        await db.purchase_invoice_lines.insert_one(line_dict)
    
    # Audit log
    await create_audit_log(current_user.id, "CREATE_PURCHASE_INVOICE", "procurement", 
                          {"invoice_id": invoice.id, "invoice_no": invoice_no})
    
    # Reload with line items for response
    invoice.line_items = [
        PurchaseInvoiceLine(
            invoice_id=invoice.id,
            line_no=line_data.line_no,
            variety=line_data.variety,
            count_value=line_data.count_value,
            custom_variety_notes=line_data.custom_variety_notes,
            custom_count_notes=line_data.custom_count_notes,
            quantity_kg=line_data.quantity_kg,
            rate=line_data.rate,
            amount=round(line_data.quantity_kg * line_data.rate, 2)
        ) for line_data in invoice_data.line_items
    ]
    return invoice

@api_router.get("/purchase-invoices")
async def get_purchase_invoices(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    payment_status: Optional[str] = None,
    invoice_status: Optional[str] = None,
    agent_name: Optional[str] = None,
    party_name: Optional[str] = None,
    search: Optional[str] = None,
    list_sub_tab: Optional[str] = Query(None, description="List slice: pending (not pushed), pushed, or audit (manually recorded)"),
    page: int = 1,
    per_page: int = 25,
    sort: str = "invoice_date:desc",
    debug_timing: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get purchase invoices with filters and pagination. Omitted or empty filters mean no restriction. Uses aggregation for metrics (fast) and paginated find for list. Use debug_timing=1 to get timing breakdown in response."""
    base_query = _build_purchase_invoice_list_query(
        from_date,
        to_date,
        payment_status,
        invoice_status,
        agent_name=agent_name,
        party_name=party_name,
        search=search,
        list_sub_tab=None,
    )
    query = _build_purchase_invoice_list_query(
        from_date,
        to_date,
        payment_status,
        invoice_status,
        agent_name=agent_name,
        party_name=party_name,
        search=search,
        list_sub_tab=list_sub_tab,
    )
    sort_field, sort_dir = sort.split(':')
    sort_direction = -1 if sort_dir == 'desc' else 1
    skip = (page - 1) * per_page

    t0 = time.perf_counter() if debug_timing else None

    # Run metrics aggregation, page fetch, all-invoices total kg, and sub-tab counts in parallel
    async def fetch_page():
        cursor = db.purchase_invoices.find(query, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(per_page)
        return await cursor.to_list(per_page)

    metrics, invoices, total_kg_all, sub_tab_counts = await asyncio.gather(
        _purchase_invoice_metrics_aggregation(query),
        fetch_page(),
        _purchase_invoice_total_kg_all(),
        _purchase_invoice_sub_tab_counts(base_query),
    )
    metrics["sub_tab_counts"] = sub_tab_counts
    metrics["total_quantity_kg_all"] = round(total_kg_all, 3)  # All invoices, not just selected filters
    total = metrics["total_count"]
    pages = (total + per_page - 1) // per_page if per_page else 0

    t1 = time.perf_counter() if debug_timing else None

    for inv in invoices:
        normalize_invoice_advance_balance(inv)

    # Batch party short_code lookup (one query instead of N)
    party_ids = list({inv["party_id"] for inv in invoices if inv.get("party_id")})
    if party_ids:
        parties = await db.parties.find({"id": {"$in": party_ids}}, {"_id": 0, "id": 1, "short_code": 1}).to_list(len(party_ids))
        party_map = {p["id"]: p.get("short_code") for p in parties}
        for inv in invoices:
            if inv.get("party_id"):
                inv["party_short_code"] = party_map.get(inv["party_id"])

    t2 = time.perf_counter() if debug_timing else None

    out = {
        "data": invoices,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "metrics": metrics,
    }
    if debug_timing and t0 is not None and t1 is not None and t2 is not None:
        out["debug_timing_ms"] = {
            "metrics_and_page_ms": round((t1 - t0) * 1000, 1),
            "party_lookup_ms": round((t2 - t1) * 1000, 1),
            "total_ms": round((t2 - t0) * 1000, 1),
        }
    return out

@api_router.get("/purchase-invoices/metrics")
async def get_purchase_invoice_metrics(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    payment_status: Optional[str] = None,
    invoice_status: Optional[str] = None,
    agent_name: Optional[str] = None,
    party_name: Optional[str] = None,
    search: Optional[str] = None,
    list_sub_tab: Optional[str] = Query(None, description="Same as GET /purchase-invoices list_sub_tab"),
    current_user: User = Depends(get_current_user)
):
    """Get metrics for purchase invoices dashboard. Uses aggregation (fast). Omitted/empty = All (no restriction)."""
    base_query = _build_purchase_invoice_list_query(
        from_date,
        to_date,
        payment_status,
        invoice_status,
        agent_name=agent_name,
        party_name=party_name,
        search=search,
        list_sub_tab=None,
    )
    query = _build_purchase_invoice_list_query(
        from_date,
        to_date,
        payment_status,
        invoice_status,
        agent_name=agent_name,
        party_name=party_name,
        search=search,
        list_sub_tab=list_sub_tab,
    )
    metrics, sub_tab_counts = await asyncio.gather(
        _purchase_invoice_metrics_aggregation(query),
        _purchase_invoice_sub_tab_counts(base_query),
    )
    # Add fields the frontend or other callers may expect from the standalone metrics endpoint
    pending_total = metrics["partial_total"]  # approximate; full pending would need extra aggregation
    pending_count = metrics["partial_count"]
    return {
        "total_count": metrics["total_count"],
        "total_value": metrics["total_value"],
        "total_quantity_kg": metrics["total_quantity_kg"],
        "pending_count": pending_count,
        "pending_total": pending_total,
        "partial_count": metrics["partial_count"],
        "partial_total": metrics["partial_total"],
        "paid_count": metrics["advances_paid_count"],
        "paid_total": metrics["advances_paid_total"],
        "advances_paid_count": metrics["advances_paid_count"],
        "advances_paid_total": metrics["advances_paid_total"],
        "outstanding_total": pending_total,
        "top_farmers": [],  # omit heavy top_farmers for speed; use list endpoint if needed
        "sub_tab_counts": sub_tab_counts,
    }


@api_router.get("/purchase-invoices/filter-options")
async def get_purchase_invoice_filter_options(current_user: User = Depends(get_current_user)):
    """Get distinct dropdown options for purchase invoice filters."""
    agent_rows = await db.purchase_invoices.aggregate([
        {"$match": {"agent_ref_name": {"$exists": True, "$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$agent_ref_name"}},
        {"$sort": {"_id": 1}},
        {"$project": {"_id": 0, "name": "$_id"}},
    ]).to_list(1000)
    party_rows = await db.purchase_invoices.aggregate([
        {"$match": {"party_name_text": {"$exists": True, "$ne": None, "$ne": ""}}},
        {"$group": {"_id": "$party_name_text"}},
        {"$sort": {"_id": 1}},
        {"$project": {"_id": 0, "name": "$_id"}},
    ]).to_list(1000)
    return {
        "agents": [r["name"] for r in agent_rows if r.get("name")],
        "parties": [r["name"] for r in party_rows if r.get("name")],
    }


def _compress_weighment_slip_image_bytes(raw: bytes) -> tuple[bytes, str]:
    """Resize large images and re-encode as JPEG for smaller storage (weighment slip scans/photos)."""
    from PIL import Image

    img = Image.open(io.BytesIO(raw))
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    elif img.mode != "RGB":
        img = img.convert("RGB")

    max_side = 2048
    w, h = img.size
    if w <= 0 or h <= 0:
        raise ValueError("Invalid image dimensions")
    if max(w, h) > max_side:
        ratio = max_side / float(max(w, h))
        img = img.resize((max(1, int(w * ratio)), max(1, int(h * ratio))), Image.Resampling.LANCZOS)

    out = io.BytesIO()
    img.save(out, format="JPEG", quality=82, optimize=True)
    data = out.getvalue()
    if len(data) > 1_500_000:
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=68, optimize=True)
        data = out.getvalue()
    return data, "image/jpeg"


def _safe_delete_weighment_upload_file(public_url: Optional[str]) -> None:
    """Remove a file under uploads/ if it matches our weighment slip naming (and exists)."""
    if not public_url:
        return
    path = _signature_path_from_config_value(public_url)
    if path is None or not path.is_file():
        return
    base = path.name
    if not base.startswith("weighment_slip_"):
        return
    try:
        path.unlink()
    except OSError:
        pass


@api_router.post("/purchase-invoices/upload-weighment-slip")
async def upload_purchase_invoice_weighment_slip(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload a weighment slip image; server compresses to JPEG (max side 2048px)."""
    raw_name = (file.filename or "").lower()
    ext = raw_name.rsplit(".", 1)[-1] if "." in raw_name else ""
    if ext not in ("png", "jpg", "jpeg", "gif", "webp", "bmp"):
        raise HTTPException(
            status_code=400,
            detail="Supported: PNG, JPG, JPEG, GIF, WebP, BMP",
        )
    content = await file.read()
    if len(content) > 12_000_000:
        raise HTTPException(status_code=400, detail="File too large (max 12 MB before compression)")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        jpeg_bytes, mime = _compress_weighment_slip_image_bytes(content)
    except Exception as e:
        logging.getLogger(__name__).warning("weighment slip compress failed: %s", e)
        raise HTTPException(status_code=400, detail="Could not read or compress image")

    fname = f"weighment_slip_{uuid.uuid4().hex}.jpg"
    dest = UPLOADS_DIR / fname
    with open(dest, "wb") as out:
        out.write(jpeg_bytes)

    public_url = f"/uploads/{fname}"
    size_kb = round(len(jpeg_bytes) / 1024.0, 2)
    await create_audit_log(
        current_user.id,
        "UPLOAD_WEIGHMENT_SLIP",
        "procurement",
        {"file_name": fname, "size_kb": size_kb},
    )
    return {
        "file_url": public_url,
        "mime_type": mime,
        "size_kb": size_kb,
        "message": "Weighment slip saved (compressed JPEG)",
    }


@api_router.get("/purchase-invoices/{invoice_id}")
async def get_purchase_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get single purchase invoice with line items"""
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    normalize_invoice_advance_balance(invoice)
    # Get line items
    lines = await db.purchase_invoice_lines.find(
        {"invoice_id": invoice_id},
        {"_id": 0}
    ).sort("line_no", 1).to_list(100)
    
    invoice['line_items'] = lines
    return invoice

@api_router.put("/purchase-invoices/{invoice_id}")
async def update_purchase_invoice(
    invoice_id: str,
    invoice_data: PurchaseInvoiceCreate,
    current_user: User = Depends(get_current_user)
):
    """Update purchase invoice (admin can edit any status; others only draft)"""
    existing = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")

    is_admin = str(getattr(current_user, "role", "")).lower() == "admin"
    if existing.get('status') != 'draft' and not is_admin:
        raise HTTPException(status_code=400, detail="Can only edit draft invoices (admin can edit any invoice)")
    
    # Recalculate totals
    totals = calculate_invoice_totals(invoice_data.line_items, invoice_data.tds_rate_pct)
    
    # Update invoice
    update_data = {
        "invoice_date": invoice_data.invoice_date.isoformat(),
        "farmer_name": invoice_data.farmer_name,
        "farmer_mobile": invoice_data.farmer_mobile,
        "farmer_location": invoice_data.farmer_location,
        "agent_ref_name": invoice_data.agent_ref_name,
        "weighment_slip_no": invoice_data.weighment_slip_no,
        "weighment_slip_file_url": invoice_data.weighment_slip_file_url,
        "weighment_slip_mime_type": invoice_data.weighment_slip_mime_type,
        "custom_field_1_label": invoice_data.custom_field_1_label,
        "custom_field_1_value": invoice_data.custom_field_1_value,
        "custom_field_2_label": invoice_data.custom_field_2_label,
        "custom_field_2_value": invoice_data.custom_field_2_value,
        "tds_rate_pct": invoice_data.tds_rate_pct,
        "advance_paid": invoice_data.advance_paid,
        "notes": invoice_data.notes,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        **totals
    }
    
    # Calculate balance due and payment status
    balance_due = totals['grand_total'] - invoice_data.advance_paid
    update_data['balance_due'] = balance_due
    
    if balance_due <= 0:
        update_data['payment_status'] = 'paid'
    elif invoice_data.advance_paid > 0:
        update_data['payment_status'] = 'partial'
    else:
        update_data['payment_status'] = 'pending'

    old_slip_url = existing.get("weighment_slip_file_url")
    new_slip_url = invoice_data.weighment_slip_file_url
    if old_slip_url and old_slip_url != new_slip_url:
        _safe_delete_weighment_upload_file(old_slip_url)
    
    await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": update_data})

    # If this invoice already created a procurement lot, keep lot header amounts aligned.
    if existing.get("status") == "pushed" and existing.get("lot_id"):
        total_qty = float(totals.get("total_quantity_kg") or 0)
        subtotal_amt = float(totals.get("subtotal") or 0)
        avg_rate = round(subtotal_amt / total_qty, 2) if total_qty > 0 else 0
        await db.procurement_lots.update_one(
            {"id": existing.get("lot_id")},
            {"$set": {
                "gross_weight_kg": total_qty,
                "net_weight_kg": total_qty,
                "no_of_tons": round(total_qty / 1000, 3),
                "rate_per_kg": avg_rate,
                "total_amount": float(totals.get("grand_total") or 0),
                "advance_paid": float(invoice_data.advance_paid or 0),
                "balance_due": float(balance_due or 0),
                "payment_status": update_data.get("payment_status"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
    
    # Delete old line items and insert new ones
    await db.purchase_invoice_lines.delete_many({"invoice_id": invoice_id})
    
    for line_data in invoice_data.line_items:
        line = PurchaseInvoiceLine(
            invoice_id=invoice_id,
            line_no=line_data.line_no,
            variety=line_data.variety,
            count_value=line_data.count_value,
            custom_variety_notes=line_data.custom_variety_notes,
            custom_count_notes=line_data.custom_count_notes,
            quantity_kg=line_data.quantity_kg,
            rate=line_data.rate,
            amount=round(line_data.quantity_kg * line_data.rate, 2)
        )
        await db.purchase_invoice_lines.insert_one(line.model_dump())
    
    # Keep party ledger bill/payment entry in sync with current invoice values when linked.
    refreshed = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if refreshed and refreshed.get("party_id"):
        try:
            await create_ledger_entry_for_invoice(refreshed, current_user.id)
        except Exception as e:
            print(f"Warning: Could not sync ledger after invoice update: {e}")

    await create_audit_log(current_user.id, "UPDATE_PURCHASE_INVOICE", "procurement",
                          {"invoice_id": invoice_id, "status": existing.get("status"), "admin_override": bool(is_admin and existing.get("status") != "draft")})
    
    return {"status": "success", "message": "Invoice updated"}

@api_router.post("/purchase-invoices/{invoice_id}/approve")
async def approve_purchase_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Approve purchase invoice (locks it for editing)"""
    # Check role permission
    if current_user.role not in ['admin', 'owner', 'procurement_manager']:
        raise HTTPException(status_code=403, detail="Not authorized to approve invoices")
    
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Can only approve draft invoices")
    
    # Update status
    await db.purchase_invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user.id,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    # Create party ledger entry as soon as invoice is approved (with invoice details)
    if invoice.get("party_id"):
        try:
            await create_ledger_entry_for_invoice(invoice, current_user.id)
        except Exception as e:
            print(f"Warning: Could not create party ledger entry on approve: {e}")
    
    await create_audit_log(current_user.id, "APPROVE_PURCHASE_INVOICE", "procurement",
                          {"invoice_id": invoice_id, "invoice_no": invoice.get('invoice_no')})
    
    return {"status": "success", "message": "Invoice approved and locked"}

@api_router.post("/purchase-invoices/{invoice_id}/push-to-procurement")
async def push_invoice_to_procurement(
    invoice_id: str,
    payload: Optional[PushInvoiceRequest] = None,
    current_user: User = Depends(get_current_user)
):
    """Push approved invoice to procurement (creates procurement lot)"""
    # Check role permission
    if current_user.role not in ['admin', 'owner', 'procurement_manager']:
        raise HTTPException(status_code=403, detail="Not authorized to push invoices")
    
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get('status') != 'approved':
        raise HTTPException(status_code=400, detail="Can only push approved invoices")
    
    if invoice.get('lot_id'):
        raise HTTPException(status_code=400, detail="Invoice already pushed")

    apply_digital_signature = bool(payload.apply_digital_signature) if payload else False
    if apply_digital_signature and current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Only admin can apply digital signature")
    
    # Get line items to determine species
    lines = await db.purchase_invoice_lines.find(
        {"invoice_id": invoice_id},
        {"_id": 0}
    ).sort("line_no", 1).to_list(100)
    
    # Parse species from first line variety (default to vannamei)
    first_variety = lines[0]['variety'].lower() if lines else "vannamei"
    species = "Vannamei"
    if "black" in first_variety or "tiger" in first_variety:
        species = "Black Tiger"
    
    # Calculate weighted average rate
    total_amount = invoice.get('subtotal', 0)
    total_qty = invoice.get('total_quantity_kg', 0)
    avg_rate = round(total_amount / total_qty, 2) if total_qty > 0 else 0
    
    # Lookup or create agent
    agent_id = None
    agent_name = invoice.get('agent_ref_name', 'Unknown')
    if agent_name and agent_name != 'Unknown':
        existing_agent = await db.agents.find_one({"name": agent_name}, {"_id": 0})
        if existing_agent:
            agent_id = existing_agent['id']
        else:
            # Create pending agent
            new_agent = {
                "id": str(uuid.uuid4()),
                "name": agent_name,
                "phone": "",
                "commission_pct": 0.0,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.agents.insert_one(new_agent)
            agent_id = new_agent['id']
    
    # Generate lot number
    today = datetime.now(timezone.utc).date()
    lot_prefix = f"PRW-{today.strftime('%Y-%m-%d')}"
    existing_lots_today = await db.procurement_lots.count_documents({
        "lot_number": {"$regex": f"^{lot_prefix}"}
    })
    lot_number = f"{lot_prefix}-{existing_lots_today + 1:03d}"
    
    # Create procurement lot
    lot = {
        "id": str(uuid.uuid4()),
        "lot_number": lot_number,
        "agent_id": agent_id or "",
        "agent_name": agent_name,
        "vehicle_number": "",
        "driver_name": "",
        "arrival_time": invoice['invoice_date'],
        "species": species,
        "count_per_kg": lines[0]['count_value'] if lines else "",
        "boxes_count": 0,
        "gross_weight_kg": total_qty,
        "ice_weight_kg": 0,
        "net_weight_kg": total_qty,
        "no_of_tons": round(total_qty / 1000, 3),
        "no_of_trays": 0,
        "rate_per_kg": avg_rate,
        "total_amount": invoice.get('grand_total', 0),
        "advance_paid": invoice.get('advance_paid', 0),
        "balance_due": invoice.get('balance_due', 0),
        "ice_ratio_pct": 0,
        "freshness_grade": "A",
        "is_rejected": False,
        "payment_status": invoice.get('payment_status', 'pending'),
        "payments": [],
        "photos": [],
        "is_update_pending_approval": False,
        "approval_status": "approved",
        "notes": f"Created from Purchase Invoice {invoice['invoice_no']}",
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "purchase_invoice_id": invoice_id,
        "purchase_invoice_no": invoice['invoice_no']
    }
    
    await db.procurement_lots.insert_one(lot)
    
    # Update invoice
    await db.purchase_invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "pushed",
            "lot_id": lot['id'],
            "pushed_at": datetime.now(timezone.utc).isoformat(),
            "pushed_by": current_user.id,
            "apply_digital_signature": apply_digital_signature,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    # Create digital signature metadata only when explicitly requested by admin.
    if apply_digital_signature:
        # Default to SECRET_KEY if INVOICE_SIGNING_SECRET is not set (dev convenience).
        # For production, set INVOICE_SIGNING_SECRET to a dedicated secret.
        invoice_signing_secret = os.environ.get("INVOICE_SIGNING_SECRET") or SECRET_KEY
        if invoice_signing_secret:
            pushed_invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
            if pushed_invoice:
                payload_bytes = _canonical_purchase_invoice_payload(pushed_invoice, lines)
                sig_meta = _hmac_sign_purchase_invoice_payload(payload_bytes, invoice_signing_secret)
                await db.purchase_invoices.update_one(
                    {"id": invoice_id},
                    {"$set": {
                        "digital_signature_algo": sig_meta.get("signature_algo"),
                        "digital_signature_payload_hash_sha256": sig_meta.get("payload_hash_sha256"),
                        "digital_signature_value_b64": sig_meta.get("signature_value_b64"),
                        "digital_signature_signed_by": current_user.id,
                        "digital_signature_signed_at": datetime.now(timezone.utc).isoformat(),
                    }}
                )
    else:
        # Explicitly clear signature fields so the invoice sign area remains empty.
        await db.purchase_invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "digital_signature_algo": None,
                "digital_signature_payload_hash_sha256": None,
                "digital_signature_value_b64": None,
                "digital_signature_signed_by": None,
                "digital_signature_signed_at": None,
            }}
        )
    
    # A5: Create ledger entry if party is linked
    if invoice.get("party_id"):
        try:
            await create_ledger_entry_for_invoice(invoice, current_user.id)
        except Exception as e:
            # Log but don't fail the push
            print(f"Warning: Could not create ledger entry: {e}")
    
    await create_audit_log(current_user.id, "PUSH_PURCHASE_INVOICE", "procurement",
                          {"invoice_id": invoice_id, "lot_id": lot['id'], "lot_number": lot_number})
    
    return {
        "status": "success",
        "message": "Invoice pushed to procurement",
        "lot_id": lot['id'],
        "lot_number": lot_number
    }

@api_router.post("/purchase-invoices/{invoice_id}/sync-ledger")
async def sync_invoice_ledger(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Ensure party ledger reflects this invoice (bill + advance payment). Idempotent; backfills advance if missing."""
    if current_user.role not in ['admin', 'owner', 'procurement_manager']:
        raise HTTPException(status_code=403, detail="Not authorized")
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if not invoice.get("party_id"):
        raise HTTPException(status_code=400, detail="Invoice has no party linked")
    try:
        await create_ledger_entry_for_invoice(invoice, current_user.id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ledger sync failed: {str(e)}")
    return {"status": "success", "message": "Ledger synced for invoice"}

@api_router.delete("/purchase-invoices/{invoice_id}")
async def delete_purchase_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete draft purchase invoice"""
    # Check role permission
    if current_user.role not in ['admin', 'procurement_manager']:
        raise HTTPException(status_code=403, detail="Not authorized to delete invoices")
    
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Can only delete draft invoices")
    
    _safe_delete_weighment_upload_file(invoice.get("weighment_slip_file_url"))
    
    # Delete line items
    await db.purchase_invoice_lines.delete_many({"invoice_id": invoice_id})
    
    # Delete invoice
    await db.purchase_invoices.delete_one({"id": invoice_id})
    
    await create_audit_log(current_user.id, "DELETE_PURCHASE_INVOICE", "procurement",
                          {"invoice_id": invoice_id, "invoice_no": invoice.get('invoice_no')})
    
    return {"status": "success", "message": "Invoice deleted"}

@api_router.patch("/purchase-invoices/{invoice_id}/manual-audit")
async def toggle_manual_audit(
    invoice_id: str,
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """Toggle manual audit recording status (A4 PATCH 10G)"""
    # Check role permission
    if current_user.role not in ['admin', 'owner', 'procurement_manager']:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    is_recorded = data.get('is_manually_recorded', False)
    
    update_data = {
        "is_manually_recorded": is_recorded,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if is_recorded:
        update_data["manually_recorded_at"] = datetime.now(timezone.utc).isoformat()
        update_data["manually_recorded_by"] = current_user.id
    else:
        update_data["manually_recorded_at"] = None
        update_data["manually_recorded_by"] = None
    
    await db.purchase_invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    await create_audit_log(current_user.id, "TOGGLE_MANUAL_AUDIT", "procurement",
                          {"invoice_id": invoice_id, "is_recorded": is_recorded})
    
    return {"status": "success", "message": f"Manual audit status updated"}

# Digitally sign purchase invoices on "push to procurement".
# We sign an HMAC over a canonical JSON payload (invoice + line items),
# and then embed the signature metadata into the PDF.
def _canonical_purchase_invoice_payload(invoice: dict, lines: List[dict]) -> bytes:
    payload = {"invoice": invoice, "lines": lines}
    payload_json = json.dumps(payload, sort_keys=True, default=str, separators=(",", ":"))
    return payload_json.encode("utf-8")


def _hmac_sign_purchase_invoice_payload(payload_bytes: bytes, secret: str) -> dict:
    """
    Returns signature metadata.
    Uses HMAC-SHA256 so we only need a shared secret (no certificate/keypair).
    """
    signature_algo = "HMAC-SHA256"
    payload_hash_sha256 = hashlib.sha256(payload_bytes).hexdigest()
    sig_bytes = hmac.new(secret.encode("utf-8"), payload_bytes, hashlib.sha256).digest()
    sig_b64 = base64.b64encode(sig_bytes).decode("ascii")
    return {
        "signature_algo": signature_algo,
        "payload_hash_sha256": payload_hash_sha256,
        "signature_value_b64": sig_b64,
    }


def _ende_canonical_signing_date(iso_dt: Optional[str]) -> str:
    """
    Formats datetime for endesive's signingdate field.
    endesive expects: YYYYmmddHHMMSS+00'00'
    """
    if not iso_dt:
        iso_dt = datetime.now(timezone.utc).isoformat()
    try:
        dt = datetime.fromisoformat(iso_dt.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(timezone.utc)
    except Exception:
        dt = datetime.now(timezone.utc)
    return dt.strftime("%Y%m%d%H%M%S") + "+00'00'"


def _parse_signature_box_env(value: Optional[str]) -> tuple:
    """
    Parses INVOICE_PDF_SIGNATURE_BOX like "350,50,550,150".
    Coordinates are page-user-space units (points).
    """
    # Keep default visible signature appearance inside footer right area.
    # Format: (x1, y1, x2, y2) in PDF points.
    default_box = (440, 66, 555, 88)
    if not value:
        return default_box
    try:
        parts = [float(x.strip()) for x in value.split(",") if x.strip()]
        if len(parts) >= 4:
            return tuple(parts[:4])
    except Exception:
        pass
    return default_box


def _maybe_sign_purchase_invoice_pdf(pdf_bytes: bytes, invoice: dict) -> bytes:
    """
    If certificate env vars are configured, apply a real cryptographic PDF signature
    (CMS/PKCS#7) using the provided PFX (PKCS#12) file.
    Otherwise returns the original PDF bytes unchanged.
    """
    pfx_path = os.environ.get("INVOICE_PDF_SIGNING_PFX_PATH")
    if not pfx_path:
        return pdf_bytes

    # Only apply cryptographic signature for pushed invoices (admin action).
    if invoice.get("status") != "pushed":
        return pdf_bytes
    if not invoice.get("apply_digital_signature"):
        return pdf_bytes

    pfx_password = os.environ.get("INVOICE_PDF_SIGNING_PFX_PASSWORD", "")
    sig_box = _parse_signature_box_env(os.environ.get("INVOICE_PDF_SIGNATURE_BOX"))

    try:
        from cryptography.hazmat import backends
        from cryptography.hazmat.primitives.serialization import pkcs12
        from endesive import pdf as endesive_pdf
    except Exception as e:
        # If endesive isn't installed in environment, keep PDF usable.
        print(f"PDF signing disabled (missing library): {e}")
        return pdf_bytes

    try:
        with open(pfx_path, "rb") as f:
            pfx_data = f.read()
        p12pk, p12pc, p12oc = pkcs12.load_key_and_certificates(
            pfx_data, (pfx_password or "").encode("utf-8"), backends.default_backend()
        )
        if not p12pk or not p12pc:
            print("PDF signing disabled: PFX did not contain key/certificate")
            return pdf_bytes
    except Exception as e:
        print(f"PDF signing disabled (PFX load failed): {e}")
        return pdf_bytes

    # Extract CN/C (if available) for signature metadata
    def _get_rdn_names(x509_name):
        # Best-effort extraction of DN components used by endesive example.
        names = {"CN": "", "C": ""}
        try:
            for rdn in x509_name.rdns:
                for attr in rdn:
                    if getattr(attr.oid, "dotted_string", "") == "2.5.4.3":  # commonName
                        names["CN"] = str(attr.value)
                    if getattr(attr.oid, "dotted_string", "") == "2.5.4.6":  # countryName
                        names["C"] = str(attr.value)
        except Exception:
            pass
        return names

    names = _get_rdn_names(p12pc.subject)
    signing_date = _ende_canonical_signing_date(invoice.get("digital_signature_signed_at"))

    reason = os.environ.get("INVOICE_PDF_SIGNING_REASON", "Signed by Prawn ERP")
    contact = os.environ.get("INVOICE_PDF_SIGNING_CONTACT", "")
    location = os.environ.get("INVOICE_PDF_SIGNING_LOCATION", "")
    if not contact:
        contact = "noreply@example.com"
    # Keep visual signature text compact so it stays inside the box.
    signature = "Digitally Signed"

    dct = {
        "sigflags": 3,
        "sigpage": 0,
        "contact": contact,
        "location": location or "—",
        "signingdate": signing_date,
        "reason": reason,
        "signature": signature,
        "signaturebox": sig_box,
    }

    # endesive signs by generating a signature block and appending it to the PDF.
    # Returning pdf_bytes + datas matches the official example's write(datau); write(datas).
    datas = endesive_pdf.cms.sign(pdf_bytes, dct, p12pk, p12pc, p12oc, "sha256")
    return pdf_bytes + datas


def _signature_path_from_config_value(raw: Optional[Any]) -> Optional[Path]:
    """
    Resolve tenant_config invoice_signature_image to a file under UPLOADS_DIR.
    Accepts '/uploads/name.ext' or a full URL containing '/uploads/...'.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Normalize separators so we accept values stored with backslashes too.
    s = s.replace("\\", "/")
    # Remove query/fragment early so basename is clean.
    s = s.split("?")[0].split("#")[0]

    # Accept any string that contains "/uploads/" (including full URLs),
    # or a value stored as "uploads/..." (no leading slash).
    low = s.lower()
    uploads_idx = low.find("/uploads/")
    if uploads_idx != -1:
        s = s[uploads_idx:]  # now starts with "/uploads/..."
    elif low.startswith("uploads/"):
        s = "/" + s  # "uploads/..." -> "/uploads/..."
    else:
        return None

    base_name = os.path.basename(s)
    if not base_name or base_name in (".", "..") or ".." in s:
        return None
    try:
        uploads_root = UPLOADS_DIR.resolve()
        full = (UPLOADS_DIR / base_name).resolve()
        full.relative_to(uploads_root)
    except (OSError, ValueError):
        return None
    if not full.is_file():
        return None
    return full


def generate_purchase_invoice_pdf(
    invoice: dict,
    lines: List[dict],
    tenant_config: dict,
    signature: Optional[dict] = None,
    pdf_generated_at: Optional[datetime] = None,
) -> bytes:
    """Generate PDF matching exact format from reference images"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.3*inch, bottomMargin=0.3*inch,
                          leftMargin=0.3*inch, rightMargin=0.3*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Page width for full-width tables
    page_width = A4[0] - 0.6*inch
    
    # Custom styles
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=3
    )
    
    company_style = ParagraphStyle(
        'CompanyStyle',
        parent=styles['Normal'],
        fontSize=20,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=3
    )
    
    address_style = ParagraphStyle(
        'AddressStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_CENTER,
        spaceAfter=2
    )
    
    # Create outer table for cyan background
    inner_content = []
    
    # Header section
    inner_content.append(Paragraph("PURCHASE INVOICE", header_style))
    inner_content.append(Spacer(1, 0.05*inch))
    inner_content.append(Paragraph(tenant_config.get('company_name', 'COMPANY NAME'), company_style))
    inner_content.append(Spacer(1, 0.05*inch))
    inner_content.append(Paragraph(tenant_config.get('company_address_1', ''), address_style))
    inner_content.append(Paragraph(tenant_config.get('company_address_2', ''), address_style))
    inner_content.append(Paragraph(
        f"Contact Number: {tenant_config.get('company_phone', '')}, Email Id: {tenant_config.get('company_email', '')}",
        address_style
    ))
    inner_content.append(Spacer(1, 0.1*inch))
    
    # Meta information table
    meta_data = [
        ['Farmer Name: ' + str(invoice.get('farmer_name') or ''), 'DATE: ' + str(invoice.get('invoice_date') or '')],
        ['Location: ' + str(invoice.get('farmer_location') or ''), 'Purchase Invoice No: ' + str(invoice.get('invoice_no') or '')],
        ['Farmer/Agent Ref Name: ' + str(invoice.get('agent_ref_name') or ''), 'Weighment Slip No: ' + str(invoice.get('weighment_slip_no') or '')],
    ]
    
    meta_table = Table(meta_data, colWidths=[page_width/2, page_width/2])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), rl_colors.HexColor('#e0f7fa')),
        ('TEXTCOLOR', (0, 0), (-1, -1), rl_colors.HexColor('#0d47a1')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#0d47a1')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    inner_content.append(meta_table)
    inner_content.append(Spacer(1, 0.05*inch))
    
    # Line items table (pad to fixed rows for stable template layout)
    line_data = [['S.NO', 'Variety', 'Count', 'Quantity\nKgs/Gms', 'Rate', 'Amount']]
    
    for idx, line in enumerate(lines, 1):
        _q = _pdf_safe_float(line.get("quantity_kg"), 0.0)
        _r = _pdf_safe_float(line.get("rate"), 0.0)
        _a = _pdf_safe_float(line.get("amount"), 0.0)
        line_data.append([
            str(idx),
            str(line.get("variety") or ""),
            str(line.get("count_value") or ""),
            f"{_q:.3f}",
            f"{_r:.2f}",
            f"{_a:.2f}",
        ])
    
    # Pad to 12 rows to avoid crowding totals/footer area
    while len(line_data) < 13:  # 12 data rows + 1 header
        line_data.append(['', '', '', '', '', '0.00'])
    
    # Add subtotal row
    line_data.append(['', '', '', '', '', f"{_pdf_safe_float(invoice.get('subtotal'), 0):.2f}"])
    
    # Add TDS row
    tds_label = f"TDS@{_pdf_safe_float(invoice.get('tds_rate_pct'), 0.1)}%"
    line_data.append(['', '', '', '', tds_label, f"{_pdf_safe_float(invoice.get('tds_amount'), 0):.2f}"])
    
    # Add Rounded Off row
    line_data.append(['', '', '', '', 'Rounded Off', f"{_pdf_safe_float(invoice.get('rounded_off'), 0):.2f}"])
    
    # Add Grand Total row (label spans first 3 cols so it fits; qty col 3; amount col 5)
    _gt = _pdf_safe_float(invoice.get('grand_total'), 0)
    line_data.append([
        'Grand Total',
        '',
        '',
        f"{_pdf_safe_float(invoice.get('total_quantity_kg'), 0):.3f}",
        '',
        f"{_gt:.2f}",
    ])
    # Advance Paid and Balance Due (so generated invoice shows them)
    advance_paid = _pdf_safe_float(invoice.get('advance_paid'), 0)
    balance_due = _pdf_safe_float(invoice.get('balance_due'), 0)
    if balance_due == 0 and advance_paid > 0:
        balance_due = _gt - advance_paid
    line_data.append(['Advance Paid', '', '', '', '', f"{advance_paid:.2f}"])
    line_data.append(['Balance Due', '', '', '', '', f"{balance_due:.2f}"])

    # Column widths fill printable width; col 0 stays narrow for S.No. — wide labels use SPAN below
    _lcw0 = 0.48 * inch
    _lcw1 = 1.42 * inch
    _lcw2 = 0.92 * inch
    _lcw3 = 1.05 * inch
    _lcw4 = 0.82 * inch
    _lcw5 = page_width - (_lcw0 + _lcw1 + _lcw2 + _lcw3 + _lcw4)
    if _lcw5 < 0.85 * inch:
        _lcw5 = 0.85 * inch
    line_col_widths = [_lcw0, _lcw1, _lcw2, _lcw3, _lcw4, _lcw5]

    line_table = Table(line_data, colWidths=line_col_widths)
    line_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#e0f7fa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.HexColor('#0d47a1')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body lines + padded rows: center; numeric amount column right-align on col 5 only
        ('BACKGROUND', (0, 1), (-1, -7), rl_colors.HexColor('#e0f7fa')),
        ('TEXTCOLOR', (0, 1), (-1, -7), rl_colors.HexColor('#0d47a1')),
        ('FONTNAME', (0, 1), (-1, -7), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -7), 9),
        ('ALIGN', (0, 1), (-1, -7), 'CENTER'),
        ('ALIGN', (5, 1), (5, -7), 'RIGHT'),
        # Subtotal row (-6)
        ('ALIGN', (5, -6), (5, -6), 'RIGHT'),
        ('FONTNAME', (5, -6), (5, -6), 'Helvetica-Bold'),
        # TDS row (-5)
        ('TEXTCOLOR', (4, -5), (4, -5), rl_colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (5, -5), (5, -5), rl_colors.HexColor('#d32f2f')),
        ('FONTNAME', (4, -5), (5, -5), 'Helvetica-Bold'),
        ('ALIGN', (4, -5), (4, -5), 'RIGHT'),
        ('ALIGN', (5, -5), (5, -5), 'RIGHT'),
        # Rounded Off row (-4)
        ('TEXTCOLOR', (4, -4), (4, -4), rl_colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (5, -4), (5, -4), rl_colors.HexColor('#d32f2f')),
        ('FONTNAME', (4, -4), (5, -4), 'Helvetica-Bold'),
        ('ALIGN', (4, -4), (4, -4), 'RIGHT'),
        ('ALIGN', (5, -4), (5, -4), 'RIGHT'),
        # Grand Total (-3): merge label across cols 0–2 so text stays inside grid
        ('SPAN', (0, -3), (2, -3)),
        ('BACKGROUND', (0, -3), (-1, -3), rl_colors.HexColor('#e0f7fa')),
        ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -3), (-1, -3), 9.5),
        ('ALIGN', (0, -3), (2, -3), 'LEFT'),
        ('LEFTPADDING', (0, -3), (2, -3), 4),
        ('ALIGN', (3, -3), (3, -3), 'RIGHT'),
        ('ALIGN', (5, -3), (5, -3), 'RIGHT'),
        # Advance Paid (-2): label spans cols 0–4
        ('SPAN', (0, -2), (4, -2)),
        ('FONTNAME', (0, -2), (-1, -2), 'Helvetica'),
        ('ALIGN', (0, -2), (4, -2), 'LEFT'),
        ('LEFTPADDING', (0, -2), (4, -2), 4),
        ('ALIGN', (5, -2), (5, -2), 'RIGHT'),
        # Balance Due (-1)
        ('SPAN', (0, -1), (4, -1)),
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#e8f5e9')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (0, -1), (4, -1), 'LEFT'),
        ('LEFTPADDING', (0, -1), (4, -1), 4),
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#0d47a1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, -3), (-1, -1), 4),
        ('BOTTOMPADDING', (0, -3), (-1, -1), 4),
    ]))
    inner_content.append(line_table)
    inner_content.append(Spacer(1, 0.1*inch))
    
    # Bottom section (wrapped Paragraphs so long "amount in words" stays inside the box)
    amount_words = amount_to_words_indian(invoice.get("grand_total"))
    _co_raw = tenant_config.get('company_name') or 'COMPANY NAME'
    _co_name = _co_raw.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    _words_safe = str(amount_words).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    words_in_cell_style = ParagraphStyle(
        'AmountWords',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_LEFT,
        spaceBefore=0,
        spaceAfter=2,
    )
    for_company_style = ParagraphStyle(
        'ForCompany',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_RIGHT,
    )
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=rl_colors.HexColor('#1565c0'),
        alignment=TA_RIGHT,
    )
    auth_sig_style = ParagraphStyle(
        'AuthSignatory',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=10,
        textColor=rl_colors.HexColor('#0d47a1'),
        alignment=TA_RIGHT,
    )

    _bottom_left_w = page_width * 0.58
    _bottom_right_w = page_width * 0.42
    words_para = Paragraph(
        f"<b>Total Amount In Words:</b> {_words_safe}",
        words_in_cell_style,
    )
    for_para = Paragraph(f"For {_co_name}", for_company_style)

    _gen_at = pdf_generated_at or datetime.now(timezone.utc)
    printed_ts_style = ParagraphStyle(
        'PrintedTs',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=rl_colors.HexColor('#455a64'),
        alignment=TA_RIGHT,
    )
    printed_ts_para = Paragraph(
        f"Printed: {_gen_at.strftime('%d-%m-%Y %H:%M:%S')} UTC",
        printed_ts_style,
    )

    # Optional scanned / uploaded owner signature image (Company Settings).
    # For pushed invoices, only embed this stamp when the user opted in to digitally sign on push;
    # otherwise the footer should stay visually empty of signature artwork (HMAC/PDF-CMS are already off).
    sig_image_holder = None
    _sig_raw = tenant_config.get("invoice_signature_image")
    _sig_path = _signature_path_from_config_value(_sig_raw)
    if _sig_raw and _sig_path is None:
        logging.getLogger(__name__).warning(
            "Purchase invoice PDF: invoice_signature_image is set (%r) but no matching file under %s. "
            "Re-upload from Company Settings or ensure backend/uploads is persisted (Docker volume).",
            _sig_raw,
            UPLOADS_DIR,
        )
    _embed_company_signature_stamp = bool(invoice.get("apply_digital_signature"))
    if _sig_path is not None and _embed_company_signature_stamp:
        try:
            with open(_sig_path, "rb") as _sig_f:
                _sig_bytes = _sig_f.read()
            rl_sig = RLImage(io.BytesIO(_sig_bytes))
            _max_w = 1.45 * inch
            if rl_sig.drawWidth > _max_w:
                _sc = _max_w / rl_sig.drawWidth
                rl_sig.drawWidth = _max_w
                rl_sig.drawHeight = rl_sig.drawHeight * _sc
            _sig_inner_w2 = max(float(_bottom_right_w) - 24, 120)
            sig_image_holder = Table(
                [[rl_sig], [printed_ts_para]],
                colWidths=[_sig_inner_w2],
            )
            sig_image_holder.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
                # Keep this compact so the rest of the footer doesn't get clipped
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
        except Exception as _sig_err:
            # Some image formats can fail ReportLab decoding even if Pillow supports them.
            # Try converting to PNG in-memory as a robust fallback.
            try:
                from PIL import Image as PILImage

                pil_img = PILImage.open(_sig_path)
                bio = io.BytesIO()
                pil_img.save(bio, format="PNG")
                bio.seek(0)
                rl_sig = RLImage(bio)

                _max_w = 1.45 * inch
                if rl_sig.drawWidth > _max_w:
                    _sc = _max_w / rl_sig.drawWidth
                    rl_sig.drawWidth = _max_w
                    rl_sig.drawHeight = rl_sig.drawHeight * _sc
                _sig_inner_w2 = max(float(_bottom_right_w) - 24, 120)
                sig_image_holder = Table(
                    [[rl_sig], [printed_ts_para]],
                    colWidths=[_sig_inner_w2],
                )
                sig_image_holder.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ]))
            except Exception as _sig_pil_err:
                logging.getLogger(__name__).warning(
                    "Purchase invoice PDF: could not embed signature image from %s. "
                    "ReportLab error=%s, Pillow fallback error=%s",
                    _sig_path,
                    _sig_err,
                    _sig_pil_err,
                )
                sig_image_holder = None

    # Visual signature block (always): line + "Authorized Signatory". When pushed with HMAC, add digital note.
    _sig_inner_w = max(float(_bottom_right_w) - 24, 120)
    sig_line_row = Table(
        [['', '']],
        colWidths=[_sig_inner_w * 0.36, _sig_inner_w * 0.64],
    )
    sig_line_row.setStyle(TableStyle([
        ('LINEBELOW', (1, 0), (1, 0), 0.75, rl_colors.HexColor('#0d47a1')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (1, 0), (1, 0), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    digital_lines = ""
    if signature and signature.get("signature_value_b64"):
        digital_lines = "Digitally Signed"
        _sat = signature.get("signed_at")
        if _sat:
            try:
                _iso = str(_sat).replace("Z", "+00:00")
                _dt = datetime.fromisoformat(_iso)
                if _dt.tzinfo is None:
                    _dt = _dt.replace(tzinfo=timezone.utc)
                digital_lines += f"<br/><font size=\"8\" color=\"#455a64\">{_dt.strftime('%d-%m-%Y %H:%M')} UTC</font>"
            except Exception:
                pass

    show_signature_block = bool(invoice.get("apply_digital_signature"))
    right_col_flowables = [for_para]
    if show_signature_block:
        right_col_flowables.extend([sig_line_row, Spacer(1, 0.04 * inch)])
        if sig_image_holder is not None:
            right_col_flowables.append(sig_image_holder)
            right_col_flowables.append(Spacer(1, 0.02 * inch))
        right_col_flowables.append(Paragraph("Authorized Signatory", auth_sig_style))
        if digital_lines:
            # Smaller spacer to avoid pushing scanned signature out of page when clipping occurs.
            right_col_flowables.extend([Spacer(1, 0.02 * inch), Paragraph(digital_lines, signature_style)])

    bottom_data = [[words_para, right_col_flowables]]

    bottom_table = Table(
        bottom_data,
        colWidths=[_bottom_left_w, _bottom_right_w],
    )
    # If the "Digitally Signed" block is present, tighten outer padding to keep everything visible.
    _outer_pad = 4 if digital_lines else 6
    bottom_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), rl_colors.HexColor('#e0f7fa')),
        ('TEXTCOLOR', (0, 0), (-1, -1), rl_colors.HexColor('#0d47a1')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#0d47a1')),
        ('TOPPADDING', (0, 0), (-1, -1), _outer_pad),
        ('BOTTOMPADDING', (0, 0), (-1, -1), _outer_pad),
        ('LEFTPADDING', (0, 0), (0, 0), 6),
        ('RIGHTPADDING', (0, 0), (0, 0), 6),
        ('LEFTPADDING', (1, 0), (1, 0), 6),
        ('RIGHTPADDING', (1, 0), (1, 0), 6),
    ]))
    inner_content.append(bottom_table)
    
    # Build PDF
    for elem in inner_content:
        elements.append(elem)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

@api_router.get("/purchase-invoices/{invoice_id}/pdf")
async def download_purchase_invoice_pdf(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download purchase invoice as PDF"""
    _log = logging.getLogger(__name__)
    try:
        invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        normalize_invoice_advance_balance(invoice)
        lines = await db.purchase_invoice_lines.find(
            {"invoice_id": invoice_id},
            {"_id": 0}
        ).sort("line_no", 1).to_list(100)

        tenant_config = await get_tenant_config_dict(use_cache=False)
        signature_context = None
        if invoice.get("status") == "pushed" and invoice.get("digital_signature_value_b64"):
            signature_context = {
                "signature_value_b64": invoice.get("digital_signature_value_b64"),
                "signed_by": invoice.get("digital_signature_signed_by"),
                "signed_at": invoice.get("digital_signature_signed_at"),
            }
        try:
            pdf_bytes = generate_purchase_invoice_pdf(
                invoice,
                lines,
                tenant_config,
                signature_context,
                pdf_generated_at=datetime.now(timezone.utc),
            )
            pdf_bytes = _maybe_sign_purchase_invoice_pdf(pdf_bytes, invoice)
        except Exception as e:
            _log.exception("purchase invoice PDF generation failed for %s", invoice_id)
            raise HTTPException(
                status_code=500,
                detail=f"Could not generate PDF: {e!s}",
            ) from e

        from fastapi.responses import Response
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename=purchase_invoice_{str(invoice.get("invoice_no") or invoice_id).replace("/", "_")}.pdf'
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        _log.exception("purchase invoice PDF download failed for %s", invoice_id)
        raise HTTPException(
            status_code=500,
            detail=f"Could not download invoice PDF: {e!s}",
        ) from e

@api_router.get("/tenant-config")
async def get_tenant_config(current_user: User = Depends(get_current_user)):
    """Get all tenant configuration (cached 60s)."""
    d = await get_tenant_config_dict()
    return [{"key": k, "value": v} for k, v in d.items()]

@api_router.post("/tenant-config")
async def update_tenant_config(
    config_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update tenant configuration (admin/owner only)"""
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    key = config_data.get('key')
    value = config_data.get('value')
    
    if not key:
        raise HTTPException(status_code=400, detail="Key is required")
    
    # Update or insert
    await db.tenant_config.update_one(
        {"key": key},
        {"$set": {"key": key, "value": value}},
        upsert=True
    )
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"tenant_config_dict:{tid}", None)
    except Exception:
        pass
    await create_audit_log(current_user.id, "UPDATE_TENANT_CONFIG", "admin",
                          {"key": key})
    return {"status": "success", "message": "Configuration updated"}


@api_router.post("/tenant-config/signature-image")
async def upload_invoice_signature_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload PNG/JPEG/WebP/GIF used on purchase invoice PDFs (admin or owner only)."""
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Not authorized")
    raw_name = file.filename or ""
    ext = raw_name.rsplit(".", 1)[-1].lower() if "." in raw_name else ""
    if ext not in ("png", "jpg", "jpeg", "gif", "webp"):
        raise HTTPException(status_code=400, detail="Use PNG, JPG, JPEG, GIF, or WebP")
    content = await file.read()
    if len(content) > 2_000_000:
        raise HTTPException(status_code=400, detail="Image too large (max 2 MB)")
    fname = f"invoice_signature_{uuid.uuid4().hex[:16]}.{ext}"
    dest = UPLOADS_DIR / fname
    with open(dest, "wb") as out:
        out.write(content)
    public_url = f"/uploads/{fname}"
    await db.tenant_config.update_one(
        {"key": "invoice_signature_image"},
        {"$set": {"key": "invoice_signature_image", "value": public_url}},
        upsert=True,
    )
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"tenant_config_dict:{tid}", None)
    except Exception:
        pass
    await create_audit_log(
        current_user.id,
        "UPDATE_TENANT_CONFIG",
        "admin",
        {"key": "invoice_signature_image", "action": "upload"},
    )
    return {"url": public_url, "message": "Signature image saved"}


@api_router.delete("/tenant-config/signature-image")
async def delete_invoice_signature_image(current_user: User = Depends(get_current_user)):
    """Remove uploaded invoice signature image from config and delete file from disk."""
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Not authorized")
    doc = await db.tenant_config.find_one({"key": "invoice_signature_image"}, {"_id": 0})
    old_val = (doc or {}).get("value") or ""
    old_path = _signature_path_from_config_value(old_val)
    if old_path and old_path.is_file():
        try:
            old_path.unlink()
        except OSError:
            pass
    await db.tenant_config.update_one(
        {"key": "invoice_signature_image"},
        {"$set": {"key": "invoice_signature_image", "value": ""}},
        upsert=True,
    )
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"tenant_config_dict:{tid}", None)
    except Exception:
        pass
    await create_audit_log(
        current_user.id,
        "UPDATE_TENANT_CONFIG",
        "admin",
        {"key": "invoice_signature_image", "action": "delete"},
    )
    return {"status": "success", "message": "Signature image removed"}


# ══════════════════════════════════════════════════════════════════════════════
# Comprehensive Reports - PDF & Excel Generation
# ══════════════════════════════════════════════════════════════════════════════

def generate_procurement_report_pdf(lots: List[dict], filters: dict) -> bytes:
    """Generate PDF report for procurement lots"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Procurement Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Filter info
    filter_text = f"Period: {filters.get('date_from', 'All')} to {filters.get('date_to', 'All')}"
    elements.append(Paragraph(filter_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Data table
    data = [['Lot #', 'Date', 'Supplier', 'Species', 'Quantity (kg)', 'Rate', 'Amount']]
    
    total_qty = 0
    total_amt = 0
    
    for lot in lots:
        data.append([
            lot.get('lot_number', 'N/A'),
            lot.get('purchase_date', '')[:10] if lot.get('purchase_date') else 'N/A',
            lot.get('supplier_name', 'N/A'),
            lot.get('species', 'N/A'),
            f"{lot.get('quantity', 0):.2f}",
            f"₹{lot.get('rate_per_kg', 0):.2f}",
            f"₹{lot.get('total_purchase_amount', 0):.2f}"
        ])
        total_qty += lot.get('quantity', 0)
        total_amt += lot.get('total_purchase_amount', 0)
    
    # Add totals row
    data.append(['', '', '', 'TOTAL', f"{total_qty:.2f}", '', f"₹{total_amt:.2f}"])
    
    table = Table(data, colWidths=[1*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.read()

def generate_production_report_pdf(orders: List[dict], filters: dict) -> bytes:
    """Generate PDF report for production orders"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Production Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Filter info
    filter_text = f"Period: {filters.get('date_from', 'All')} to {filters.get('date_to', 'All')}"
    elements.append(Paragraph(filter_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Data table
    data = [['Order #', 'Date', 'Type', 'Input (kg)', 'Output (kg)', 'Yield %', 'Status']]
    
    total_input = 0
    total_output = 0
    
    for order in orders:
        input_qty = order.get('input_quantity', 0)
        output_qty = order.get('output_quantity', 0)
        yield_pct = (output_qty / input_qty * 100) if input_qty > 0 else 0
        
        data.append([
            order.get('order_number', 'N/A'),
            order.get('start_date', '')[:10] if order.get('start_date') else 'N/A',
            order.get('process_type', 'N/A'),
            f"{input_qty:.2f}",
            f"{output_qty:.2f}",
            f"{yield_pct:.1f}%",
            order.get('status', 'N/A')
        ])
        total_input += input_qty
        total_output += output_qty
    
    # Add totals row
    avg_yield = (total_output / total_input * 100) if total_input > 0 else 0
    data.append(['', '', 'TOTAL', f"{total_input:.2f}", f"{total_output:.2f}", f"{avg_yield:.1f}%", ''])
    
    table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.read()

@api_router.get("/reports/procurement/pdf")
async def download_procurement_report_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    supplier: Optional[str] = None,
    species: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Generate and download procurement report as PDF"""
    # Build query
    query = {}
    if date_from:
        query['purchase_date'] = {'$gte': date_from}
    if date_to:
        if 'purchase_date' in query:
            query['purchase_date']['$lte'] = date_to
        else:
            query['purchase_date'] = {'$lte': date_to}
    if supplier:
        query['supplier_name'] = supplier
    if species:
        query['species'] = species
    
    # Fetch data
    lots = await db.procurement_lots.find(query, {"_id": 0}).sort("purchase_date", -1).to_list(1000)
    
    # Generate PDF
    filters = {
        'date_from': date_from or 'All',
        'date_to': date_to or 'All',
        'supplier': supplier or 'All',
        'species': species or 'All'
    }
    pdf_bytes = generate_procurement_report_pdf(lots, filters)
    
    # Return as downloadable file
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=procurement_report_{datetime.now().strftime("%Y%m%d")}.pdf'
        }
    )

@api_router.get("/reports/production/pdf")
async def download_production_report_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    process_type: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Generate and download production report as PDF"""
    # Build query
    query = {}
    if date_from:
        query['start_date'] = {'$gte': date_from}
    if date_to:
        if 'start_date' in query:
            query['start_date']['$lte'] = date_to
        else:
            query['start_date'] = {'$lte': date_to}
    if process_type:
        query['process_type'] = process_type
    if status:
        query['status'] = status
    
    # Fetch data
    orders = await db.production_orders.find(query, {"_id": 0}).sort("start_date", -1).to_list(1000)
    
    # Generate PDF
    filters = {
        'date_from': date_from or 'All',
        'date_to': date_to or 'All',
        'process_type': process_type or 'All',
        'status': status or 'All'
    }
    pdf_bytes = generate_production_report_pdf(orders, filters)
    
    # Return as downloadable file
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=production_report_{datetime.now().strftime("%Y%m%d")}.pdf'
        }
    )

@api_router.get("/reports/cold-storage/pdf")
async def download_cold_storage_report_pdf(
    chamber_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Generate and download cold storage inventory report as PDF"""
    # Build query
    query = {}
    if chamber_id:
        query['chamber_id'] = chamber_id
    
    # Fetch inventory data
    inventory = await db.cold_storage_inventory.find(query, {"_id": 0}).to_list(1000)
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Cold Storage Inventory Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Data table
    data = [['Item', 'Lot #', 'Chamber', 'Slot', 'Quantity (kg)', 'Entry Date', 'Status']]
    
    total_qty = 0
    
    for item in inventory:
        data.append([
            item.get('item_name', 'N/A'),
            item.get('lot_number', 'N/A'),
            item.get('chamber_name', 'N/A'),
            item.get('slot_number', 'N/A'),
            f"{item.get('quantity', 0):.2f}",
            item.get('entry_date', '')[:10] if item.get('entry_date') else 'N/A',
            item.get('status', 'N/A')
        ])
        total_qty += item.get('quantity', 0)
    
    # Add totals row
    data.append(['', '', '', '', f"{total_qty:.2f}", '', ''])
    
    table = Table(data, colWidths=[1.3*inch, 1*inch, 1*inch, 0.8*inch, 1.1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    pdf_bytes = buffer.read()
    
    # Return as downloadable file
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={
            'Content-Disposition': f'attachment; filename=cold_storage_report_{datetime.now().strftime("%Y%m%d")}.pdf'
        }
    )

# ══════════════════════════════════════════════════════════════════════════════
# Audit Trail Interface
# ══════════════════════════════════════════════════════════════════════════════

@api_router.get("/audit-logs")
async def get_audit_logs(
    module: Optional[str] = None,
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(get_current_user)
):
    """Get audit logs with filters"""
    # Build query
    query = {}
    
    if module:
        query['module'] = module
    if action:
        query['action'] = action
    if user_id:
        query['user_id'] = user_id
    if date_from or date_to:
        query['timestamp'] = {}
        if date_from:
            query['timestamp']['$gte'] = date_from
        if date_to:
            query['timestamp']['$lte'] = date_to
    
    # Fetch logs
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    
    # Get total count
    total = await db.audit_logs.count_documents(query)
    
    # Batch enrich logs with user names (one query instead of N)
    user_ids = list({log["user_id"] for log in logs if log.get("user_id")})
    user_map = {}
    if user_ids:
        users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(len(user_ids))
        user_map = {u["id"]: u for u in users}
    for log in logs:
        if log.get("user_id"):
            user = user_map.get(log["user_id"])
            if user:
                log["user_name"] = user.get("name", "Unknown")
                log["user_email"] = user.get("email", "Unknown")
    
    return {
        "logs": logs,
        "total": total,
        "limit": limit,
        "skip": skip
    }

@api_router.get("/audit-logs/modules")
async def get_audit_modules(current_user: User = Depends(get_current_user)):
    """Get list of modules with audit logs"""
    modules = await db.audit_logs.distinct("module")
    return {"modules": sorted(modules)}

@api_router.get("/audit-logs/actions")
async def get_audit_actions(
    module: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get list of actions with audit logs"""
    query = {}
    if module:
        query['module'] = module
    
    actions = await db.audit_logs.distinct("action", query)
    return {"actions": sorted(actions)}

# Universal Attachments & Notes
@api_router.post("/attachments/upload")
async def upload_attachment(
    file: UploadFile = File(...),
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    category: str = Form(...),
    description: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    """Upload a file attachment for any entity"""
    # Validate file size (max 10MB)
    contents = await file.read()
    file_size_kb = len(contents) / 1024
    if file_size_kb > 10240:  # 10MB limit
        raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")
    
    # Generate unique filename
    file_ext = Path(file.filename).suffix
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Create attachment record
    attachment = Attachment(
        entity_type=entity_type,
        entity_id=entity_id,
        file_name=file.filename,
        file_url=f"/uploads/{unique_filename}",
        file_size_kb=round(file_size_kb, 2),
        mime_type=file.content_type or "application/octet-stream",
        category=category,
        description=description,
        uploaded_by=current_user.id
    )
    
    attachment_dict = attachment.model_dump()
    attachment_dict['created_at'] = attachment_dict['created_at'].isoformat()
    
    await db.attachments.insert_one(attachment_dict)
    return {
        "id": attachment.id,
        "file_name": attachment.file_name,
        "file_url": attachment.file_url,
        "file_size_kb": attachment.file_size_kb,
        "category": attachment.category
    }

@api_router.delete("/attachments/{attachment_id}")
async def delete_attachment(attachment_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete an attachment"""
    result = await db.attachments.update_one(
        {"id": attachment_id},
        {"$set": {"is_deleted": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Attachment not found")
    return {"status": "deleted"}

@api_router.post("/attachments", response_model=Attachment)
async def create_attachment(attachment_data: AttachmentCreate, current_user: User = Depends(get_current_user)):
    attachment = Attachment(
        entity_type=attachment_data.entity_type,
        entity_id=attachment_data.entity_id,
        file_name=attachment_data.file_name,
        file_url=f"/uploads/{attachment_data.file_name}",
        file_size_kb=0.0,
        mime_type="application/octet-stream",
        category=attachment_data.category,
        description=attachment_data.description,
        uploaded_by=current_user.id
    )
    
    attachment_dict = attachment.model_dump()
    attachment_dict['created_at'] = attachment_dict['created_at'].isoformat()
    
    await db.attachments.insert_one(attachment_dict)
    return attachment

@api_router.get("/attachments/{entity_type}/{entity_id}", response_model=List[Attachment])
async def get_attachments(entity_type: str, entity_id: str, current_user: User = Depends(get_current_user)):
    attachments = await db.attachments.find(
        {"entity_type": entity_type, "entity_id": entity_id, "is_deleted": False},
        {"_id": 0}
    ).to_list(1000)
    for attachment in attachments:
        if isinstance(attachment.get('created_at'), str):
            attachment['created_at'] = datetime.fromisoformat(attachment['created_at'])
    return attachments

@api_router.post("/notes", response_model=Note)
async def create_note(note_data: NoteCreate, current_user: User = Depends(get_current_user)):
    is_admin_note = current_user.role in [UserRole.admin, UserRole.owner]
    
    note = Note(
        entity_type=note_data.entity_type,
        entity_id=note_data.entity_id,
        note_text=note_data.note_text,
        is_pinned=note_data.is_pinned,
        is_admin_note=is_admin_note,
        authored_by=current_user.id,
        author_name=current_user.name
    )
    
    note_dict = note.model_dump()
    note_dict['created_at'] = note_dict['created_at'].isoformat()
    
    await db.notes.insert_one(note_dict)
    return note

@api_router.get("/notes/{entity_type}/{entity_id}", response_model=List[Note])
async def get_notes(entity_type: str, entity_id: str, current_user: User = Depends(get_current_user)):
    notes = await db.notes.find(
        {"entity_type": entity_type, "entity_id": entity_id, "is_deleted": False},
        {"_id": 0}
    ).sort([("is_pinned", -1), ("created_at", -1)]).to_list(1000)
    for note in notes:
        if isinstance(note.get('created_at'), str):
            note['created_at'] = datetime.fromisoformat(note['created_at'])
    return notes



# Approval Workflow Endpoints
@api_router.get("/admin/pending-approvals", response_model=List[PendingApproval])
async def get_pending_approvals(current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.admin, UserRole.owner, UserRole.production_supervisor]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    approvals = await db.pending_approvals.find(
        {"approval_status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    for approval in approvals:
        if isinstance(approval.get('created_at'), str):
            approval['created_at'] = datetime.fromisoformat(approval['created_at'])
    return approvals

@api_router.post("/admin/approve-action")
async def handle_approval(action_data: ApprovalAction, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.admin, UserRole.owner, UserRole.production_supervisor]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    approval = await db.pending_approvals.find_one({"id": action_data.approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    
    # Update approval status
    new_status = ApprovalStatus.approved if action_data.action == "approve" else ApprovalStatus.rejected
    
    await db.pending_approvals.update_one(
        {"id": action_data.approval_id},
        {"$set": {
            "approval_status": new_status.value,
            "approved_by": current_user.id,
            "approval_notes": action_data.notes
        }}
    )
    
    # If approved, apply changes to the actual entity
    if action_data.action == "approve":
        entity_type = approval['entity_type']
        entity_id = approval['entity_id']
        new_data = approval['new_data']
        
        if entity_type == "procurement_lot":
            await db.procurement_lots.update_one(
                {"id": entity_id},
                {"$set": {
                    **new_data,
                    "is_update_pending_approval": False,
                    "approval_status": "approved",
                    "approved_by": current_user.id,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
    
    await create_audit_log(current_user.id, f"APPROVAL_{action_data.action.upper()}", "admin", {
        "approval_id": action_data.approval_id,
        "entity_type": approval['entity_type'],
        "entity_id": approval['entity_id']
    })
    
    return {"message": f"Successfully {action_data.action}d", "status": new_status.value}


# File Upload Endpoint
@api_router.post("/upload-file")
async def upload_file(
    file: UploadFile = File(...),
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    entity_display: str = Form(...),
    stage: str = Form(...),
    count_per_kg_visible: Optional[str] = Form(None),
    tray_count_visible: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    # Generate unique filename
    file_extension = file.filename.split('.')[-1]
    unique_filename = f"{entity_type}_{entity_id}_{uuid.uuid4()}.{file_extension}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    # Create photo tracker record
    photo_url = f"/uploads/{unique_filename}"
    tray_count = int(tray_count_visible) if tray_count_visible and tray_count_visible.isdigit() else None
    
    photo = PhotoTracker(
        entity_type=entity_type,
        entity_id=entity_id,
        entity_display=entity_display,
        stage=stage,
        photo_url=photo_url,
        count_per_kg_visible=count_per_kg_visible,
        tray_count_visible=tray_count,
        uploaded_by=current_user.id,
        uploader_name=current_user.name
    )
    
    photo_dict = photo.model_dump()
    photo_dict['created_at'] = photo_dict['created_at'].isoformat()
    
    await db.photo_tracker.insert_one(photo_dict)
    
    # Update entity's photos array
    collection_map = {
        "procurement_lot": "procurement_lots",
        "preprocessing_batch": "preprocessing_batches",
        "production_order": "production_orders"
    }
    
    if entity_type in collection_map:
        collection = db[collection_map[entity_type]]
        await collection.update_one(
            {"id": entity_id},
            {"$push": {"photos": photo_url}}
        )
    
    return {"photo_url": photo_url, "photo_id": photo.id, "message": "File uploaded successfully"}


# ============================================================================
# WASTAGE TRACKING ENDPOINTS (v4.0)
# ============================================================================

# Yield Benchmarks CRUD
@api_router.post("/yield-benchmarks", response_model=YieldBenchmark)
async def create_yield_benchmark(
    benchmark_data: YieldBenchmarkCreate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    benchmark = YieldBenchmark(**benchmark_data.model_dump(), set_by=current_user.id)
    benchmark_dict = benchmark.model_dump()
    benchmark_dict['created_at'] = benchmark_dict['created_at'].isoformat()
    if benchmark_dict.get('updated_at'):
        benchmark_dict['updated_at'] = benchmark_dict['updated_at'].isoformat()
    
    await db.yield_benchmarks.insert_one(benchmark_dict)
    return benchmark

@api_router.get("/yield-benchmarks", response_model=List[YieldBenchmark])
async def get_yield_benchmarks(current_user: User = Depends(get_current_user)):
    benchmarks = await db.yield_benchmarks.find({}, {"_id": 0}).to_list(1000)
    return benchmarks

@api_router.get("/yield-benchmarks/{benchmark_id}", response_model=YieldBenchmark)
async def get_yield_benchmark(benchmark_id: str, current_user: User = Depends(get_current_user)):
    benchmark = await db.yield_benchmarks.find_one({"id": benchmark_id}, {"_id": 0})
    if not benchmark:
        raise HTTPException(status_code=404, detail="Benchmark not found")
    return benchmark

@api_router.put("/yield-benchmarks/{benchmark_id}", response_model=YieldBenchmark)
async def update_yield_benchmark(
    benchmark_id: str,
    benchmark_data: YieldBenchmarkCreate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_dict = benchmark_data.model_dump()
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.yield_benchmarks.update_one(
        {"id": benchmark_id},
        {"$set": update_dict}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Benchmark not found")
    
    updated = await db.yield_benchmarks.find_one({"id": benchmark_id}, {"_id": 0})
    return updated

@api_router.delete("/yield-benchmarks/{benchmark_id}")
async def delete_yield_benchmark(benchmark_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.yield_benchmarks.delete_one({"id": benchmark_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Benchmark not found")
    
    return {"message": "Benchmark deleted successfully"}

# Market Rates CRUD
@api_router.post("/market-rates", response_model=MarketRate)
async def create_market_rate(
    rate_data: MarketRateCreate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    rate = MarketRate(**rate_data.model_dump(), set_by=current_user.id)
    rate_dict = rate.model_dump()
    rate_dict['created_at'] = rate_dict['created_at'].isoformat()
    rate_dict['effective_from'] = rate_dict['effective_from'].isoformat()
    if rate_dict.get('effective_to'):
        rate_dict['effective_to'] = rate_dict['effective_to'].isoformat()
    
    await db.market_rates.insert_one(rate_dict)
    return rate

@api_router.get("/market-rates", response_model=List[MarketRate])
async def get_market_rates(current_user: User = Depends(get_current_user)):
    rates = await db.market_rates.find({}, {"_id": 0}).sort("effective_from", -1).to_list(1000)
    return rates

@api_router.get("/market-rates/active", response_model=List[MarketRate])
async def get_active_market_rates(current_user: User = Depends(get_current_user)):
    today = date.today().isoformat()
    rates = await db.market_rates.find({
        "effective_from": {"$lte": today},
        "$or": [
            {"effective_to": None},
            {"effective_to": {"$gte": today}}
        ]
    }, {"_id": 0}).to_list(1000)
    return rates

@api_router.put("/market-rates/{rate_id}", response_model=MarketRate)
async def update_market_rate(
    rate_id: str,
    rate_data: MarketRateCreate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    existing = await db.market_rates.find_one({"id": rate_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Market rate not found")
    
    update_data = rate_data.model_dump()
    update_data['effective_from'] = update_data['effective_from'].isoformat()
    if update_data.get('effective_to'):
        update_data['effective_to'] = update_data['effective_to'].isoformat()
    
    await db.market_rates.update_one(
        {"id": rate_id},
        {"$set": update_data}
    )
    
    updated = await db.market_rates.find_one({"id": rate_id}, {"_id": 0})
    return updated

@api_router.delete("/market-rates/{rate_id}")
async def delete_market_rate(
    rate_id: str,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ['admin', 'owner']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    existing = await db.market_rates.find_one({"id": rate_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Market rate not found")
    
    await db.market_rates.delete_one({"id": rate_id})
    return {"message": "Market rate deleted successfully"}

# Lot Stage Wastage
@api_router.post("/lot-stage-wastage", response_model=LotStageWastage)
async def create_lot_stage_wastage(
    wastage_data: LotStageWastageCreate,
    current_user: User = Depends(get_current_user)
):
    # Calculate derived fields
    wastage_kg = wastage_data.input_weight_kg - wastage_data.output_weight_kg
    yield_pct = (wastage_data.output_weight_kg / wastage_data.input_weight_kg * 100) if wastage_data.input_weight_kg > 0 else 0
    
    # Look up benchmark thresholds
    lot = await db.procurement_lots.find_one({"id": wastage_data.lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    benchmark = await db.yield_benchmarks.find_one({
        "species": lot['species'],
        "process_type": wastage_data.process_type,
        "is_active": True
    }, {"_id": 0})
    
    # Calculate threshold status
    threshold_status = "info"
    min_yield_pct = None
    optimal_yield_pct = None
    
    if benchmark:
        min_yield_pct = benchmark.get('min_yield_pct')
        optimal_yield_pct = benchmark.get('optimal_yield_pct')
        
        if min_yield_pct and optimal_yield_pct:
            if yield_pct >= optimal_yield_pct:
                threshold_status = "green"
            elif yield_pct >= min_yield_pct:
                threshold_status = "amber"
            else:
                threshold_status = "red"
    
    # Look up market rate
    rate_per_kg = None
    if benchmark and benchmark.get('reference_rate_per_kg'):
        rate_per_kg = benchmark['reference_rate_per_kg']
    else:
        rate_per_kg = lot.get('rate_per_kg', 0)
    
    # Calculate revenue loss
    revenue_loss_inr = wastage_kg * rate_per_kg if rate_per_kg else 0
    
    wastage = LotStageWastage(
        **wastage_data.model_dump(),
        wastage_kg=wastage_kg,
        yield_pct=round(yield_pct, 2),
        min_yield_pct=min_yield_pct,
        optimal_yield_pct=optimal_yield_pct,
        threshold_status=threshold_status,
        rate_per_kg_used=rate_per_kg,
        revenue_loss_inr=revenue_loss_inr,
        net_loss_inr=revenue_loss_inr,
        is_alert=(threshold_status == "red"),
        recorded_by=current_user.id
    )
    
    wastage_dict = wastage.model_dump()
    wastage_dict['created_at'] = wastage_dict['created_at'].isoformat()
    if wastage_dict.get('updated_at'):
        wastage_dict['updated_at'] = wastage_dict['updated_at'].isoformat()
    if wastage_dict.get('alert_ack_at'):
        wastage_dict['alert_ack_at'] = wastage_dict['alert_ack_at'].isoformat()
    
    await db.lot_stage_wastage.insert_one(wastage_dict)
    
    # Create alert notification if red
    if wastage.is_alert:
        notification = Notification(
            title=f"🔴 Yield Alert: {wastage.stage_name} below minimum threshold",
            message=f"Lot {lot['lot_number']} {wastage.stage_name}: {yield_pct:.2f}% yield (min: {min_yield_pct}%)\nWastage: {wastage_kg:.2f} kg | Revenue loss: ₹{revenue_loss_inr:.2f}",
            type="alert",
            priority="urgent",
            target_roles=['admin', 'owner', 'production_supervisor'],
            created_by=current_user.id
        )
        notif_dict = notification.model_dump()
        notif_dict['created_at'] = notif_dict['created_at'].isoformat()
        await db.notifications.insert_one(notif_dict)
    
    return wastage

@api_router.get("/lot-stage-wastage/{lot_id}", response_model=List[LotStageWastage])
async def get_lot_wastage(lot_id: str, current_user: User = Depends(get_current_user)):
    wastage_records = await db.lot_stage_wastage.find(
        {"lot_id": lot_id},
        {"_id": 0}
    ).sort("stage_sequence", 1).to_list(1000)
    return wastage_records

# Wastage Dashboard
@api_router.get("/wastage/dashboard-stats", response_model=WastageDashboardStats)
async def get_wastage_dashboard_stats(current_user: User = Depends(get_current_user)):
    today = date.today()
    month_start = date(today.year, today.month, 1)
    
    # Today's wastage
    today_wastage = await db.lot_stage_wastage.aggregate([
        {
            "$match": {
                "created_at": {"$regex": f"^{today.isoformat()}"}
            }
        },
        {
            "$group": {
                "_id": None,
                "total_wastage_kg": {"$sum": "$wastage_kg"},
                "unique_lots": {"$addToSet": "$lot_id"}
            }
        }
    ]).to_list(1)
    
    today_wastage_kg = today_wastage[0]['total_wastage_kg'] if today_wastage else 0
    today_lots_count = len(today_wastage[0]['unique_lots']) if today_wastage else 0
    
    # This month's revenue loss
    month_loss = await db.lot_stage_wastage.aggregate([
        {
            "$match": {
                "created_at": {"$regex": f"^{month_start.isoformat()[:7]}"}
            }
        },
        {
            "$group": {
                "_id": None,
                "total_loss": {"$sum": "$net_loss_inr"},
                "byproduct_revenue": {"$sum": "$byproduct_revenue_inr"},
                "unique_lots": {"$addToSet": "$lot_id"}
            }
        }
    ]).to_list(1)
    
    month_revenue_loss = month_loss[0]['total_loss'] if month_loss else 0
    byproduct_revenue = month_loss[0]['byproduct_revenue'] if month_loss else 0
    month_lots_count = len(month_loss[0]['unique_lots']) if month_loss else 0
    
    # Active red alerts
    active_alerts = await db.lot_stage_wastage.count_documents({
        "is_alert": True,
        "alert_acknowledged": False
    })
    
    return WastageDashboardStats(
        today_wastage_kg=today_wastage_kg,
        today_lots_count=today_lots_count,
        month_revenue_loss_inr=month_revenue_loss,
        month_lots_count=month_lots_count,
        active_red_alerts=active_alerts,
        byproduct_revenue_inr=byproduct_revenue
    )

@api_router.get("/wastage/breach-alerts", response_model=List[WastageBreachAlert])
async def get_wastage_breach_alerts(current_user: User = Depends(get_current_user)):
    alerts = await db.lot_stage_wastage.find({
        "is_alert": True,
        "alert_acknowledged": False
    }, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Batch fetch lots (one query instead of N)
    lot_ids = list({a["lot_id"] for a in alerts if a.get("lot_id")})
    lot_map = {}
    if lot_ids:
        lots = await db.procurement_lots.find({"id": {"$in": lot_ids}}, {"_id": 0, "id": 1, "lot_number": 1, "species": 1}).to_list(len(lot_ids))
        lot_map = {l["id"]: l for l in lots}
    
    result = []
    for alert in alerts:
        lot = lot_map.get(alert["lot_id"]) if alert.get("lot_id") else None
        if lot:
            variance_pct = alert['yield_pct'] - alert['min_yield_pct'] if alert.get('min_yield_pct') else 0
            result.append(WastageBreachAlert(
                id=alert['id'],
                lot_id=alert['lot_id'],
                lot_number=lot['lot_number'],
                stage_name=alert['stage_name'],
                species=lot['species'],
                actual_yield_pct=alert['yield_pct'],
                min_threshold_pct=alert.get('min_yield_pct', 0),
                variance_pct=variance_pct,
                loss_inr=alert['net_loss_inr'],
                created_at=alert['created_at'] if isinstance(alert['created_at'], datetime) else datetime.fromisoformat(alert['created_at'])
            ))
    
    return result

@api_router.post("/wastage/acknowledge/{wastage_id}")
async def acknowledge_wastage_alert(wastage_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ['admin', 'owner', 'production_supervisor']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.lot_stage_wastage.update_one(
        {"id": wastage_id},
        {
            "$set": {
                "alert_acknowledged": True,
                "alert_ack_by": current_user.id,
                "alert_ack_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Wastage record not found")
    
    return {"message": "Alert acknowledged successfully"}

@api_router.get("/wastage/stage-summary")
async def get_stage_wastage_summary(current_user: User = Depends(get_current_user)):
    # Get this month's stage-wise wastage
    month_start = date.today().replace(day=1).isoformat()
    
    summary = await db.lot_stage_wastage.aggregate([
        {
            "$match": {
                "created_at": {"$regex": f"^{month_start[:7]}"}
            }
        },
        {
            "$group": {
                "_id": "$stage_name",
                "total_input_kg": {"$sum": "$input_weight_kg"},
                "total_wastage_kg": {"$sum": "$wastage_kg"},
                "avg_yield_pct": {"$avg": "$yield_pct"},
                "red_count": {
                    "$sum": {"$cond": [{"$eq": ["$threshold_status", "red"]}, 1, 0]}
                },
                "amber_count": {
                    "$sum": {"$cond": [{"$eq": ["$threshold_status", "amber"]}, 1, 0]}
                },
                "green_count": {
                    "$sum": {"$cond": [{"$eq": ["$threshold_status", "green"]}, 1, 0]}
                }
            }
        },
        {"$sort": {"_id": 1}}
    ]).to_list(100)
    
    return summary


@api_router.get("/wastage/lot-waterfall/{lot_id}")
async def get_lot_waterfall(lot_id: str, current_user: User = Depends(get_current_user)):
    """Get complete waterfall view of a lot showing all stages and wastage"""
    # Get lot details
    lot = await db.procurement_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    # Get all wastage records for this lot
    wastage_records = await db.lot_stage_wastage.find(
        {"lot_id": lot_id}, 
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    # Calculate total losses
    total_wastage_kg = sum(w.get("wastage_kg", 0) for w in wastage_records)
    total_revenue_loss = sum(w.get("revenue_loss_inr", 0) for w in wastage_records)
    total_net_loss = sum(w.get("net_loss_inr", 0) for w in wastage_records)
    
    # Build waterfall stages
    stages = []
    current_weight = lot.get("gross_weight_kg", 0)
    
    for record in wastage_records:
        stages.append({
            "stage_name": record.get("stage_name", "Unknown"),
            "process_type": record.get("process_type", "Unknown"),
            "input_weight_kg": record.get("input_weight_kg", 0),
            "output_weight_kg": record.get("output_weight_kg", 0),
            "wastage_kg": record.get("wastage_kg", 0),
            "yield_pct": record.get("yield_pct", 0),
            "threshold_status": record.get("threshold_status", "info"),
            "revenue_loss_inr": record.get("revenue_loss_inr", 0),
            "net_loss_inr": record.get("net_loss_inr", 0),
            "created_at": record.get("created_at", "")
        })
    
    return {
        "lot_id": lot_id,
        "lot_number": lot.get("lot_number", ""),
        "species": lot.get("species", ""),
        "agent_name": lot.get("agent_name", ""),
        "initial_weight_kg": lot.get("gross_weight_kg", 0),
        "final_weight_kg": stages[-1].get("output_weight_kg", 0) if stages else lot.get("net_weight_kg", 0),
        "total_wastage_kg": total_wastage_kg,
        "total_revenue_loss_inr": total_revenue_loss,
        "total_net_loss_inr": total_net_loss,
        "stages": stages
    }




# Photo Tracker Endpoints
@api_router.post("/admin/photos", response_model=PhotoTracker)
async def upload_photo_tracking(photo_data: PhotoUpload, current_user: User = Depends(get_current_user)):
    photo = PhotoTracker(
        entity_type=photo_data.entity_type,
        entity_id=photo_data.entity_id,
        entity_display=photo_data.entity_display,
        stage=photo_data.stage,
        photo_url=photo_data.photo_url,
        count_per_kg_visible=photo_data.count_per_kg_visible,
        tray_count_visible=photo_data.tray_count_visible,
        uploaded_by=current_user.id,
        uploader_name=current_user.name,
        notes=photo_data.notes
    )
    
    photo_dict = photo.model_dump()
    photo_dict['created_at'] = photo_dict['created_at'].isoformat()
    
    await db.photo_tracker.insert_one(photo_dict)
    
    # Also update the entity's photos array
    collection_map = {
        "procurement_lot": "procurement_lots",
        "preprocessing_batch": "preprocessing_batches",
        "production_order": "production_orders"
    }
    
    if photo_data.entity_type in collection_map:
        collection = db[collection_map[photo_data.entity_type]]
        await collection.update_one(
            {"id": photo_data.entity_id},
            {"$push": {"photos": photo_data.photo_url}}
        )
    
    return photo

@api_router.get("/admin/photos", response_model=List[PhotoTracker])
async def get_all_photos(stage: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"stage": stage} if stage else {}
    photos = await db.photo_tracker.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for photo in photos:
        if isinstance(photo.get('created_at'), str):
            photo['created_at'] = datetime.fromisoformat(photo['created_at'])
    return photos

# Live Price Tracking
@api_router.get("/live-prices", response_model=List[LivePriceData])
async def get_live_prices(current_user: User = Depends(get_current_user)):
    # Mock data for now - in production, this would fetch from external API or database
    mock_prices = [
        {
            "id": str(uuid.uuid4()),
            "category": "Vannamei 30/40",
            "price_per_kg": 420.0,
            "location": "Andhra Pradesh",
            "market": "Nellore",
            "date": datetime.now(timezone.utc),
            "source": "Market Data"
        },
        {
            "id": str(uuid.uuid4()),
            "category": "Vannamei 40/60",
            "price_per_kg": 380.0,
            "location": "Andhra Pradesh",
            "market": "Kakinada",
            "date": datetime.now(timezone.utc),
            "source": "Market Data"
        },
        {
            "id": str(uuid.uuid4()),
            "category": "Vannamei 60/80",
            "price_per_kg": 340.0,
            "location": "Andhra Pradesh",
            "market": "Bhimavaram",
            "date": datetime.now(timezone.utc),
            "source": "Market Data"
        },
        {
            "id": str(uuid.uuid4()),
            "category": "Black Tiger 20/30",
            "price_per_kg": 650.0,
            "location": "Andhra Pradesh",
            "market": "Nellore",
            "date": datetime.now(timezone.utc),
            "source": "Market Data"
        },
        {
            "id": str(uuid.uuid4()),
            "category": "Black Tiger 30/40",
            "price_per_kg": 580.0,
            "location": "Andhra Pradesh",
            "market": "Kakinada",
            "date": datetime.now(timezone.utc),
            "source": "Market Data"
        }
    ]
    
    return mock_prices

# Traceability View
@api_router.get("/traceability/{lot_number}", response_model=TraceabilityRecord)
async def get_traceability(lot_number: str, current_user: User = Depends(get_current_user)):
    # Get procurement lot
    lot = await db.procurement_lots.find_one({"lot_number": lot_number}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    # Get preprocessing batches
    batches = await db.preprocessing_batches.find({"procurement_lot_id": lot['id']}, {"_id": 0}).to_list(1000)
    
    # Get production orders
    batch_ids = [b['id'] for b in batches]
    orders = []
    if batch_ids:
        orders = await db.production_orders.find(
            {"preprocessing_batch_ids": {"$in": batch_ids}},
            {"_id": 0}
        ).to_list(1000)
    
    # Get finished goods
    order_ids = [o['id'] for o in orders]
    finished_goods = []
    if order_ids:
        finished_goods = await db.finished_goods.find(
            {"production_order_id": {"$in": order_ids}},
            {"_id": 0}
        ).to_list(1000)
    
    # Get cold storage inventory
    fg_ids = [fg['id'] for fg in finished_goods]
    cold_storage = []
    if fg_ids:
        cold_storage = await db.cold_storage_inventory.find(
            {"fg_id": {"$in": fg_ids}},
            {"_id": 0}
        ).to_list(1000)
    
    # Get shipments
    shipment = None
    if order_ids:
        shipment = await db.shipments.find_one(
            {"sales_order_id": {"$in": order_ids}},
            {"_id": 0}
        )
    
    # Track count per kg changes
    count_changes = [lot.get('count_per_kg', 'N/A')]
    for order in orders:
        if order.get('target_size_count'):
            count_changes.append(order['target_size_count'])
    
    # Calculate total trays
    total_trays = lot.get('no_of_trays', 0)
    for batch in batches:
        total_trays += batch.get('no_of_trays', 0)
    
    # Get photos count
    photos_count = await db.photo_tracker.count_documents({"entity_id": lot['id']})
    
    return TraceabilityRecord(
        lot_number=lot_number,
        procurement_data=lot,
        preprocessing_data=batches,
        production_data=orders,
        finished_goods_data=finished_goods,
        cold_storage_data=cold_storage,
        shipment_data=shipment,
        total_count_per_kg_changes=count_changes,
        total_tray_count=total_trays,
        photos_count=photos_count
    )

# Update procurement lot endpoint to require approval
@api_router.put("/procurement/lots/{lot_id}")
async def update_procurement_lot(
    lot_id: str,
    updates: Dict[str, Any],
    current_user: User = Depends(get_current_user)
):
    lot = await db.procurement_lots.find_one({"id": lot_id}, {"_id": 0})
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    
    # Create pending approval
    pending_approval = PendingApproval(
        entity_type="procurement_lot",
        entity_id=lot_id,
        entity_display=lot['lot_number'],
        change_type="update",
        old_data=lot,
        new_data=updates,
        requested_by=current_user.id,
        requester_name=current_user.name
    )
    
    approval_dict = pending_approval.model_dump()
    approval_dict['created_at'] = approval_dict['created_at'].isoformat()
    
    await db.pending_approvals.insert_one(approval_dict)
    
    # Mark lot as pending approval
    await db.procurement_lots.update_one(
        {"id": lot_id},
        {"$set": {"is_update_pending_approval": True}}
    )
    
    return {
        "message": "Update request submitted for approval",
        "approval_id": pending_approval.id,
        "requires_approval_from": ["admin", "owner", "production_supervisor"]
    }

# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT ANNOUNCEMENTS ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════
@api_router.get("/announcements/active")
async def get_active_announcements(request: Request, current_user: dict = Depends(get_current_user)):
    """Get active announcements for the current tenant"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        return []
    
    now = datetime.now(timezone.utc)
    
    # Fetch announcements from MongoDB
    announcements = await db.active_announcements.find({
        "$and": [
            {
                "$or": [
                    {"target_all": True},
                    {"target_tenant_ids": tenant_id}
                ]
            },
            {"show_from": {"$lte": now.isoformat()}},
            {
                "$or": [
                    {"show_until": None},
                    {"show_until": {"$gte": now.isoformat()}}
                ]
            }
        ]
    }).to_list(length=20)
    
    # Get dismissals for this tenant
    dismissals = await db.announcement_dismissals.find({
        "tenant_id": tenant_id
    }).to_list(length=100)
    
    dismissed_ids = {d["announcement_id"] for d in dismissals}
    
    # Filter out dismissed announcements
    active = []
    for ann in announcements:
        if ann.get("announcement_id") not in dismissed_ids:
            active.append({
                "id": ann.get("announcement_id"),
                "title": ann.get("title"),
                "body": ann.get("body"),
                "type": ann.get("announcement_type", "info"),
                "show_from": ann.get("show_from"),
                "show_until": ann.get("show_until")
            })
    
    return active

@api_router.post("/announcements/{announcement_id}/dismiss")
async def dismiss_announcement(
    announcement_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Dismiss an announcement for the current tenant"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant ID required")
    
    # Record dismissal in MongoDB
    await db.announcement_dismissals.update_one(
        {"tenant_id": tenant_id, "announcement_id": announcement_id},
        {
            "$set": {
                "tenant_id": tenant_id,
                "announcement_id": announcement_id,
                "dismissed_at": datetime.utcnow().isoformat(),
                "dismissed_by": str(current_user.get("_id", ""))
            }
        },
        upsert=True
    )
    
    return {"message": "Announcement dismissed"}


# ══════════════════════════════════════════════════════════════════════════════
# Amendment A5: Party Ledger API Endpoints
# ══════════════════════════════════════════════════════════════════════════════

# ── PARTY MASTER CRUD ─────────────────────────────────────────────────────────

@api_router.post("/parties", response_model=Party)
async def create_party(party: PartyCreate, current_user: User = Depends(get_current_user)):
    """Create a new party"""
    # Check for duplicate party name
    existing = await db.parties.find_one({"party_name": party.party_name})
    if existing:
        raise HTTPException(status_code=400, detail="Party with this name already exists")
    
    party_data = party.model_dump()
    party_id = str(uuid.uuid4())
    party_data["id"] = party_id
    party_data["created_by"] = current_user.id
    party_data["created_at"] = datetime.now(timezone.utc)
    party_data["updated_at"] = datetime.now(timezone.utc)
    party_data["is_active"] = True
    
    await db.parties.insert_one(party_data)
    _invalidate_parties_cache()
    # Auto-create ledger account with opening balance 0 for current FY
    current_fy = get_financial_year(date.today())
    ledger_id = str(uuid.uuid4())
    ledger_account = {
        "id": ledger_id,
        "party_id": party_id,
        "financial_year": current_fy,
        "opening_balance": 0.0,
        "closing_balance": 0.0,
        "total_billed": 0.0,
        "total_tds": 0.0,
        "total_payments": 0.0,
        "is_locked": False,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "created_by": current_user.id
    }
    await db.party_ledger_accounts.insert_one(ledger_account)
    
    # Create opening balance entry
    fy_start, _ = get_fy_date_range(current_fy)
    # Convert date to datetime for MongoDB
    if isinstance(fy_start, date) and not isinstance(fy_start, datetime):
        fy_start = datetime.combine(fy_start, datetime.min.time()).replace(tzinfo=timezone.utc)
    
    opening_entry = {
        "id": str(uuid.uuid4()),
        "ledger_id": ledger_id,
        "party_id": party_id,
        "entry_date": fy_start,
        "entry_type": "opening",
        "entry_order": 0,
        "description": f"Opening Balance FY {current_fy}",
        "balance_after": 0.0,
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.id
    }
    await db.party_ledger_entries.insert_one(opening_entry)
    
    return Party(**party_data)

def _invalidate_parties_cache():
    try:
        tid = tenant_context.get_tenant()
        _response_cache.pop(f"parties_list:{tid}", None)
    except Exception:
        pass


@api_router.get("/parties", response_model=List[Party])
async def list_parties(
    search: Optional[str] = None,
    is_active: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """List all parties with optional search. Full list (no filters) cached 30s."""
    query = {}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"party_alias": {"$regex": search, "$options": "i"}},
            {"short_code": {"$regex": search, "$options": "i"}}
        ]
    if is_active is not None:
        query["is_active"] = is_active
    
    use_cache = not query  # cache only unfiltered list (dropdown / default load)
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"parties_list:{tid}"
    if use_cache:
        cached = _cache_get(cache_key, ttl_sec=30)
        if cached is not None:
            return cached
    
    parties = await db.parties.find(query, {"_id": 0}).sort("party_name", 1).to_list(1000)
    
    # Batch fetch current FY ledger balances (one query instead of N)
    current_fy = get_financial_year(date.today())
    party_ids = [p["id"] for p in parties]
    ledger_map = {}
    if party_ids:
        ledgers = await db.party_ledger_accounts.find(
            {"party_id": {"$in": party_ids}, "financial_year": current_fy},
            {"_id": 0, "party_id": 1, "closing_balance": 1}
        ).to_list(len(party_ids))
        ledger_map = {lg["party_id"]: lg for lg in ledgers}
    for party in parties:
        ledger = ledger_map.get(party["id"])
        party["current_fy_balance"] = ledger["closing_balance"] if ledger else 0.0
    
    result = [Party(**p) for p in parties]
    if use_cache:
        _cache_set(cache_key, result, ttl_sec=30)
    return result


@api_router.get("/parties/insights")
async def get_party_insights(
    fy: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """Party Master insights: FY KG supplied, top parties, and declining parties."""
    if not fy:
        fy = get_financial_year(date.today())
    fy_start, fy_end = get_fy_date_range(fy)
    fy_start_s = fy_start.isoformat()
    fy_end_s = fy_end.isoformat()

    # Rolling comparison windows (recent 90 days vs previous 90 days).
    today = date.today()
    recent_from = (today - timedelta(days=89)).isoformat()
    recent_to = today.isoformat()
    prev_from = (today - timedelta(days=179)).isoformat()
    prev_to = (today - timedelta(days=90)).isoformat()

    parties = await db.parties.find({}, {"_id": 0, "id": 1, "party_name": 1, "short_code": 1, "is_active": 1}).to_list(2000)
    party_map = {p["id"]: p for p in parties if p.get("id")}

    fy_rows = await db.purchase_invoices.aggregate([
        {
            "$match": {
                "party_id": {"$exists": True, "$ne": None, "$ne": ""},
                "invoice_date": {"$gte": fy_start_s, "$lte": fy_end_s},
            }
        },
        {
            "$group": {
                "_id": "$party_id",
                "kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}},
                "invoice_count": {"$sum": 1},
            }
        },
    ]).to_list(3000)

    recent_rows = await db.purchase_invoices.aggregate([
        {
            "$match": {
                "party_id": {"$exists": True, "$ne": None, "$ne": ""},
                "invoice_date": {"$gte": recent_from, "$lte": recent_to},
            }
        },
        {"$group": {"_id": "$party_id", "kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}}}},
    ]).to_list(3000)

    prev_rows = await db.purchase_invoices.aggregate([
        {
            "$match": {
                "party_id": {"$exists": True, "$ne": None, "$ne": ""},
                "invoice_date": {"$gte": prev_from, "$lte": prev_to},
            }
        },
        {"$group": {"_id": "$party_id", "kg": {"$sum": {"$ifNull": ["$total_quantity_kg", 0]}}}},
    ]).to_list(3000)

    fy_map = {r.get("_id"): {"kg": float(r.get("kg") or 0), "invoice_count": int(r.get("invoice_count") or 0)} for r in fy_rows}
    recent_map = {r.get("_id"): float(r.get("kg") or 0) for r in recent_rows}
    prev_map = {r.get("_id"): float(r.get("kg") or 0) for r in prev_rows}

    by_party = []
    for pid, pdata in party_map.items():
        fy_kg = fy_map.get(pid, {}).get("kg", 0.0)
        by_party.append({
            "party_id": pid,
            "party_name": pdata.get("party_name") or pid,
            "short_code": pdata.get("short_code"),
            "is_active": bool(pdata.get("is_active", True)),
            "fy_kg": round(fy_kg, 3),
            "fy_invoice_count": fy_map.get(pid, {}).get("invoice_count", 0),
            "recent_90d_kg": round(recent_map.get(pid, 0.0), 3),
            "prev_90d_kg": round(prev_map.get(pid, 0.0), 3),
        })

    top_performing = sorted(
        [p for p in by_party if p["fy_kg"] > 0],
        key=lambda x: x["fy_kg"],
        reverse=True,
    )[:10]

    declining = []
    for p in by_party:
        prev_kg = p["prev_90d_kg"]
        curr_kg = p["recent_90d_kg"]
        if prev_kg > 0 and curr_kg < prev_kg:
            drop_pct = ((prev_kg - curr_kg) / prev_kg) * 100
            declining.append({
                **p,
                "drop_pct": round(drop_pct, 1),
                "drop_kg": round(prev_kg - curr_kg, 3),
            })
    declining = sorted(declining, key=lambda x: x["drop_pct"], reverse=True)[:10]

    total_fy_kg = round(sum(p["fy_kg"] for p in by_party), 3)
    active_suppliers = sum(1 for p in by_party if p["fy_kg"] > 0)

    return {
        "fy": fy,
        "summary": {
            "total_fy_kg": total_fy_kg,
            "active_suppliers": active_suppliers,
            "declining_count": len(declining),
            "top_performing_count": len(top_performing),
        },
        "top_performing_parties": top_performing,
        "declining_parties": declining,
        "by_party": by_party,
    }

@api_router.get("/parties/{party_id}", response_model=Party)
async def get_party(party_id: str, current_user: User = Depends(get_current_user)):
    """Get a single party by ID"""
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    return Party(**party)

@api_router.put("/parties/{party_id}", response_model=Party)
async def update_party(party_id: str, party: PartyCreate, current_user: User = Depends(get_current_user)):
    """Update a party"""
    existing = await db.parties.find_one({"id": party_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Check for duplicate name (excluding current party)
    duplicate = await db.parties.find_one({"party_name": party.party_name, "id": {"$ne": party_id}})
    if duplicate:
        raise HTTPException(status_code=400, detail="Another party with this name already exists")
    
    update_data = party.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.parties.update_one({"id": party_id}, {"$set": update_data})
    _invalidate_parties_cache()
    updated = await db.parties.find_one({"id": party_id}, {"_id": 0})
    return Party(**updated)

@api_router.delete("/parties/{party_id}")
async def delete_party(party_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete a party (set is_active = False)"""
    existing = await db.parties.find_one({"id": party_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Check if party has ledger entries
    has_entries = await db.party_ledger_entries.find_one({"party_id": party_id})
    if has_entries:
        # Soft delete only
        await db.parties.update_one(
            {"id": party_id}, 
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
        )
        _invalidate_parties_cache()
        return {"message": "Party deactivated (has ledger history)"}
    
    # Hard delete if no entries
    await db.parties.delete_one({"id": party_id})
    _invalidate_parties_cache()
    return {"message": "Party deleted"}


# ── PARTY LEDGER ACCOUNTS ─────────────────────────────────────────────────────

@api_router.get("/party-ledger/available-fys")
async def get_available_fys(current_user: User = Depends(get_current_user)):
    """Get list of financial years that have ledger records"""
    fys = await db.party_ledger_accounts.distinct("financial_year")
    
    # Add current FY if not present
    current_fy = get_financial_year(date.today())
    if current_fy not in fys:
        fys.append(current_fy)
    
    return sorted(fys, reverse=True)

@api_router.get("/party-ledger/parties")
async def list_party_ledgers_with_details(
    fy: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    FIX-1: Get all parties with their ledger account info for current FY
    Returns: all rows from parties table JOIN party_ledger_accounts (for current FY)
    """
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get all active parties
    query = {"is_active": True}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"party_alias": {"$regex": search, "$options": "i"}},
            {"short_code": {"$regex": search, "$options": "i"}}
        ]
    
    parties = await db.parties.find(query, {"_id": 0}).sort("party_name", 1).to_list(1000)
    
    # Batch fetch ledger accounts for this FY (one query instead of N)
    party_ids = [p["id"] for p in parties]
    ledger_map = {}
    if party_ids:
        ledgers = await db.party_ledger_accounts.find(
            {"party_id": {"$in": party_ids}, "financial_year": fy},
            {"_id": 0}
        ).to_list(len(party_ids))
        ledger_map = {lg["party_id"]: lg for lg in ledgers}
    
    default_ledger = {
        "financial_year": fy,
        "opening_balance": 0.0,
        "closing_balance": 0.0,
        "total_billed": 0.0,
        "total_tds": 0.0,
        "total_payments": 0.0
    }
    result = []
    for party in parties:
        ledger = ledger_map.get(party["id"]) or default_ledger
        
        result.append({
            "id": party["id"],
            "party_name": party["party_name"],
            "party_alias": party.get("party_alias"),
            "short_code": party.get("short_code"),
            "mobile": party.get("mobile"),
            "address": party.get("address"),
            "ledger": {
                "financial_year": ledger.get("financial_year", fy),
                "opening_balance": ledger.get("opening_balance", 0.0),
                "closing_balance": ledger.get("closing_balance", 0.0),
                "total_billed": ledger.get("total_billed", 0.0),
                "total_tds": ledger.get("total_tds", 0.0),
                "total_payments": ledger.get("total_payments", 0.0)
            }
        })
    
    return result

@api_router.get("/party-ledger")
async def list_party_ledgers(
    fy: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """List all party ledger accounts for a financial year (legacy endpoint)"""
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get all ledgers for this FY
    query = {"financial_year": fy}
    ledgers = await db.party_ledger_accounts.find(query, {"_id": 0}).to_list(1000)
    if not ledgers:
        return []
    # Batch fetch parties (avoids N+1)
    party_ids = [lg["party_id"] for lg in ledgers]
    parties = await db.parties.find({"id": {"$in": party_ids}}, {"_id": 0, "id": 1, "party_name": 1, "party_alias": 1, "short_code": 1}).to_list(len(party_ids))
    party_map = {p["id"]: p for p in parties}
    result = []
    for ledger in ledgers:
        party = party_map.get(ledger["party_id"])
        if not party:
            continue
        if search and search.lower() not in (party.get("party_name") or "").lower():
            continue
        result.append({
            **ledger,
            "party_name": party.get("party_name"),
            "party_alias": party.get("party_alias"),
            "short_code": party.get("short_code")
        })
    return result

@api_router.get("/party-ledger/parties/{party_id}/ledger")
async def get_party_ledger_detail(
    party_id: str,
    fy: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    FIX-1: Get detailed ledger for a specific party with entries formatted as per spec
    Returns entries with bill_no (=invoice_no), lines array, and running balance
    """
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get party
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Get or create ledger account
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    
    if not ledger:
        # No ledger exists for this FY
        return {
            "party": {
                "id": party["id"],
                "party_name": party["party_name"],
                "party_alias": party.get("party_alias"),
                "short_code": party.get("short_code")
            },
            "ledger_account": None,
            "entries": []
        }
    
    # Get all entries (recompute only on write, not on every read - big perf win)
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger["id"]},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    # Batch-fetch invoice lines for all bill entries (avoids N+1)
    invoice_ids = list({e["invoice_id"] for e in entries if e.get("entry_type") == "bill" and e.get("invoice_id")})
    lines_by_invoice = {}
    if invoice_ids:
        all_lines = await db.purchase_invoice_lines.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"_id": 0, "invoice_id": 1, "line_no": 1, "count_value": 1, "quantity_kg": 1, "rate": 1, "amount": 1}
        ).sort("line_no", 1).to_list(5000)
        for line in all_lines:
            iid = line.get("invoice_id")
            if iid not in lines_by_invoice:
                lines_by_invoice[iid] = []
            lines_by_invoice[iid].append({
                "count_value": line.get("count_value", ""),
                "quantity_kg": line.get("quantity_kg", 0),
                "rate": line.get("rate", 0),
                "amount": line.get("amount", 0)
            })
    
    # Format entries as per spec
    formatted_entries = []
    for entry in entries:
        formatted_entry = {
            "id": entry["id"],
            "entry_date": entry["entry_date"].isoformat() if isinstance(entry["entry_date"], datetime) else entry["entry_date"],
            "entry_type": entry["entry_type"],
            "bill_no": entry.get("invoice_no", ""),
            "entry_order": entry["entry_order"],
            "balance_after": entry["balance_after"]
        }
        if entry.get("entry_type") == "bill" and entry.get("invoice_id"):
            formatted_entry["lines"] = lines_by_invoice.get(entry["invoice_id"], [])
            formatted_entry["total_bill"] = entry.get("bill_subtotal", 0)
            formatted_entry["tds_rate_pct"] = entry.get("tds_rate_pct", 0.1)
            formatted_entry["tds_amount"] = entry.get("tds_amount", 0)
            formatted_entry["tds_after_bill"] = entry.get("tds_after_bill", 0)
        else:
            formatted_entry["lines"] = []
        
        # For payment entries
        if entry.get("entry_type") == "payment":
            formatted_entry["payment_amount"] = entry.get("payment_amount", 0)
            formatted_entry["payment_date"] = entry.get("payment_date", entry["entry_date"])
            formatted_entry["paid_to"] = entry.get("paid_to", "")
            formatted_entry["payment_mode"] = entry.get("payment_mode", "")
            formatted_entry["description"] = entry.get("description", "Payment Received")
        
        # For manual entries
        if entry.get("entry_type") in ("manual_debit", "manual_credit"):
            formatted_entry["description"] = entry.get("description", "")
            if entry.get("entry_type") == "manual_debit":
                formatted_entry["total_bill"] = entry.get("tds_after_bill", 0) or entry.get("bill_subtotal", 0)
                formatted_entry["tds_after_bill"] = entry.get("tds_after_bill", 0) or entry.get("bill_subtotal", 0)
            else:
                formatted_entry["payment_amount"] = entry.get("payment_amount", 0)
        
        formatted_entries.append(formatted_entry)
    
    return {
        "party": {
            "id": party["id"],
            "party_name": party["party_name"],
            "party_alias": party.get("party_alias"),
            "short_code": party.get("short_code")
        },
        "ledger_account": {
            "id": ledger["id"],
            "financial_year": ledger["financial_year"],
            "opening_balance": ledger["opening_balance"],
            "closing_balance": ledger["closing_balance"],
            "total_billed": ledger["total_billed"],
            "total_tds": ledger["total_tds"],
            "total_payments": ledger["total_payments"],
            "is_locked": ledger.get("is_locked", False)
        },
        "entries": formatted_entries
    }

@api_router.get("/party-ledger/{party_id}")
async def get_party_ledger(
    party_id: str,
    fy: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get full ledger detail for a party in a financial year (legacy endpoint)"""
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get party
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Get or create ledger account
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    if not ledger:
        ledger = await _create_party_ledger_for_fy(
            party_id=party_id,
            fy=fy,
            created_by=getattr(current_user, "id", None),
        )
    
    # Get all entries
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger["id"]},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    # Batch-fetch invoice line items for bill entries (avoids N+1)
    invoice_ids = list({e["invoice_id"] for e in entries if e.get("entry_type") == "bill" and e.get("invoice_id")})
    lines_by_invoice = {}
    if invoice_ids:
        all_lines = await db.purchase_invoice_lines.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"_id": 0}
        ).sort("line_no", 1).to_list(5000)
        for line in all_lines:
            iid = line.get("invoice_id")
            if iid not in lines_by_invoice:
                lines_by_invoice[iid] = []
            lines_by_invoice[iid].append(line)
    for entry in entries:
        if entry.get("entry_type") == "bill" and entry.get("invoice_id"):
            entry["line_items"] = lines_by_invoice.get(entry["invoice_id"], [])
    
    tenant_config = await get_tenant_config_dict()
    return {
        "party": party,
        "ledger": ledger,
        "entries": entries,
        "tenant_config": tenant_config,
        "financial_year": fy
    }

@api_router.put("/party-ledger/{party_id}/opening-balance")
async def set_opening_balance(
    party_id: str,
    fy: str,
    opening_balance: float,
    current_user: User = Depends(get_current_user)
):
    """Set opening balance for a ledger (admin override for lock/entries-exist checks)."""
    is_admin = str(getattr(current_user, "role", "")).lower() == "admin"
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy}
    )
    if not ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")
    
    if ledger.get("is_locked") and not is_admin:
        raise HTTPException(status_code=400, detail="Ledger is locked")
    
    # Check for existing entries (excluding opening entry)
    entry_count = await db.party_ledger_entries.count_documents({
        "ledger_id": ledger["id"],
        "entry_type": {"$ne": "opening"}
    })
    if entry_count > 0 and not is_admin:
        raise HTTPException(status_code=400, detail="Cannot change opening balance after entries exist")
    
    # Update opening balance
    await db.party_ledger_accounts.update_one(
        {"id": ledger["id"]},
        {"$set": {
            "opening_balance": opening_balance,
            "closing_balance": opening_balance,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Create or update opening entry
    fy_start, _ = get_fy_date_range(fy)
    await db.party_ledger_entries.update_one(
        {"ledger_id": ledger["id"], "entry_type": "opening"},
        {"$set": {
            "id": str(uuid.uuid4()),
            "ledger_id": ledger["id"],
            "party_id": party_id,
            "entry_date": fy_start.isoformat(),  # Convert date to string
            "entry_type": "opening",
            "description": f"Opening Balance b/f",
            "balance_after": opening_balance,
            "entry_order": 0,
            "created_by": current_user.id,
            "created_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    
    return {"message": "Opening balance set", "opening_balance": opening_balance}


# ── LEDGER ENTRY MANAGEMENT ───────────────────────────────────────────────────

async def recompute_ledger_balances(ledger_id: str):
    """
    Recompute balance_after for all entries in a ledger.
    Must be called after any INSERT/UPDATE/DELETE on entries.
    Single sequential scan - O(n) not O(n²).
    """
    ledger = await db.party_ledger_accounts.find_one({"id": ledger_id}, {"_id": 0})
    if not ledger:
        return
    
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger_id},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    running = ledger.get("opening_balance", 0.0)
    total_billed = 0.0
    total_tds = 0.0
    total_payments = 0.0
    
    for entry in entries:
        entry_type = entry.get("entry_type")
        
        if entry_type == "opening":
            # Opening balance entry just shows the opening, doesn't change running
            pass
        elif entry_type in ("bill", "manual_debit"):
            tds_after = entry.get("tds_after_bill") or entry.get("bill_subtotal", 0)
            running += tds_after
            total_billed += entry.get("bill_subtotal", 0)
            total_tds += entry.get("tds_amount", 0)
        elif entry_type in ("payment", "manual_credit"):
            running -= entry.get("payment_amount", 0)
            total_payments += entry.get("payment_amount", 0)
        
        # Update entry's balance_after
        await db.party_ledger_entries.update_one(
            {"id": entry["id"]},
            {"$set": {"balance_after": round(running, 2)}}
        )
    
    # Update ledger totals
    await db.party_ledger_accounts.update_one(
        {"id": ledger_id},
        {"$set": {
            "closing_balance": round(running, 2),
            "total_billed": round(total_billed, 2),
            "total_tds": round(total_tds, 4),
            "total_payments": round(total_payments, 2),
            "updated_at": datetime.now(timezone.utc)
        }}
    )

async def create_ledger_entry_for_invoice(invoice: dict, current_user_id: str):
    """
    Create a party ledger (bill) entry from invoice details.
    Called when an invoice is approved and optionally when pushed to procurement (if not already created).
    Uses invoice subtotal/grand_total; falls back to summing line items if amounts are missing/zero.
    """
    normalize_invoice_advance_balance(invoice)
    party_id = invoice.get("party_id")
    if not party_id:
        return None  # No party linked, no ledger entry

    invoice_id = invoice.get("id")
    if not invoice_id:
        return None

    # Avoid duplicate bill entry for the same invoice
    existing_bill = await db.party_ledger_entries.find_one(
        {"invoice_id": invoice_id, "entry_type": "bill"}
    )
    if existing_bill:
        # Bill already exists; ensure advance_paid is reflected (backfill if missing)
        advance_paid = float(invoice.get("advance_paid") or 0)
        if advance_paid > 0:
            existing_payment = await db.party_ledger_entries.find_one(
                {"invoice_id": invoice_id, "entry_type": "payment"}
            )
            if not existing_payment:
                # Use same ledger as the bill so payment is counted in this ledger's Total Payments
                ledger_id_for_payment = existing_bill["ledger_id"]
                ledger = await db.party_ledger_accounts.find_one(
                    {"id": ledger_id_for_payment},
                    {"_id": 0}
                )
                if ledger and not ledger.get("is_locked"):
                    invoice_date = invoice.get("invoice_date")
                    if isinstance(invoice_date, str):
                        invoice_date = date.fromisoformat(invoice_date.split("T")[0][:10]) if invoice_date else date.today()
                    invoice_date_str = invoice_date.isoformat() if hasattr(invoice_date, "isoformat") else str(invoice_date)
                    max_order = await db.party_ledger_entries.find_one(
                        {"ledger_id": ledger_id_for_payment},
                        sort=[("entry_order", -1)]
                    )
                    next_order = (max_order.get("entry_order", 0) if max_order else 0) + 1
                    payment_entry = {
                        "id": str(uuid.uuid4()),
                        "ledger_id": ledger_id_for_payment,
                        "party_id": party_id,
                        "entry_date": invoice_date_str,
                        "entry_type": "payment",
                        "payment_amount": round(advance_paid, 2),
                        "payment_date": invoice_date_str,
                        "paid_to": invoice.get("party_name_text") or "",
                        "payment_mode": None,
                        "payment_reference": f"Advance against invoice {invoice.get('invoice_no') or invoice_id}",
                        "invoice_id": invoice_id,
                        "description": f"Advance against invoice {invoice.get('invoice_no') or invoice_id}",
                        "balance_after": 0,
                        "entry_order": next_order,
                        "created_by": current_user_id,
                        "created_at": datetime.now(timezone.utc)
                    }
                    await db.party_ledger_entries.insert_one(payment_entry)
                    await recompute_ledger_balances(ledger_id_for_payment)
        return existing_bill

    invoice_date = invoice.get("invoice_date")
    if isinstance(invoice_date, str):
        invoice_date = date.fromisoformat(invoice_date.split("T")[0] if "T" in invoice_date else invoice_date)
    if not isinstance(invoice_date, date):
        invoice_date = date.today()
    invoice_date_str = invoice_date.isoformat()

    fy = get_financial_year(invoice_date)

    # Resolve bill amount from invoice: prefer subtotal, then grand_total, then sum of line items
    subtotal = float(invoice.get("subtotal") or 0) or float(invoice.get("grand_total") or 0)
    if subtotal <= 0:
        lines = await db.purchase_invoice_lines.find(
            {"invoice_id": invoice_id},
            {"_id": 0, "amount": 1, "quantity_kg": 1, "rate": 1}
        ).to_list(500)
        for line in lines:
            amt = line.get("amount")
            if amt is not None:
                subtotal += float(amt)
            else:
                qty = float(line.get("quantity_kg") or 0)
                rate = float(line.get("rate") or 0)
                subtotal += round(qty * rate, 2)
        subtotal = round(subtotal, 2)

    tds_rate = float(invoice.get("tds_rate_pct") or 0.1)
    tds_amount = float(invoice.get("tds_amount")) if invoice.get("tds_amount") is not None else round(subtotal * tds_rate / 100, 4)
    tds_after_bill = round(subtotal - tds_amount, 2)

    # Get or create ledger account
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy}
    )
    if not ledger:
        ledger = await _create_party_ledger_for_fy(
            party_id=party_id,
            fy=fy,
            created_by=current_user_id,
            opening_entry_date=invoice_date,
        )
    elif ledger.get("is_locked"):
        raise HTTPException(status_code=400, detail=f"Ledger for FY {fy} is locked")

    max_order = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger["id"]},
        sort=[("entry_order", -1)]
    )
    next_order = (max_order.get("entry_order", 0) if max_order else 0) + 1

    entry = {
        "id": str(uuid.uuid4()),
        "ledger_id": ledger["id"],
        "party_id": party_id,
        "entry_date": invoice_date_str,
        "entry_type": "bill",
        "invoice_id": invoice_id,
        "invoice_no": invoice.get("invoice_no"),
        "bill_subtotal": round(subtotal, 2),
        "tds_rate_pct": tds_rate,
        "tds_amount": round(tds_amount, 4),
        "tds_after_bill": tds_after_bill,
        "balance_after": 0,
        "entry_order": next_order,
        "created_by": current_user_id,
        "created_at": datetime.now(timezone.utc)
    }
    await db.party_ledger_entries.insert_one(entry)

    # If invoice has advance paid, add a payment entry so ledger balance reflects it
    advance_paid = float(invoice.get("advance_paid") or 0)
    if advance_paid > 0:
        next_order += 1
        payment_entry = {
            "id": str(uuid.uuid4()),
            "ledger_id": ledger["id"],
            "party_id": party_id,
            "entry_date": invoice_date_str,
            "entry_type": "payment",
            "payment_amount": round(advance_paid, 2),
            "payment_date": invoice_date_str,
            "paid_to": invoice.get("party_name_text") or "",
            "payment_mode": None,
            "payment_reference": f"Advance against invoice {invoice.get('invoice_no') or invoice_id}",
            "invoice_id": invoice_id,
            "description": f"Advance against invoice {invoice.get('invoice_no') or invoice_id}",
            "balance_after": 0,
            "entry_order": next_order,
            "created_by": current_user_id,
            "created_at": datetime.now(timezone.utc)
        }
        await db.party_ledger_entries.insert_one(payment_entry)

    await recompute_ledger_balances(ledger["id"])
    return entry



# ── FIX-3: ADD PAYMENT ENDPOINT ──────────────────────────────────────────────

@api_router.post("/party-ledger/parties/{party_id}/payments")
async def add_party_payment(
    party_id: str,
    payment_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    FIX-3: Add payment entry to party ledger
    Body: { financial_year, payment_date, payment_amount, paid_to, payment_mode, payment_reference, invoice_id? }
    """
    # Get ledger account
    fy = payment_data.get("financial_year") or get_financial_year(date.today())
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    
    if not ledger:
        raise HTTPException(status_code=404, detail=f"No ledger account found for party in FY {fy}")
    
    # Get last entry's balance
    last_entry = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger["id"]},
        {"_id": 0},
        sort=[("entry_order", -1)]
    )
    
    prev_balance = last_entry["balance_after"] if last_entry else ledger["opening_balance"]
    payment_amount = float(payment_data["payment_amount"])
    
    # Create payment entry
    entry_id = str(uuid.uuid4())
    entry_order = (last_entry["entry_order"] + 1) if last_entry else 1
    
    payment_entry = {
        "id": entry_id,
        "ledger_id": ledger["id"],
        "party_id": party_id,
        "entry_date": payment_data["payment_date"],
        "entry_type": "payment",
        "entry_order": entry_order,
        "payment_amount": payment_amount,
        "payment_date": payment_data["payment_date"],
        "paid_to": payment_data.get("paid_to", ""),
        "payment_mode": payment_data.get("payment_mode", ""),
        "payment_reference": payment_data.get("payment_reference", ""),
        "invoice_id": payment_data.get("invoice_id"),
        "balance_after": round(prev_balance - payment_amount, 2),
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.id
    }
    
    await db.party_ledger_entries.insert_one(payment_entry)
    
    # Update ledger account totals
    await db.party_ledger_accounts.update_one(
        {"id": ledger["id"]},
        {
            "$set": {
                "closing_balance": payment_entry["balance_after"],
                "updated_at": datetime.now(timezone.utc)
            },
            "$inc": {"total_payments": payment_amount}
        }
    )
    
    # If invoice_id provided, update invoice payment status
    if payment_data.get("invoice_id"):
        invoice_id = payment_data["invoice_id"]
        invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
        if invoice:
            new_advance_paid = invoice.get("advance_paid", 0) + payment_amount
            balance_due = invoice.get("grand_total", 0) - new_advance_paid
            
            payment_status = "paid" if balance_due <= 0 else ("partial" if new_advance_paid > 0 else "pending")
            
            await db.purchase_invoices.update_one(
                {"id": invoice_id},
                {
                    "$set": {
                        "advance_paid": new_advance_paid,
                        "balance_due": balance_due,
                        "payment_status": payment_status,
                        "updated_at": datetime.now(timezone.utc)
                    }
                }
            )
    
    return {"entry_id": entry_id, "balance_after": payment_entry["balance_after"]}

@api_router.post("/party-ledger/payment")
async def add_payment(payment: PaymentCreate, current_user: User = Depends(get_current_user)):
    """Legacy payment endpoint"""
    # Implementation here for backward compatibility
    pass


# ── FIX-6: MANUAL ENTRY ENDPOINT ─────────────────────────────────────────────

@api_router.post("/party-ledger/parties/{party_id}/manual-entry")
async def add_manual_entry(
    party_id: str,
    entry_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    FIX-6: Add manual debit/credit entry
    Body: { financial_year, entry_date, entry_type('manual_debit'|'manual_credit'), amount, description }
    """
    # Get ledger account
    fy = entry_data.get("financial_year") or get_financial_year(date.today())
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    
    if not ledger:
        raise HTTPException(status_code=404, detail=f"No ledger account found for party in FY {fy}")
    
    # Get last entry's balance
    last_entry = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger["id"]},
        {"_id": 0},
        sort=[("entry_order", -1)]
    )
    
    prev_balance = last_entry["balance_after"] if last_entry else ledger["opening_balance"]
    amount = float(entry_data["amount"])
    entry_type = entry_data["entry_type"]
    
    # Create manual entry
    entry_id = str(uuid.uuid4())
    entry_order = (last_entry["entry_order"] + 1) if last_entry else 1
    
    if entry_type == "manual_debit":
        # Debit adds to balance (increases outstanding)
        new_balance = round(prev_balance + amount, 2)
        manual_entry = {
            "id": entry_id,
            "ledger_id": ledger["id"],
            "party_id": party_id,
            "entry_date": entry_data["entry_date"],
            "entry_type": "manual_debit",
            "entry_order": entry_order,
            "description": entry_data.get("description", "Manual Debit"),
            "bill_subtotal": amount,
            "tds_amount": 0,
            "tds_after_bill": amount,
            "balance_after": new_balance,
            "created_at": datetime.now(timezone.utc),
            "created_by": current_user.id
        }
        
        # Update ledger totals
        await db.party_ledger_entries.insert_one(manual_entry)
        await db.party_ledger_accounts.update_one(
            {"id": ledger["id"]},
            {
                "$set": {
                    "closing_balance": new_balance,
                    "updated_at": datetime.now(timezone.utc)
                },
                "$inc": {"total_billed": amount}
            }
        )
    else:  # manual_credit
        # Credit reduces balance (reduces outstanding)
        new_balance = round(prev_balance - amount, 2)
        manual_entry = {
            "id": entry_id,
            "ledger_id": ledger["id"],
            "party_id": party_id,
            "entry_date": entry_data["entry_date"],
            "entry_type": "manual_credit",
            "entry_order": entry_order,
            "description": entry_data.get("description", "Manual Credit"),
            "payment_amount": amount,
            "balance_after": new_balance,
            "created_at": datetime.now(timezone.utc),
            "created_by": current_user.id
        }
        
        # Update ledger totals
        await db.party_ledger_entries.insert_one(manual_entry)
        await db.party_ledger_accounts.update_one(
            {"id": ledger["id"]},
            {
                "$set": {
                    "closing_balance": new_balance,
                    "updated_at": datetime.now(timezone.utc)
                },
                "$inc": {"total_payments": amount}
            }
        )
    
    # Recompute all balances after this entry
    await recompute_ledger_balances(ledger["id"])
    
    return {"entry_id": entry_id, "balance_after": new_balance}

# ── FIX-7: SET OPENING BALANCE ───────────────────────────────────────────────

@api_router.post("/party-ledger/parties/{party_id}/opening-balance")
async def set_party_opening_balance(
    party_id: str,
    ob_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    FIX-7: Set opening balance for a party in a specific FY
    Body: { financial_year, opening_balance }
    Can only be set when there are no bill/payment entries yet for that FY
    """
    fy = ob_data.get("financial_year") or get_financial_year(date.today())
    opening_balance = float(ob_data["opening_balance"])
    
    # Check if ledger exists
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    
    is_admin = str(getattr(current_user, "role", "")).lower() == "admin"
    if ledger:
        # Check if there are any entries (except opening)
        entries = await db.party_ledger_entries.find(
            {"ledger_id": ledger["id"], "entry_type": {"$ne": "opening"}},
            {"_id": 0}
        ).limit(1).to_list(1)
        
        if entries and not is_admin:
            raise HTTPException(
                status_code=400,
                detail="Cannot set opening balance when entries already exist. Delete all entries first."
            )
        
        # Update existing ledger
        await db.party_ledger_accounts.update_one(
            {"id": ledger["id"]},
            {
                "$set": {
                    "opening_balance": opening_balance,
                    "closing_balance": opening_balance,
                    "updated_at": datetime.now(timezone.utc)
                }
            }
        )
        ledger_id = ledger["id"]
    else:
        # Create new ledger account
        ledger_id = str(uuid.uuid4())
        ledger = {
            "id": ledger_id,
            "party_id": party_id,
            "financial_year": fy,
            "opening_balance": opening_balance,
            "closing_balance": opening_balance,
            "total_billed": 0.0,
            "total_tds": 0.0,
            "total_payments": 0.0,
            "is_locked": False,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
            "created_by": current_user.id
        }
        await db.party_ledger_accounts.insert_one(ledger)
    
    # Create opening entry if it doesn't exist
    opening_entry = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger_id, "entry_type": "opening"},
        {"_id": 0}
    )
    
    if not opening_entry:
        # Get FY start date
        fy_parts = fy.split("-")
        start_year = int("20" + fy_parts[0])
        fy_start_date = f"{start_year}-04-01"
        
        opening_entry = {
            "id": str(uuid.uuid4()),
            "ledger_id": ledger_id,
            "party_id": party_id,
            "entry_date": fy_start_date,
            "entry_type": "opening",
            "entry_order": 0,
            "description": f"Opening Balance FY {fy}",
            "balance_after": opening_balance,
            "created_at": datetime.now(timezone.utc),
            "created_by": current_user.id
        }
        await db.party_ledger_entries.insert_one(opening_entry)
    else:
        # Update existing opening entry
        await db.party_ledger_entries.update_one(
            {"id": opening_entry["id"]},
            {"$set": {"balance_after": opening_balance}}
        )
    
    return {"ledger_id": ledger_id, "opening_balance": opening_balance}

    """Add a payment entry to party ledger"""
    party = await db.parties.find_one({"id": payment.party_id})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    fy = get_financial_year(payment.entry_date)
    
    # Get or create ledger
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": payment.party_id, "financial_year": fy}
    )
    if not ledger:
        ledger = await _create_party_ledger_for_fy(
            party_id=payment.party_id,
            fy=fy,
            created_by=current_user.id,
            opening_entry_date=payment.entry_date,
        )
    elif ledger.get("is_locked"):
        raise HTTPException(status_code=400, detail=f"Ledger for FY {fy} is locked")
    
    # Get next entry order
    max_order = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger["id"]},
        sort=[("entry_order", -1)]
    )
    next_order = (max_order.get("entry_order", 0) if max_order else 0) + 1
    
    # Create payment entry - Convert date objects to ISO strings for MongoDB
    entry_date_str = payment.entry_date.isoformat() if isinstance(payment.entry_date, date) else str(payment.entry_date)
    payment_date_val = payment.payment_date or payment.entry_date
    payment_date_str = payment_date_val.isoformat() if isinstance(payment_date_val, date) else str(payment_date_val)
    
    entry = {
        "id": str(uuid.uuid4()),
        "ledger_id": ledger["id"],
        "party_id": payment.party_id,
        "entry_date": entry_date_str,
        "entry_type": "payment",
        "payment_amount": payment.payment_amount,
        "payment_date": payment_date_str,
        "paid_to": payment.paid_to or party.get("short_code"),
        "payment_mode": payment.payment_mode.value if payment.payment_mode else None,
        "payment_reference": payment.payment_reference,
        "invoice_id": payment.invoice_id,
        "description": payment.notes,
        "balance_after": 0,
        "entry_order": next_order,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc)
    }
    await db.party_ledger_entries.insert_one(entry)
    
    # If linked to invoice, update invoice payment status
    if payment.invoice_id:
        invoice = await db.purchase_invoices.find_one({"id": payment.invoice_id})
        if invoice:
            new_advance = invoice.get("advance_paid", 0) + payment.payment_amount
            grand_total = invoice.get("grand_total", 0)
            new_balance = grand_total - new_advance
            
            if new_balance <= 0:
                new_status = "paid"
            elif new_advance > 0:
                new_status = "partial"
            else:
                new_status = "pending"
            
            await db.purchase_invoices.update_one(
                {"id": payment.invoice_id},
                {"$set": {
                    "advance_paid": new_advance,
                    "balance_due": max(0, new_balance),
                    "payment_status": new_status,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
    
    # Recompute balances
    await recompute_ledger_balances(ledger["id"])
    
    return {"message": "Payment recorded", "entry_id": entry["id"]}

@api_router.post("/party-ledger/manual-entry")
async def add_manual_entry(entry_data: ManualEntryCreate, current_user: User = Depends(get_current_user)):
    """Add a manual debit or credit entry"""
    if entry_data.entry_type not in ("manual_debit", "manual_credit"):
        raise HTTPException(status_code=400, detail="Entry type must be manual_debit or manual_credit")
    
    party = await db.parties.find_one({"id": entry_data.party_id})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    fy = get_financial_year(entry_data.entry_date)
    
    # Get or create ledger
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": entry_data.party_id, "financial_year": fy}
    )
    if not ledger:
        ledger = await _create_party_ledger_for_fy(
            party_id=entry_data.party_id,
            fy=fy,
            created_by=current_user.id,
            opening_entry_date=entry_data.entry_date,
        )
    elif ledger.get("is_locked"):
        raise HTTPException(status_code=400, detail=f"Ledger for FY {fy} is locked")
    
    # Get next entry order
    max_order = await db.party_ledger_entries.find_one(
        {"ledger_id": ledger["id"]},
        sort=[("entry_order", -1)]
    )
    next_order = (max_order.get("entry_order", 0) if max_order else 0) + 1
    
    # Create entry - Convert date objects to ISO strings for MongoDB
    entry_date_str = entry_data.entry_date.isoformat() if isinstance(entry_data.entry_date, date) else str(entry_data.entry_date)
    
    entry = {
        "id": str(uuid.uuid4()),
        "ledger_id": ledger["id"],
        "party_id": entry_data.party_id,
        "entry_date": entry_date_str,
        "entry_type": entry_data.entry_type,
        "description": entry_data.description,
        "balance_after": 0,
        "entry_order": next_order,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc)
    }
    
    if entry_data.entry_type == "manual_debit":
        entry["bill_subtotal"] = entry_data.amount
        entry["tds_after_bill"] = entry_data.amount
    else:
        entry["payment_amount"] = entry_data.amount
    
    await db.party_ledger_entries.insert_one(entry)
    
    # Recompute balances
    await recompute_ledger_balances(ledger["id"])
    
    return {"message": "Manual entry recorded", "entry_id": entry["id"]}

@api_router.delete("/party-ledger/entry/{entry_id}")
async def delete_ledger_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    """Delete a ledger entry (admin can delete any non-opening entry)."""
    entry = await db.party_ledger_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    is_admin = str(getattr(current_user, "role", "")).lower() == "admin"
    if entry.get("entry_type") == "opening":
        raise HTTPException(status_code=400, detail="Opening entry cannot be deleted")
    if entry.get("entry_type") not in ("manual_debit", "manual_credit", "payment") and not is_admin:
        raise HTTPException(status_code=400, detail="Only manual entries and payments can be deleted")
    
    ledger = await db.party_ledger_accounts.find_one({"id": entry["ledger_id"]})
    if ledger and ledger.get("is_locked") and not is_admin:
        raise HTTPException(status_code=400, detail="Ledger is locked")
    
    await db.party_ledger_entries.delete_one({"id": entry_id})
    
    # Recompute balances
    if ledger:
        await recompute_ledger_balances(ledger["id"])
    
    return {"message": "Entry deleted"}


@api_router.put("/party-ledger/entry/{entry_id}")
async def update_ledger_entry(entry_id: str, payload: dict, current_user: User = Depends(get_current_user)):
    """Update ledger entry fields. Admin-only for full ledger book corrections."""
    is_admin = str(getattr(current_user, "role", "")).lower() == "admin"
    if not is_admin:
        raise HTTPException(status_code=403, detail="Only admin can edit ledger entries")

    entry = await db.party_ledger_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.get("entry_type") == "bill":
        raise HTTPException(status_code=400, detail="Bill entries are derived from purchase invoices. Edit invoice instead.")

    ledger = await db.party_ledger_accounts.find_one({"id": entry.get("ledger_id")}, {"_id": 0})
    if not ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")

    update_data = {"updated_at": datetime.now(timezone.utc)}
    entry_type = entry.get("entry_type")
    if payload.get("entry_date") is not None:
        update_data["entry_date"] = payload.get("entry_date")
    if payload.get("description") is not None:
        update_data["description"] = str(payload.get("description") or "")

    if entry_type == "opening":
        if payload.get("amount") is None:
            raise HTTPException(status_code=400, detail="amount is required for opening entry")
        opening_amount = float(payload.get("amount") or 0)
        update_data["balance_after"] = opening_amount
        await db.party_ledger_accounts.update_one(
            {"id": ledger["id"]},
            {"$set": {"opening_balance": opening_amount, "updated_at": datetime.now(timezone.utc)}}
        )
    elif entry_type == "payment":
        if payload.get("amount") is not None:
            update_data["payment_amount"] = float(payload.get("amount") or 0)
        if payload.get("payment_date") is not None:
            update_data["payment_date"] = payload.get("payment_date")
        if payload.get("paid_to") is not None:
            update_data["paid_to"] = str(payload.get("paid_to") or "")
        if payload.get("payment_mode") is not None:
            update_data["payment_mode"] = payload.get("payment_mode")
        if payload.get("payment_reference") is not None:
            update_data["payment_reference"] = str(payload.get("payment_reference") or "")
    elif entry_type == "manual_debit":
        if payload.get("amount") is not None:
            amt = float(payload.get("amount") or 0)
            update_data["bill_subtotal"] = amt
            update_data["tds_after_bill"] = amt
            update_data["tds_amount"] = 0
    elif entry_type == "manual_credit":
        if payload.get("amount") is not None:
            update_data["payment_amount"] = float(payload.get("amount") or 0)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported entry type for edit: {entry_type}")

    await db.party_ledger_entries.update_one({"id": entry_id}, {"$set": update_data})
    await recompute_ledger_balances(ledger["id"])
    return {"status": "success", "message": "Ledger entry updated"}


# ── FY CARRY FORWARD ──────────────────────────────────────────────────────────

@api_router.post("/party-ledger/carry-forward")
async def carry_forward_fy(
    from_fy: str,
    current_user: User = Depends(get_current_user)
):
    """
    Carry forward all party ledgers from one FY to the next.
    Idempotent: existing target ledgers are not overwritten.
    """
    to_fy = get_next_fy(from_fy)
    to_fy_start, _ = get_fy_date_range(to_fy)

    source_ledgers = await db.party_ledger_accounts.find(
        {"financial_year": from_fy},
        {"_id": 0, "party_id": 1, "closing_balance": 1},
    ).to_list(10000)

    created = 0
    skipped_existing = 0
    for src in source_ledgers:
        party_id = src.get("party_id")
        if not party_id:
            continue
        existing_target = await db.party_ledger_accounts.find_one(
            {"party_id": party_id, "financial_year": to_fy},
            {"_id": 0, "id": 1},
        )
        if existing_target:
            skipped_existing += 1
            continue
        await _create_party_ledger_for_fy(
            party_id=party_id,
            fy=to_fy,
            created_by=getattr(current_user, "id", None),
            opening_balance=float(src.get("closing_balance") or 0.0),
            opening_entry_date=to_fy_start,
        )
        created += 1

    return {
        "status": "success",
        "from_fy": from_fy,
        "to_fy": to_fy,
        "source_ledgers": len(source_ledgers),
        "created_ledgers": created,
        "skipped_existing": skipped_existing,
    }


# ── FIX-5: CSV/EXCEL EXPORT ──────────────────────────────────────────────────

@api_router.get("/party-ledger/parties/{party_id}/export")
async def export_party_ledger(
    party_id: str,
    fy: Optional[str] = None,
    format: str = "csv",
    current_user: User = Depends(get_current_user)
):
    """
    FIX-5: Export ledger as CSV or Excel
    Query params: fy (financial year), format ('csv' or 'excel')
    """
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get ledger data using the detail endpoint logic
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    if not ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")
    
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger["id"]},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    # Batch-fetch invoice line items for bill entries (avoids N+1)
    invoice_ids = list({e["invoice_id"] for e in entries if e.get("entry_type") == "bill" and e.get("invoice_id")})
    lines_by_invoice = {}
    if invoice_ids:
        all_lines = await db.purchase_invoice_lines.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"_id": 0}
        ).sort("line_no", 1).to_list(5000)
        for line in all_lines:
            iid = line.get("invoice_id")
            if iid not in lines_by_invoice:
                lines_by_invoice[iid] = []
            lines_by_invoice[iid].append(line)
    for entry in entries:
        if entry.get("entry_type") == "bill" and entry.get("invoice_id"):
            entry["line_items"] = lines_by_invoice.get(entry["invoice_id"], [])
    
    # Helper functions
    def format_date(d):
        if not d: return ""
        if isinstance(d, str):
            d = date.fromisoformat(d.split('T')[0])
        return d.strftime('%d-%b-%y')
    
    def format_curr(amt):
        return f"₹{amt:,.2f}" if amt else ""
    
    # Build CSV data
    rows = []
    rows.append(["DATE", "BILL NO", "COUNT", "QTY", "RATE", "AMOUNT", "TOTAL BILL", "TDS@0.1%", "TDS AFTER BILL", "PAYMENT", "PAYMENT DATE", "BALANCE", "PAID TO"])
    rows.append(["OPENING BALANCE", "", "", "", "", "", "", "", "", "", "", format_curr(ledger["opening_balance"]), ""])
    
    for entry in entries:
        if entry.get('entry_type') == 'opening':
            continue  # Skip opening entry as we already added it
        elif entry.get('entry_type') == 'bill':
            line_items = entry.get('line_items', [])
            if line_items:
                for idx, line in enumerate(line_items):
                    is_last = idx == len(line_items) - 1
                    rows.append([
                        format_date(entry['entry_date']) if idx == 0 else '',
                        entry.get('invoice_no', '') if idx == 0 else '',
                        line.get('count_value', ''),
                        f"{line.get('quantity_kg', 0):.3f}",
                        f"{line.get('rate', 0):.2f}",
                        format_curr(line.get('amount', 0)),
                        format_curr(entry['bill_subtotal']) if is_last else '',
                        format_curr(entry['tds_amount']) if is_last else '',
                        format_curr(entry['tds_after_bill']) if is_last else '',
                        '', '',
                        format_curr(entry['balance_after']) if is_last else '',
                        ''
                    ])
            else:
                rows.append([
                    format_date(entry['entry_date']),
                    entry.get('invoice_no', ''),
                    '', '', '', '',
                    format_curr(entry['bill_subtotal']),
                    format_curr(entry['tds_amount']),
                    format_curr(entry['tds_after_bill']),
                    '', '',
                    format_curr(entry['balance_after']),
                    ''
                ])
        elif entry.get('entry_type') == 'payment':
            rows.append([
                format_date(entry['entry_date']),
                '', '', '', '', '', '', '', '',
                format_curr(entry['payment_amount']),
                format_date(entry.get('payment_date')),
                format_curr(entry['balance_after']),
                entry.get('paid_to', '')
            ])
        elif entry.get('entry_type') in ('manual_debit', 'manual_credit'):
            amt_col = format_curr(entry.get('tds_after_bill') or entry.get('bill_subtotal', 0)) if entry.get('entry_type') == 'manual_debit' else ''
            pay_col = format_curr(entry.get('payment_amount', 0)) if entry.get('entry_type') == 'manual_credit' else ''
            rows.append([
                format_date(entry['entry_date']),
                entry.get('description', ''),
                '', '', '', '', '', '', amt_col, pay_col, '',
                format_curr(entry['balance_after']),
                ''
            ])
    
    # Totals row
    rows.append([
        'TOTAL', '', '', '', '', '',
        format_curr(ledger['total_billed']),
        format_curr(ledger['total_tds']),
        '',
        format_curr(ledger['total_payments']),
        '',
        format_curr(ledger['closing_balance']),
        ''
    ])
    
    if format == "csv":
        # Generate CSV
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        csv_content = output.getvalue()
        
        filename = f"{party['party_name'].replace(' ', '_')}_{fy}.csv"
        
        from starlette.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:  # Excel format
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ledger_{fy}"
        
        # Add rows
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                
                # Header row styling
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
        
        # Save to buffer
        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"{party['party_name'].replace(' ', '_')}_{fy}.xlsx"
        
        from starlette.responses import StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

# ── PDF EXPORT (EXISTING) ─────────────────────────────────────────────────────


# ── PDF & EXCEL EXPORT ────────────────────────────────────────────────────────

@api_router.get("/party-ledger/{party_id}/export-pdf")
async def export_ledger_pdf(
    party_id: str,
    fy: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export party ledger as PDF (Landscape A4)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get ledger data
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    if not ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")
    
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger["id"]},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    # Batch-fetch invoice line items for bill entries (avoids N+1)
    invoice_ids = list({e["invoice_id"] for e in entries if e.get("entry_type") == "bill" and e.get("invoice_id")})
    lines_by_invoice = {}
    if invoice_ids:
        all_lines = await db.purchase_invoice_lines.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"_id": 0}
        ).sort("line_no", 1).to_list(5000)
        for line in all_lines:
            iid = line.get("invoice_id")
            if iid not in lines_by_invoice:
                lines_by_invoice[iid] = []
            lines_by_invoice[iid].append(line)
    for entry in entries:
        if entry.get("entry_type") == "bill" and entry.get("invoice_id"):
            entry["line_items"] = lines_by_invoice.get(entry["invoice_id"], [])
    
    tenant_config = await get_tenant_config_dict()
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(tenant_config.get('company_name', 'KRISH AQUA TRADERS'), title_style))
    elements.append(Paragraph(f"{tenant_config.get('company_address_1', '')}", styles['Normal']))
    elements.append(Spacer(1, 0.1*inch))
    
    # Party info header
    party_name = party['party_name']
    if party.get('party_alias'):
        party_name += f" ({party['party_alias']})"
    
    page_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    header_data = [
        [f"PARTY: {party_name}", f"FY: {fy}", f"Opening Balance: ₹{ledger['opening_balance']:,.2f}"]
    ]
    header_table = Table(
        header_data,
        colWidths=[page_width * 0.56, page_width * 0.14, page_width * 0.30]
    )
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e0f2fe')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.15*inch))
    
    # Ledger table
    wrap_style = ParagraphStyle(
        'LedgerWrap',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.4,
        leading=7.2,
        alignment=TA_LEFT,
        wordWrap='CJK',
        splitLongWords=True,
    )

    def _cell_text(v):
        return str(v or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def wrap_cell(v):
        return Paragraph(_cell_text(v), wrap_style)

    table_data = [
        ['DATE', 'BILL NO', 'COUNT', 'QTY', 'RATE', 'AMOUNT', 'TOTAL', 'TDS@0.1%', 'AFTER TDS', 'PAYMENT', 'PAY DATE', 'BALANCE', 'PAID TO']
    ]
    
    def format_date(d):
        if not d: return ''
        if isinstance(d, str):
            d = date.fromisoformat(d.split('T')[0])
        return d.strftime('%d-%b-%y')
    
    def format_curr(amt):
        return f"₹{amt:,.2f}" if amt else ''
    
    for entry in entries:
        if entry.get('entry_type') == 'opening':
            table_data.append([wrap_cell('OPENING BALANCE'), '', '', '', '', '', '', '', '', '', '', format_curr(entry['balance_after']), ''])
        elif entry.get('entry_type') == 'bill':
            line_items = entry.get('line_items', [])
            if line_items:
                for idx, line in enumerate(line_items):
                    is_last = idx == len(line_items) - 1
                    table_data.append([
                        format_date(entry['entry_date']) if idx == 0 else '',
                        wrap_cell(entry.get('invoice_no', '')) if idx == 0 else '',
                        wrap_cell(line.get('count_value', '')),
                        f"{_pdf_safe_float(line.get('quantity_kg'), 0):.3f}",
                        f"{_pdf_safe_float(line.get('rate'), 0):.2f}",
                        format_curr(line.get('amount', 0)),
                        format_curr(entry['bill_subtotal']) if is_last else '',
                        format_curr(entry['tds_amount']) if is_last else '',
                        format_curr(entry['tds_after_bill']) if is_last else '',
                        '',
                        '',
                        format_curr(entry['balance_after']) if is_last else '',
                        ''
                    ])
            else:
                table_data.append([
                    format_date(entry['entry_date']),
                    wrap_cell(entry.get('invoice_no', '')),
                    '', '', '', '',
                    format_curr(entry['bill_subtotal']),
                    format_curr(entry['tds_amount']),
                    format_curr(entry['tds_after_bill']),
                    '', '',
                    format_curr(entry['balance_after']),
                    ''
                ])
        elif entry.get('entry_type') == 'payment':
            table_data.append([
                format_date(entry['entry_date']),
                '', '', '', '', '', '', '', '',
                format_curr(entry['payment_amount']),
                format_date(entry.get('payment_date')),
                format_curr(entry['balance_after']),
                wrap_cell(entry.get('paid_to', ''))
            ])
        elif entry.get('entry_type') in ('manual_debit', 'manual_credit'):
            amt_col = format_curr(entry.get('tds_after_bill') or entry.get('bill_subtotal', 0)) if entry.get('entry_type') == 'manual_debit' else ''
            pay_col = format_curr(entry.get('payment_amount', 0)) if entry.get('entry_type') == 'manual_credit' else ''
            table_data.append([
                format_date(entry['entry_date']),
                wrap_cell(entry.get('description', '')),
                '', '', '', '', '', '', amt_col, pay_col, '',
                format_curr(entry['balance_after']),
                ''
            ])
    
    # Totals row
    table_data.append([
        'TOTAL', '', '', '', '', '', 
        format_curr(ledger['total_billed']),
        format_curr(ledger['total_tds']),
        '',
        format_curr(ledger['total_payments']),
        '',
        format_curr(ledger['closing_balance']),
        ''
    ])
    
    # Use full printable width and wider text columns to avoid clipping.
    col_fracs = [0.07, 0.13, 0.05, 0.05, 0.05, 0.07, 0.08, 0.07, 0.08, 0.08, 0.07, 0.08, 0.12]
    col_widths = [page_width * f for f in col_fracs]
    
    ledger_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ledger_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 6.4),
        ('ALIGN', (3, 1), (5, -1), 'RIGHT'),  # QTY, RATE, AMOUNT
        ('ALIGN', (6, 1), (-3, -1), 'RIGHT'),  # TOTAL, TDS, AFTER, PAYMENT, BALANCE
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Totals row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ccfbf1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 7),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(ledger_table)
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"Ledger_{party['party_name'].replace(' ', '_')}_{fy}.pdf"
    
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/party-ledger/{party_id}/export-excel")
async def export_ledger_excel(
    party_id: str,
    fy: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export party ledger as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not fy:
        fy = get_financial_year(date.today())
    
    # Get ledger data
    party = await db.parties.find_one({"id": party_id}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    ledger = await db.party_ledger_accounts.find_one(
        {"party_id": party_id, "financial_year": fy},
        {"_id": 0}
    )
    if not ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")
    
    entries = await db.party_ledger_entries.find(
        {"ledger_id": ledger["id"]},
        {"_id": 0}
    ).sort("entry_order", 1).to_list(10000)
    
    # Batch-fetch invoice line items for bill entries (avoids N+1)
    invoice_ids = list({e["invoice_id"] for e in entries if e.get("entry_type") == "bill" and e.get("invoice_id")})
    lines_by_invoice = {}
    if invoice_ids:
        all_lines = await db.purchase_invoice_lines.find(
            {"invoice_id": {"$in": invoice_ids}},
            {"_id": 0}
        ).sort("line_no", 1).to_list(5000)
        for line in all_lines:
            iid = line.get("invoice_id")
            if iid not in lines_by_invoice:
                lines_by_invoice[iid] = []
            lines_by_invoice[iid].append(line)
    for entry in entries:
        if entry.get("entry_type") == "bill" and entry.get("invoice_id"):
            entry["line_items"] = lines_by_invoice.get(entry["invoice_id"], [])
    
    tenant_config = await get_tenant_config_dict()
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ledger_{fy}"
    
    # Styles
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="1e3a8a")
    totals_fill = PatternFill(start_color="ccfbf1", end_color="ccfbf1", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = tenant_config.get('company_name', 'KRISH AQUA TRADERS')
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    
    # Party info
    ws.merge_cells('A2:M2')
    party_name = party['party_name']
    if party.get('party_alias'):
        party_name += f" ({party['party_alias']})"
    ws['A2'].value = f"PARTY: {party_name} | FY: {fy} | Opening Balance: ₹{ledger['opening_balance']:,.2f}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['DATE', 'BILL NO', 'COUNT', 'QTY (KG)', 'RATE', 'AMOUNT', 'TOTAL BILL', 'TDS@0.1%', 'TDS AFTER', 'PAYMENT', 'PAY DATE', 'BALANCE', 'PAID TO']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 5
    
    def format_date_excel(d):
        if not d: return ''
        if isinstance(d, str):
            d = date.fromisoformat(d.split('T')[0])
        return d.strftime('%d-%b-%y')
    
    for entry in entries:
        if entry.get('entry_type') == 'opening':
            ws.cell(row=row_num, column=1, value='OPENING BALANCE')
            ws.cell(row=row_num, column=12, value=entry['balance_after'])
            row_num += 1
        elif entry.get('entry_type') == 'bill':
            line_items = entry.get('line_items', [])
            if line_items:
                for idx, line in enumerate(line_items):
                    is_last = idx == len(line_items) - 1
                    if idx == 0:
                        ws.cell(row=row_num, column=1, value=format_date_excel(entry['entry_date']))
                        ws.cell(row=row_num, column=2, value=entry.get('invoice_no', ''))
                    ws.cell(row=row_num, column=3, value=line.get('count_value', ''))
                    ws.cell(row=row_num, column=4, value=line.get('quantity_kg', 0))
                    ws.cell(row=row_num, column=5, value=line.get('rate', 0))
                    ws.cell(row=row_num, column=6, value=line.get('amount', 0))
                    if is_last:
                        ws.cell(row=row_num, column=7, value=entry['bill_subtotal'])
                        ws.cell(row=row_num, column=8, value=entry['tds_amount'])
                        ws.cell(row=row_num, column=9, value=entry['tds_after_bill'])
                        ws.cell(row=row_num, column=12, value=entry['balance_after'])
                    row_num += 1
            else:
                ws.cell(row=row_num, column=1, value=format_date_excel(entry['entry_date']))
                ws.cell(row=row_num, column=2, value=entry.get('invoice_no', ''))
                ws.cell(row=row_num, column=7, value=entry['bill_subtotal'])
                ws.cell(row=row_num, column=8, value=entry['tds_amount'])
                ws.cell(row=row_num, column=9, value=entry['tds_after_bill'])
                ws.cell(row=row_num, column=12, value=entry['balance_after'])
                row_num += 1
        elif entry.get('entry_type') == 'payment':
            ws.cell(row=row_num, column=1, value=format_date_excel(entry['entry_date']))
            ws.cell(row=row_num, column=10, value=entry['payment_amount'])
            ws.cell(row=row_num, column=11, value=format_date_excel(entry.get('payment_date')))
            ws.cell(row=row_num, column=12, value=entry['balance_after'])
            ws.cell(row=row_num, column=13, value=entry.get('paid_to', ''))
            row_num += 1
        elif entry.get('entry_type') in ('manual_debit', 'manual_credit'):
            ws.cell(row=row_num, column=1, value=format_date_excel(entry['entry_date']))
            ws.cell(row=row_num, column=2, value=entry.get('description', ''))
            if entry.get('entry_type') == 'manual_debit':
                ws.cell(row=row_num, column=9, value=entry.get('tds_after_bill') or entry.get('bill_subtotal', 0))
            else:
                ws.cell(row=row_num, column=10, value=entry.get('payment_amount', 0))
            ws.cell(row=row_num, column=12, value=entry['balance_after'])
            row_num += 1
    
    # Totals row
    ws.cell(row=row_num, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row_num, column=7, value=ledger['total_billed']).font = Font(bold=True)
    ws.cell(row=row_num, column=8, value=ledger['total_tds']).font = Font(bold=True)
    ws.cell(row=row_num, column=10, value=ledger['total_payments']).font = Font(bold=True)
    ws.cell(row=row_num, column=12, value=ledger['closing_balance']).font = Font(bold=True)
    
    for col_num in range(1, 14):
        ws.cell(row=row_num, column=col_num).fill = totals_fill
        ws.cell(row=row_num, column=col_num).border = border
    
    # Format all data cells
    for row in ws.iter_rows(min_row=5, max_row=row_num, min_col=1, max_col=13):
        for cell in row:
            cell.border = border
            if cell.column in [4, 5, 6, 7, 8, 9, 10, 12]:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '₹#,##0.00'
    
    # Auto-fit columns
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Ledger_{party['party_name'].replace(' ', '_')}_{fy}.xlsx"
    
    # Return as streaming response
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/health", tags=["Health"])
async def api_health():
    """Public health check: process up and MongoDB reachable (use this when login fails unexpectedly)."""
    try:
        ping = await db.command("ping")
        if ping.get("ok") != 1:
            raise RuntimeError(f"unexpected ping reply: {ping}")
        return {
            "status": "ok",
            "database": "connected",
            "db_name": os.environ.get("DB_NAME", ""),
        }
    except Exception as e:
        logging.getLogger(__name__).exception("GET /api/health failed")
        raise HTTPException(status_code=503, detail=f"Database unavailable: {str(e)}")


app.include_router(api_router)

# ═══════════════════════════════════════════════════════════════════════════════
# INTERNAL SAAS HOOK ENDPOINTS (A3)
# Receives pushes FROM super admin. Protected by API key.
# ═══════════════════════════════════════════════════════════════════════════════
import hashlib

internal_router = APIRouter(prefix="/internal/saas-hook", tags=["Internal SAAS Hook"])

async def verify_saas_api_key(request: Request):
    """Verify the X-SAAS-API-Key header against stored hash, or allow localhost requests"""
    # Allow requests from localhost (Super Admin API on same server)
    client_host = request.client.host if request.client else None
    if client_host in ["127.0.0.1", "localhost", "::1"]:
        return True
    
    api_key = request.headers.get("X-SAAS-API-Key")
    if not api_key:
        raise HTTPException(status_code=401, detail="Missing API key")
    
    # Get stored hash from tenant_config
    config = await db.tenant_config.find_one({"key": "api_key_hash"})
    if not config:
        raise HTTPException(status_code=401, detail="Client not linked")
    
    # Verify hash
    provided_hash = hashlib.sha256(api_key.encode()).hexdigest()
    if provided_hash != config.get("value"):
        # Log failed attempt
        await db.audit_logs.insert_one({
            "entity_type": "saas_hook",
            "action": "auth_failed",
            "details": {"provided_hash": provided_hash[:16] + "..."},
            "created_at": datetime.utcnow().isoformat()
        })
        raise HTTPException(status_code=401, detail="Invalid API key")
    
    return True

@internal_router.post("/handshake")
async def saas_handshake(request: Request, body: dict):
    """Complete the link handshake with super admin"""
    # Verify API key
    api_key = request.headers.get("X-SAAS-API-Key")
    if not api_key:
        raise HTTPException(status_code=401, detail="Missing API key")
    
    client_id = body.get("client_id")
    tenant_id = body.get("tenant_id")
    api_key_hash = body.get("api_key_hash")
    branding = body.get("branding", {})
    plan = body.get("plan")
    
    # Verify the hash matches what we compute
    computed_hash = hashlib.sha256(api_key.encode()).hexdigest()
    if computed_hash != api_key_hash:
        raise HTTPException(status_code=401, detail="API key hash mismatch")
    
    # Store all config values
    configs = [
        {"key": "client_id", "value": client_id},
        {"key": "tenant_id", "value": tenant_id},
        {"key": "api_key_hash", "value": api_key_hash},
        {"key": "plan", "value": plan},
        {"key": "linked", "value": "true"},
        {"key": "linked_at", "value": datetime.utcnow().isoformat()},
    ]
    
    # Add branding configs
    for key, value in branding.items():
        configs.append({"key": key, "value": str(value) if value else ""})
    
    # Upsert all configs
    for cfg in configs:
        await db.tenant_config.update_one(
            {"key": cfg["key"]},
            {"$set": {"key": cfg["key"], "value": cfg["value"], "synced_at": datetime.utcnow().isoformat()}},
            upsert=True
        )
    
    return {"status": "linked", "erp_version": "5.0"}

@internal_router.post("/features")
async def saas_push_features(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Receive feature flags push from super admin"""
    features = body.get("features", {})
    # Use tenant_id from body if provided (so Super Admin can push without prior link); else from config
    tenant_id = body.get("tenant_id")
    if not tenant_id:
        config = await db.tenant_config.find_one({"key": "tenant_id"})
        tenant_id = config.get("value") if config else "default"
    
    # Clear existing flags and insert new ones
    await db.feature_flags.delete_many({"tenant_id": tenant_id})
    
    count = 0
    for feature_code, is_enabled in features.items():
        await db.feature_flags.insert_one({
            "tenant_id": tenant_id,
            "feature_code": feature_code,
            "is_enabled": bool(is_enabled),
            "synced_at": datetime.utcnow().isoformat()
        })
        count += 1
    
    # Invalidate Redis cache
    try:
        import redis
        redis_client = redis.Redis(host='localhost', port=6379, db=0)
        redis_client.delete(f"flags:{tenant_id}")
    except:
        pass
    
    return {"synced": count}

@internal_router.post("/branding")
async def saas_push_branding(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Receive branding config push from super admin"""
    branding_keys = ["company_name", "primary_color", "logo_url", "favicon_url", 
                     "login_bg_color", "sidebar_label"]
    
    count = 0
    for key in branding_keys:
        if key in body:
            await db.tenant_config.update_one(
                {"key": key},
                {"$set": {"key": key, "value": str(body[key]) if body[key] else "", "synced_at": datetime.utcnow().isoformat()}},
                upsert=True
            )
            count += 1
    
    return {"updated": count}

@internal_router.post("/provision-user")
async def saas_provision_user(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Create or update a user from super admin"""
    user_id = body.get("user_id", str(uuid.uuid4()))
    full_name = body.get("full_name")
    email = body.get("email")
    role = body.get("role", "admin")
    temp_password_hash = body.get("temp_password_hash")
    send_welcome_email = body.get("send_welcome_email", False)
    
    # Get tenant_id
    config = await db.tenant_config.find_one({"key": "tenant_id"})
    tenant_id = config.get("value") if config else "default"
    
    # Check if user exists
    existing = await db.users.find_one({"email": email, "tenant_id": tenant_id})
    
    if existing:
        # Update existing user
        await db.users.update_one(
            {"email": email, "tenant_id": tenant_id},
            {"$set": {
                "name": full_name,
                "role": role,
                "updated_at": datetime.utcnow().isoformat()
            }}
        )
        status = "updated"
    else:
        # Create new user
        user_doc = {
            "id": user_id,
            "email": email,
            "name": full_name,
            "role": role,
            "password_hash": temp_password_hash or pwd_context.hash("TempPass123!"),
            "tenant_id": tenant_id,
            "is_active": True,
            "created_at": datetime.utcnow().isoformat(),
            "provisioned_by_saas": True
        }
        await db.users.insert_one(user_doc)
        status = "created"
    
    # Log the action
    await db.audit_logs.insert_one({
        "entity_type": "saas_hook",
        "action": f"user_{status}",
        "entity_id": user_id,
        "details": {"email": email, "role": role},
        "created_at": datetime.utcnow().isoformat()
    })
    
    return {"user_id": user_id, "status": status}

@internal_router.patch("/update-user")
async def saas_update_user(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Update a user from super admin"""
    user_id = body.get("user_id")
    
    update_fields = {}
    if "role" in body:
        update_fields["role"] = body["role"]
    if "is_active" in body:
        update_fields["is_active"] = body["is_active"]
    
    if update_fields:
        update_fields["updated_at"] = datetime.utcnow().isoformat()
        await db.users.update_one(
            {"id": user_id},
            {"$set": update_fields}
        )
    
    return {"updated": True}

@internal_router.delete("/delete-user")
async def saas_delete_user(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Soft delete a user from super admin"""
    user_id = body.get("user_id")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_active": False, "deactivated_at": datetime.utcnow().isoformat()}}
    )
    
    return {"deactivated": True}

@internal_router.get("/health")
async def saas_health(request: Request, _=Depends(verify_saas_api_key)):
    """Return health status for super admin ping"""
    config = await db.tenant_config.find_one({"key": "tenant_id"})
    tenant_id = config.get("value") if config else "unknown"
    
    linked_config = await db.tenant_config.find_one({"key": "linked_at"})
    linked_at = linked_config.get("value") if linked_config else None
    
    # Count active users today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    active_users = await db.users.count_documents({
        "tenant_id": tenant_id,
        "is_active": True
    })
    
    return {
        "status": "ok",
        "tenant_id": tenant_id,
        "linked_at": linked_at,
        "db_size_mb": 0,  # Would need actual calculation
        "active_users_today": active_users
    }

@internal_router.post("/announcement")
async def saas_push_announcement(request: Request, body: dict, _=Depends(verify_saas_api_key)):
    """Receive announcement push from super admin"""
    await db.active_announcements.update_one(
        {"announcement_id": body.get("id")},
        {"$set": {
            "announcement_id": body.get("id"),
            "title": body.get("title"),
            "body": body.get("body"),
            "announcement_type": body.get("type", "info"),
            "show_from": body.get("show_from"),
            "show_until": body.get("show_until"),
            "synced_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    
    return {"synced": True}

app.include_router(internal_router)

# Public config endpoint for client ERP branding
@app.get("/api/config")
async def get_public_config():
    """Return public config for client ERP branding (no auth required)"""
    config_keys = ["company_name", "sidebar_label", "primary_color", "login_bg_color",
                   "logo_url", "favicon_url", "tenant_id"]
    
    result = {}
    for key in config_keys:
        doc = await db.tenant_config.find_one({"key": key})
        result[key] = doc.get("value") if doc else None
    
    return result

# Mount the integrated super admin router (from super_admin.py - 3-Feature Upgrade)
app.include_router(super_admin_router)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Default client ERP admin (seeded if no users exist)
DEFAULT_CLIENT_ADMIN_EMAIL = "admin@prawnexport.com"
DEFAULT_CLIENT_ADMIN_PASSWORD = "admin123"
DEFAULT_TENANT_ID = "cli_001"

# Simple TTL cache for rarely changing data: key -> (value, expiry_ts)
_response_cache: Dict[str, tuple] = {}
def _cache_get(key: str, ttl_sec: int = 60):
    now = datetime.now(timezone.utc).timestamp()
    if key in _response_cache:
        val, expiry = _response_cache[key]
        if now < expiry:
            return val
        del _response_cache[key]
    return None
def _cache_set(key: str, value: Any, ttl_sec: int = 60):
    _response_cache[key] = (value, datetime.now(timezone.utc).timestamp() + ttl_sec)


def _cache_invalidate(key_prefix: str = ""):
    """Remove cache entries whose key starts with key_prefix (or exact key if no wildcard)."""
    if not key_prefix:
        return
    to_remove = [k for k in _response_cache if k == key_prefix or k.startswith(key_prefix + ":")]
    for k in to_remove:
        _response_cache.pop(k, None)


async def get_tenant_config_dict(use_cache: bool = True) -> Dict[str, Any]:
    """Load full tenant config as key->value dict. Cached 60s per tenant to avoid repeated DB hits (ledger/PDF/export)."""
    try:
        tid = tenant_context.get_tenant()
    except Exception:
        tid = "default"
    cache_key = f"tenant_config_dict:{tid}"
    if use_cache:
        cached = _cache_get(cache_key, ttl_sec=60)
        if cached is not None:
            return cached
    config_docs = await db.tenant_config.find({}, {"_id": 0, "key": 1, "value": 1}).to_list(100)
    out: Dict[str, Any] = {}
    for doc in config_docs:
        k = doc.get("key")
        if k is None or str(k).strip() == "":
            continue
        out[str(k)] = doc.get("value", "")
    _cache_set(cache_key, out, ttl_sec=60)
    return out


async def ensure_erp_indexes():
    """Create indexes on hot-path collections for faster queries."""
    try:
        # Core master data
        await db.agents.create_index("id", unique=True)
        await db.agents.create_index("name")
        await db.agents.create_index("created_at")

        await db.parties.create_index("id", unique=True)
        await db.parties.create_index("party_name")
        await db.parties.create_index("short_code")

        await db.notifications.create_index([("target_roles", 1), ("created_at", -1)])

        # Procurement / processing / production
        await db.procurement_lots.create_index("created_at")
        await db.procurement_lots.create_index("agent_id")
        await db.procurement_lots.create_index("tenant_id")
        await db.procurement_lots.create_index("id", unique=True)

        await db.preprocessing_batches.create_index("created_at")
        await db.preprocessing_batches.create_index([("end_time", 1)])
        await db.preprocessing_batches.create_index("procurement_lot_id")  # traceability

        await db.production_orders.create_index("created_at")
        await db.production_orders.create_index("start_date")  # reports sort
        await db.party_ledger_entries.create_index([("ledger_id", 1), ("entry_order", 1)])
        await db.party_ledger_entries.create_index([("invoice_id", 1), ("entry_type", 1)])
        await db.party_ledger_accounts.create_index([("party_id", 1), ("financial_year", 1)], unique=True)

        await db.purchase_invoice_lines.create_index([("invoice_id", 1), ("line_no", 1)])
        await db.purchase_invoices.create_index("id", unique=True)
        await db.purchase_invoices.create_index([("status", 1), ("invoice_date", 1)])
        await db.purchase_invoices.create_index("invoice_date")  # date range filter + sort
        await db.purchase_invoices.create_index("payment_status")
        await db.purchase_invoices.create_index("party_id")
        # List endpoint: filter by date + status + payment_status, sort by invoice_date
        await db.purchase_invoices.create_index([("invoice_date", 1), ("status", 1), ("payment_status", 1)])

        # Finished goods / QC
        await db.finished_goods.create_index("qc_status")
        await db.finished_goods.create_index("created_at")
        await db.production_orders.create_index("qc_status")

        await db.qc_inspections.create_index("created_at")
        await db.qc_inspections.create_index([("entity_type", 1), ("entity_id", 1)])

        # Cold storage
        await db.cold_storage_chambers.create_index("id", unique=True)
        await db.cold_storage_slots.create_index("id", unique=True)
        await db.cold_storage_slots.create_index("chamber_id")
        await db.cold_storage_inventory.create_index("slot_id")
        await db.cold_storage_inventory.create_index("chamber_id")

        # Sales & shipments
        await db.buyers.create_index("id", unique=True)
        await db.buyers.create_index("name")
        await db.sales_orders.create_index("id", unique=True)
        await db.sales_orders.create_index("created_at")
        await db.shipments.create_index("id", unique=True)
        await db.shipments.create_index("created_at")

        # Wage bills / accounts
        await db.wage_bills.create_index("id", unique=True)
        await db.wage_bills.create_index("created_at")
        await db.wage_bills.create_index("status")

        # Attachments / notes
        await db.attachments.create_index([("entity_type", 1), ("entity_id", 1)])
        await db.attachments.create_index("created_at")
        await db.notes.create_index([("entity_type", 1), ("entity_id", 1)])
        await db.notes.create_index("created_at")

        # Tenant config
        await db.tenant_config.create_index("key")

        # Wastage / traceability
        await db.lot_stage_wastage.create_index("lot_id")
        await db.lot_stage_wastage.create_index([("lot_id", 1), ("stage_sequence", 1)])
        await db.lot_stage_wastage.create_index("created_at")

        # Audit logs (list sorted by timestamp)
        await db.audit_logs.create_index("timestamp")
        await db.audit_logs.create_index([("module", 1), ("timestamp", -1)])

        # Temperature logs (sort by recorded_at)
        await db.temperature_logs.create_index("recorded_at")
        await db.temperature_logs.create_index("chamber_id")

        # Market rates (sort effective_from)
        await db.market_rates.create_index("effective_from")

        # Photo tracker (by entity)
        await db.photo_tracker.create_index([("entity_id", 1), ("created_at", -1)])

        logger.info("ERP collection indexes ensured")
    except Exception as e:
        logger.warning("ERP index creation skipped or partial: %s", e)


@app.on_event("startup")
async def startup_event():
    """Initialize super admin module, create indexes, and seed default client admin if needed"""
    set_super_admin_db(db)
    set_super_admin_feature_service(feature_service)
    await create_super_admin_indexes()
    await ensure_erp_indexes()
    # Ensure default client ERP admin exists (by email) so login always works
    existing_admin = await db.users.find_one({"email": DEFAULT_CLIENT_ADMIN_EMAIL})
    if not existing_admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": DEFAULT_CLIENT_ADMIN_EMAIL,
            "name": "Admin User",
            "role": UserRole.admin.value,
            "phone": None,
            "password": get_password_hash(DEFAULT_CLIENT_ADMIN_PASSWORD),
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "tenant_id": DEFAULT_TENANT_ID,
        })
        logger.info("Seeded default client ERP admin: %s / %s", DEFAULT_CLIENT_ADMIN_EMAIL, DEFAULT_CLIENT_ADMIN_PASSWORD)
    logger.info("Super Admin module initialized with database indexes")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()